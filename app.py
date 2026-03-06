# MAPS MONITORING WIFI PREMIUM V3.1
# Copyrigth (c) 2025 PEYCELL GROUP

import json
import os
import psutil
import time
import threading
import shutil
import requests
import subprocess
import platform
import tempfile
import random
import py_compile
from datetime import datetime, timedelta
import calendar
from flask import Flask, render_template, jsonify, request, send_file, redirect, url_for, make_response, send_from_directory
import sys
import math
import importlib.util
try:
    import openpyxl
    OPENPYXL_READY = True
except ImportError:
    openpyxl = None
    OPENPYXL_READY = False
    print("[INIT] Excel Engine : WARNING (openpyxl not installed. Import/Export limited)")

LAST_BACKUP_DATE = "" # Global variable to track last auto-backup date
LAST_BILLING_NOTIF_DATE = "" # Global variable to track last billing notification date

def get_month_name(month_num, lang='id'):
    """Helper for localized month name"""
    months_id = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", 
                 "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    months_en = ["January", "February", "March", "April", "May", "June",
                 "July", "August", "September", "October", "November", "December"]
    if lang == 'en': return months_en[month_num - 1]
    return months_id[month_num - 1]
import zipfile
import re
from db_manager import DBManager

CURRENT_VERSION = '3.4.1' # Versi Aplikasi Saat Ini
# Link Folder Pusat (Obfuscated Base64)
_B64_FOLDER = "aHR0cHM6Ly9kcml2ZS5nb29nbGUuY29tL2RyaXZlL2ZvbGRlcnMvMW5PQUFULUNiOVFGRjlpb2l6Yzl5OGtJcWxxV1ByS1FX"

def get_pusat_url():
    """Mengembalikan URL pusat yang sudah didekripsi."""
    import base64
    try: return base64.b64decode(_B64_FOLDER).decode('utf-8')
    except: return ""

def extract_gdrive_id(url):
    """Mengekstrak file ID dari berbagai format link Google Drive."""
    pats = [
        r'id=([a-zA-Z0-9_-]+)',
        r'file/d/([a-zA-Z0-9_-]+)',
        r'/d/([a-zA-Z0-9_-]+)',
        r'folders/([a-zA-Z0-9_-]+)',
        r'open\?id=([a-zA-Z0-9_-]+)'
    ]
    for p in pats:
        m = re.search(p, url)
        if m: return m.group(1)
    return None

def download_gdrive_file(file_id, timeout=30):
    """Download file dari GDrive dengan handle virus warning/large file confirmation."""
    URL = "https://docs.google.com/uc?export=download"
    session = requests.Session()
    # Use a real browser User-Agent to avoid getting blocked or different layouts
    session.headers.update({'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'})
    
    # Try first download attempt
    response = session.get(URL, params={'id': file_id}, timeout=timeout)
    
    def get_form_fields(html):
        fields = {}
        # Find all hidden inputs (common in GDrive confirmation pages)
        inputs = re.findall(r'<input type="hidden" name="([^"]+)" value="([^"]+)">', html)
        for name, value in inputs:
            fields[name] = value
        
        # Also check for confirm= in URL/links as backup for some layouts
        if 'confirm' not in fields:
            m = re.search(r'confirm=([a-zA-Z0-9_-]+)', html)
            if m: fields['confirm'] = m.group(1)
            
        return fields

    fields = get_form_fields(response.text)
    if 'confirm' in fields:
        # Use drive.usercontent.google.com for confirmation (standard for GDrive)
        confirm_url = "https://drive.usercontent.google.com/download"
        # Ensure we have essential fields
        fields['id'] = file_id
        response = session.get(confirm_url, params=fields, stream=True, timeout=timeout)
        
    return response
        
    return response

# CRITICAL: Add this script's directory to path FIRST
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

BASE_DIR = SCRIPT_DIR

# Try importing license_utils with detailed error reporting
try:
    # Use standard import to support both .py and .so (Nuitka)
    import license_utils
    
    verify_license = license_utils.verify_license
    get_machine_id = license_utils.get_machine_id
    
except Exception as e:
    err_msg = str(e)
    # Fallback dummy jika license_utils hilang
    def verify_license(k): return False, "Module Missing"
    def get_machine_id(): return f"IMPORT-ERR: {err_msg}"

def get_machine_id_cached():
    """Optimized: Cache machine ID as it never changes during runtime."""
    global _MACHINE_ID_CACHE
    if _MACHINE_ID_CACHE is None:
        _MACHINE_ID_CACHE = get_machine_id()
    return _MACHINE_ID_CACHE

# --- THREAD LOCKS ---
db_lock = threading.Lock()
log_lock = threading.Lock()
billing_lock = threading.Lock()

# DATA PATHS
LOG_FILE = os.path.join(SCRIPT_DIR, 'logs.json')
CONFIG_FILE = os.path.join(SCRIPT_DIR, 'config.json')
SETTINGS_FILE = os.path.join(SCRIPT_DIR, 'settings.json')
settings_lock = threading.Lock()

FINANCE_FILE = os.path.join(SCRIPT_DIR, 'finance.json')
LICENSE_FILE = os.path.join(SCRIPT_DIR, 'license.key')
BLACKLIST_FILE = os.path.join(SCRIPT_DIR, '.blacklist_cache')
BLACKLIST_URL = "https://drive.google.com/file/d/1S701eAR5OWcH_AzKuA2FEo6GWFBP6oFs/view?usp=sharing"
_cached_blacklist = []
_last_blacklist_sync = 0

# --- V3.3.9 PERFORMANCE CACHE ---
_MACHINE_ID_CACHE = None
_LICENSE_CACHE = {"valid": False, "info": {}, "expiry": 0}
_SYSTEM_STATS_CACHE = {"data": {}, "expiry": 0}
_TOPO_CACHE = {"data": None, "expiry": 0}
_SETTINGS_CACHE = {"data": None, "expiry": 0}

def load_cached_blacklist():
    global _cached_blacklist
    if os.path.exists(BLACKLIST_FILE):
        try:
            with open(BLACKLIST_FILE, 'r') as f:
                _cached_blacklist = json.load(f)
        except: _cached_blacklist = []

def perform_blacklist_sync(force=False):
    """Melakukan sinkronisasi blacklist dari URL yang ditanam (Satu kali) dengan throttle 10 menit."""
    global _cached_blacklist, _last_blacklist_sync
    now = time.time()
    # Throttle: Jangan hajar Drive terlalu sering (maks 1x per 10 menit unless force)
    if not force and (now - _last_blacklist_sync) < 600:
        return
    _last_blacklist_sync = now
    
    try:
        f_id = extract_gdrive_id(BLACKLIST_URL)
        if f_id:
            f_url = f"https://docs.google.com/uc?export=download&id={f_id}"
            
            # 2. Download isinya
            f_resp = requests.get(f_url, timeout=15)
            if f_resp.status_code == 200:
                try:
                    data = f_resp.json()
                    if isinstance(data, list):
                        _cached_blacklist = data
                        with open(BLACKLIST_FILE, 'w') as f:
                            json.dump(data, f)
                except:
                    pass
    except:
        pass

def blacklist_sync_loop():
    """Loop background untuk sinkronisasi blacklist setiap 1 jam."""
    while True:
        perform_blacklist_sync()
        time.sleep(3600)

load_cached_blacklist()
# Jalankan loop background sekali saja saat startup
threading.Thread(target=blacklist_sync_loop, daemon=True).start()
BILLING_FILE = os.path.join(SCRIPT_DIR, 'billing.json')
PHOTO_DIR = os.path.join(SCRIPT_DIR, 'static', 'photos')
TEMP_FOLDER = os.path.join(SCRIPT_DIR, 'temp_wa')
if not os.path.exists(TEMP_FOLDER): os.makedirs(TEMP_FOLDER)
temp_folder = TEMP_FOLDER # Compatibility alias
DB_FILE = os.path.join(SCRIPT_DIR, 'topology.db')

# Global DB Manager instance
db = DBManager(DB_FILE)

def _load_settings_raw():
    """ Raw loader to avoid circular dependencies """
    if not os.path.exists(SETTINGS_FILE):
        return dict(DEFAULT_SETTINGS)
    data = _parse_json_file_loose(SETTINGS_FILE, DEFAULT_SETTINGS)
    out = dict(DEFAULT_SETTINGS)
    if isinstance(data, dict):
        out.update(data)
    return out

def _save_settings_raw(data):
    """ Raw saver for internal use """
    with settings_lock:
        _safe_write_json(SETTINGS_FILE, data, critical=True)

def load_billing_config():
    """ 
    Logic Gabungan: Mengambil data billing dari settings.json.
    Jika dari legacy billing.json, bersihkan min_payment_percentage & hapus file lamanya.
    """
    with billing_lock:
        settings = _load_settings_raw()
        billing = settings.get('billing')
        
        if billing:
            # Pastikan bersih dari field lama (permintaan User)
            if isinstance(billing, dict) and 'min_payment_percentage' in billing:
                del billing['min_payment_percentage']
            return billing
            
        if os.path.exists(BILLING_FILE):
            try:
                with open(BILLING_FILE, 'r') as f:
                    legacy = json.load(f)
                
                if 'min_payment_percentage' in legacy: del legacy['min_payment_percentage']
                
                # Merge into settings.json
                settings = _load_settings_raw()
                settings['billing'] = legacy
                _save_settings_raw(settings)
                
                # Cleanup legacy file
                try: os.remove(BILLING_FILE)
                except: pass
                
                return legacy
            except:
                pass
        
        return DEFAULT_SETTINGS['billing']

def save_billing_config(data):
    """ Simpan konfigurasi billing langsung ke settings.json """
    with billing_lock:
        settings = _load_settings_raw()
        if 'billing' not in settings: settings['billing'] = {}
        settings['billing'].update(data)
        # Hapus field minimum pembayaran jika ada (Sesuai permintaan Boss)
        if 'min_payment_percentage' in settings['billing']:
            del settings['billing']['min_payment_percentage']
        return save_settings(settings)

SERVICE_NAME = os.environ.get('NMS_SERVICE', 'monitoring-wifi.service')

if not os.path.exists(PHOTO_DIR): os.makedirs(PHOTO_DIR)

# INITIALIZATION

# --- DISK SYNC HELPER ---
def force_disk_sync(file_obj=None, global_sync=False):
    """ 
    Sinkronisasi disk lintas platform (Windows/Linux).
    Gunakan global_sync=True hanya untuk operasi kritikal agar tidak membebani sistem.
    """
    try:
        if file_obj:
            file_obj.flush()
            os.fsync(file_obj.fileno())
    except: pass
    
    # Global sync for Linux/Unix (Hanya jika benar-benar perlu)
    if global_sync and hasattr(os, 'sync'):
        try: os.sync()
        except: pass

def _safe_write_json(path, data, critical=False):
    """
    Standard tingkat tinggi untuk menyimpan JSON secara atomik & aman.
    Menjamin tidak ada file 0 bytes (corrupt) saat reboot.
    """
    try:
        # 1. Gunakan nama file sementara unik (menghindari tabrakan thread)
        rid = random.randint(1000, 9999)
        tmp = f"{path}.tmp.{rid}"
        
        # 2. Tulis data dengan explicit UTF-8
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4)
            # 3. Paksakan data tertulis ke disk sebelum tutup
            f.flush()
            os.fsync(f.fileno())
        
        # 4. Atomic Swap (rename lebih aman di Linux daripada shutil.move)
        if platform.system().lower() == 'windows':
            if os.path.exists(path): os.remove(path)
            shutil.move(tmp, path)
        else:
            os.rename(tmp, path)
            
        # 5. Backup Otomatis untuk file kritikal
        if critical:
            try:
                shutil.copy2(path, path + '.stable')
            except: pass
            
        return True
    except Exception as e:
        print(f"[DISK ERROR] Gagal menyimpan {path}: {e}")
        # Hapus sisa file sementara jika gagal
        if 'tmp' in locals() and os.path.exists(tmp):
            try: os.remove(tmp)
            except: pass
        return False

# 1. Cek Library Ping (icmplib)
try:
    from icmplib import multiping as turbo_ping
    ICMP_READY = True
except ImportError:
    ICMP_READY = False
    print("[INIT] Ping Engine   : WARNING (icmplib not installed. Slow Mode active)")

# 2. Cek Library Mikrotik (routeros_api)
try:
    import routeros_api
except ImportError:
    routeros_api = None
    print("[INIT] Mikrotik API  : WARNING (routeros_api not installed)")

# 3. Cek Library Kompresi (flask_compress) - UTAMA UNTUK PERFORMA
try:
    from flask_compress import Compress
    COMPRESS_READY = True
except ImportError:
    COMPRESS_READY = False
    print("[INIT] Compression   : WARNING (flask-compress not installed. Standard Speed)")

# --- FLASK APP SETUP ---
app = Flask(__name__)
app.secret_key = 'peycell_nms_final_super_secret'
app.permanent_session_lifetime = timedelta(days=7)

# Aktifkan Kompresi (Jika Library Ada)
if COMPRESS_READY:
    Compress(app)
    app.config['COMPRESS_MIMETYPES'] = ['application/json', 'text/html', 'text/css', 'application/javascript']
    app.config['COMPRESS_LEVEL'] = 6
    app.config['COMPRESS_MIN_SIZE'] = 500

@app.context_processor
def inject_version():
    return dict(app_version=CURRENT_VERSION)

# --- NO CACHE ROUTE FOR LANG DICT (User Request) ---
@app.route('/static/js/lang_dict.js')
def serve_lang_dict_nocache():
    """Force no-cache for language dictionary"""
    # Use absolute path to ensure correct file serving
    js_dir = os.path.join(app.root_path, 'static', 'js')
    response = make_response(send_from_directory(js_dir, 'lang_dict.js'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# LICENSE SYSTEM
def is_licensed():
    """Check if system is licensed. Optimized: signature is cached, but file existence is real-time."""
    global _LICENSE_CACHE
    now = time.time()
    
    # --- REAL-TIME CHECK: Jika file dihapus, harus langsung tidak aktif ---
    if not os.path.exists(LICENSE_FILE): 
        _LICENSE_CACHE = {"valid": False, "info": {}, "expiry": 0}
        return False

    # Return from cache if verification is still fresh (300s = 5 mins)
    if _LICENSE_CACHE['expiry'] > now:
        return _LICENSE_CACHE['valid']

    mid = str(get_machine_id_cached()).strip().upper()
    
    # 1. Cek Blacklist Lokal (Cache)
    bl_clean = [str(x).strip().upper() for x in _cached_blacklist]
    if mid in bl_clean:
        _LICENSE_CACHE = {"valid": False, "info": {}, "expiry": now + 600}
        return False
        
    try:
        with open(LICENSE_FILE, 'r') as f:
            key = f.read().strip()
        valid, info = verify_license(key)
        
        # Cache the result of verification
        _LICENSE_CACHE = {
            "valid": valid,
            "info": info if valid else {},
            "expiry": now + 300 # 5 minutes cache for expensive verification
        }
        return valid
    except:
        return False

@app.before_request
def check_license_gate():
    """Middleware: Unlicensed = Read Only Mode"""
    # 1. Exempt static files
    if request.endpoint == 'static': return None
    
    # 2. Licensi Valid? (Optimasi V3.3.9: is_licensed() gunakan cache)
    if is_licensed(): return None

    # --- JIKA UNLICENSED (LOCKDOWN MODE) ---

    # Allow License Page & Activation
    if request.endpoint in ['license_page', 'activate_license']: return None

    # Block API calls (Return JSON 403)
    if request.path.startswith('/api/'):
        return jsonify({"status": "error", "msg": "LICENSE REQUIRED: System Locked. Please activate license."}), 403

    # Redirect all other pages (Dashboard, Maps, etc) to License Page
    return redirect(url_for('license_page'))

# --- LICENSE ROUTES ---
@app.route('/license')
def license_page():
    # RECOVERY: Jika user terdampar di sini (mungkin karena diblokir), 
    # paksa cek ulang ke cloud tanpa nunggu 10 menit.
    perform_blacklist_sync(force=True)
    
    # Ambil data license saat ini jika ada
    cur_data = None
    active = False
    if is_licensed():
        try:
            with open(LICENSE_FILE, 'r') as f:
                valid, info = verify_license(f.read().strip())
                if valid: 
                    cur_data = info
                    active = True
        except: pass
        
    return render_template('license.html', machine_id=get_machine_id_cached(), license_data=cur_data, active=active)

@app.route('/api/license/activate', methods=['POST'])
def activate_license():
    data = request.json or {}
    key = data.get('key', '').strip()
    
    valid, info = verify_license(key)
    if valid:
        # Simpan key ke file secara aman
        if _safe_replace_file(LICENSE_FILE, key.encode('utf-8')):
            return jsonify({"status": "ok", "msg": "Activated", "info": info})
        else:
            return jsonify({"status": "error", "msg": "Gagal menyimpan file lisensi"})
    else:
        return jsonify({"status": "error", "msg": info})

@app.route('/api/check_license')
def check_license_api():
    """API endpoint to check current license status for UI badges"""
    if not is_licensed():
        return jsonify({"active": False, "owner": "Trial Mode"})
    
    try:
        with open(LICENSE_FILE, 'r') as f:
            valid, info = verify_license(f.read().strip())
            if valid:
                return jsonify({
                    "active": True,
                    "owner": info.get('cli', 'Authorized User'),
                    "type": "LICENSED (PRO)"
                })
    except: pass
    
    return jsonify({"active": False, "owner": "Trial Mode"})
    
@app.route('/api/migrate_v2', methods=['POST'])
def api_migrate_v2():
    """Converts uploaded V2.9 topology.json to SQLite database"""
    if 'file' not in request.files:
        return jsonify({"status": "error", "msg": "No file part"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"status": "error", "msg": "No selected file"}), 400
        
    try:
        # 1. Parse JSON
        raw_data = file.read().decode('utf-8')
        topology_data = json.loads(raw_data)
        
        # 2. Basic Validation
        if not isinstance(topology_data, dict) or 'server' not in topology_data:
            return jsonify({"status": "error", "msg": "Invalid topology.json format"}), 400
            
        # 3. Save to SQLite via DBManager
        success = db.save_full_topology(topology_data)
        
        if success:
            return jsonify({
                "status": "ok", 
                "msg": "Migrasi Berhasil! Data V2.9 telah dipindahkan ke SQLite.",
                "summary": {
                    "clients": len(topology_data.get('clients', [])),
                    "odps": len(topology_data.get('odps', [])),
                    "routers": len(topology_data.get('extra_routers', []))
                }
            })
        else:
            return jsonify({"status": "error", "msg": "Gagal menyimpan ke database SQLite"}), 500
            
    except Exception as e:
        return jsonify({"status": "error", "msg": f"Error: {str(e)}"}), 500



# ==============================================================================
#  KONFIGURASI OTOMATIS (AUTO-GENERATE)
# ==============================================================================
DEFAULT_CONFIG = {
    "admin_password": "admin",
    "viewer_password": "tamu",
    "telegram_bot_token": "",
    "telegram_chat_id": "",
    "service_name": "monitoring-wifi.service",
    "app_port": 5002
}

DEFAULT_SETTINGS = {
    "web_title": "MAPS MONITORING WIFI",
    "refresh_rate": 10,
    "wa_template": "Halo Kak *{name}*, tagihan internet bulan ini sudah terbit. Mohon pembayarannya ya. Terima kasih.",
    "wa_template_auto": "Yth. {name}, tagihan internet Anda sudah memasuki masa tenggang. Layanan akan dinonaktifkan otomatis dalam 3 hari jika belum ada pembayaran. Terima kasih.",
    "wa_template_payment": "Terima kasih, pembayaran wifi a.n *{name}* sebesar Rp {amount} pada {date} telah diterima. Layanan Anda tetap aktif hingga {expired}. Terima kasih.",
    "wa_template_isolir": "Yth. *{name}*, layanan internet Anda telah diisolir sementara karena keterlambatan pembayaran sebesar Rp {price}. Silakan lakukan pembayaran agar layanan kembali normal.",
    "wa_template_reactivate": "Halo *{name}*, pembayaran telah diterima dan layanan internet Anda telah diaktifkan kembali. Selamat berinternet!",
    "map_animation": True,
    "inventory": [],
    "billing": {
        "auto_isolir_enabled": True,
        "default_billing_day": 5,
        "grace_period_days": 3,
        "isolir_profile": "ISOLIR",
        "send_wa_notification": False,
        "billing_check_interval_hours": 24
    },
    "automation": {
        "backup": {
            "enabled": True,
            "schedule_time": "02:00",
            "keep_days": 7,
            "include_files": ["topology.db", "settings.json", "finance.json", "config.json", "license.key", "app.py", "db_manager.py", "license_utils.py", "wa-bridge.js", "package.json", "keygen.py", "CARA_PAKAI_KEYGEN.md", "DISTRIBUSI_KLIEN.md", "templates", "static"]
        },
        "telegram": {
            "enabled": False,
            "bot_token": "",
            "chat_id": "",
            "notifications": {
                "offline": True,
                "online": True,
                "daily_report": False,
                "backup_report": True,
                "startup_report": True
            }
        }
    },
    "print_header": "TERIMA KASIH TELAH MENGGUNAKAN LAYANAN KAMI",
    "print_footer": "Simpan struk ini sebagai bukti pembayaran sah.",
    "print_store_name": "NMS PREMIUM WIFI",
    "print_paper": "58mm",
    "print_auto": False,
    "print_show_logo": True,
    "print_method": "browser",
    "print_template": """<div style="text-align:center; font-family:monospace;">
    <p style="margin:0; font-weight:bold; font-size:16px;">STRUK PEMBAYARAN</p>
    <p style="margin:5px 0;">{header}</p>
    <hr style="border-top:1px dashed #000;">
    <table style="width:100%; font-size:14px;">
        <tr><td style="text-align:left;">Tgl</td><td style="text-align:right;">{date}</td></tr>
        <tr><td style="text-align:left;">ID</td><td style="text-align:right;">{id}</td></tr>
        <tr><td style="text-align:left;">Nama</td><td style="text-align:right;">{name}</td></tr>
        <tr><td style="text-align:left;">Paket</td><td style="text-align:right;">{packet}</td></tr>
        <tr><td style="text-align:left; font-weight:bold;">TOTAL</td><td style="text-align:right; font-weight:bold;">Rp {amount}</td></tr>
    </table>
    <hr style="border-top:1px dashed #000;">
    <p style="margin:5px 0;">Berlaku S/D: <b>{expired}</b></p>
    <hr style="border-top:1px dashed #000;">
    <p style="margin:5px 0; font-size:12px;">{footer}</p>
</div>""",
    "print_template_text": """{store_name}
--------------------------------
Tanggal : {date}
Nama    : {name}
Paket   : {packet}
Harga   : {amount}
--------------------------------
Terima Kasih

{footer}"""
}

def load_config():
    if not os.path.exists(CONFIG_FILE):
        # Coba restore dari .stable jika ada
        stable = CONFIG_FILE + '.stable'
        if os.path.exists(stable):
            try:
                with open(stable, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    _safe_write_json(CONFIG_FILE, data, critical=True)
                    return data
            except: pass
            
        print(f"[SYSTEM] Membuat file konfigurasi baru: {CONFIG_FILE}")
        _safe_write_json(CONFIG_FILE, DEFAULT_CONFIG, critical=True)
        return DEFAULT_CONFIG
    
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"[ERROR] Config rusak, mencoba recovery .stable: {e}")
        stable = CONFIG_FILE + '.stable'
        if os.path.exists(stable):
            try:
                with open(stable, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    _safe_write_json(CONFIG_FILE, data, critical=True)
                    return data
            except: pass
        return DEFAULT_CONFIG

def load_settings():
    """Optimized V3.3.9: Cache settings for 30s to reduce disk I/O."""
    global _SETTINGS_CACHE
    now = time.time()
    if _SETTINGS_CACHE['data'] is not None and _SETTINGS_CACHE['expiry'] > now:
        return _SETTINGS_CACHE['data']
    
    settings_data = _load_settings_internal()
    _SETTINGS_CACHE = {"data": settings_data, "expiry": now + 30}
    return settings_data

def _load_settings_internal():
    if not os.path.exists(SETTINGS_FILE):
        # Coba restore dari .stable jika ada
        stable = SETTINGS_FILE + '.stable'
        if os.path.exists(stable):
            try:
                with open(stable, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    _safe_write_json(SETTINGS_FILE, data, critical=True)
                    return load_settings() # Reload with merged logic
            except: pass
            
        _safe_write_json(SETTINGS_FILE, DEFAULT_SETTINGS, critical=True)
        return DEFAULT_SETTINGS
    data = _parse_json_file_loose(SETTINGS_FILE, DEFAULT_SETTINGS)
    out = dict(DEFAULT_SETTINGS)
    if isinstance(data, dict):
        out.update(data)
    try:
        out['refresh_rate'] = int(out.get('refresh_rate', 10))
    except Exception:
        out['refresh_rate'] = 10
    out['map_animation'] = bool(out.get('map_animation', True))
    
    # Merge Billing Config
    out['billing'] = load_billing_config()
    
    # BUGFIX: Inject Telegram from config.json to ensure UI display is correct
    # (Single Source of Truth: config.json for Bot & Port)
    cfg_tmp = load_config()
    if 'automation' not in out: out['automation'] = {}
    if 'telegram' not in out['automation']: out['automation']['telegram'] = {}
    
    out['automation']['telegram']['bot_token'] = cfg_tmp.get('telegram_bot_token', '')
    out['automation']['telegram']['chat_id'] = cfg_tmp.get('telegram_chat_id', '')
    
    return out

def save_settings(new_settings):
    try:
        current = _load_settings_raw()
        merged = dict(current)
        if isinstance(new_settings, dict):
            # Consolidate ALL settings including billing
            merged.update(new_settings)

        try:
            merged['refresh_rate'] = int(merged.get('refresh_rate', 10))
        except Exception:
            merged['refresh_rate'] = 10
        merged['map_animation'] = bool(merged.get('map_animation', True))
        
        # Save ALL settings (Billing is now part of settings.json)
        core_settings = dict(merged)
        
        # BUGFIX: Exclude Infra Config from settings.json (Already in config.json)
        # Port & Service Name
        if 'app_port' in core_settings: del core_settings['app_port']
        if 'service_name' in core_settings: del core_settings['service_name']
        
        # Telegram Bot Token & Chat ID
        if 'automation' in core_settings and 'telegram' in core_settings['automation']:
            # Capture for config.json if they were updated in this payload
            changed_cfg = False
            cur_cfg = load_config()
            
            tg = core_settings['automation']['telegram']
            if 'bot_token' in tg:
                cur_cfg['telegram_bot_token'] = tg['bot_token']
                changed_cfg = True
            if 'chat_id' in tg:
                cur_cfg['telegram_chat_id'] = tg['chat_id']
                changed_cfg = True
                
            if changed_cfg:
                _safe_write_json(CONFIG_FILE, cur_cfg, critical=True)
                reload_config_globals()

        # Invalidate Settings Cache V3.3.9
        global _SETTINGS_CACHE
        with settings_lock:
            _SETTINGS_CACHE = {"data": None, "expiry": 0}
            _safe_write_json(SETTINGS_FILE, core_settings, critical=True)
        
        # Reset billing tracking to allow immediate re-test of automation time
        global LAST_BILLING_NOTIF_DATE
        LAST_BILLING_NOTIF_DATE = None
        reload_config_globals()
        
        return merged
    except Exception:
        return load_settings()

def save_db(incoming, preserve_live=True):
    """
    Saves the full topology database while optionally preserving live status
    to prevent race conditions with monitoring threads.
    Optimized V3.3.9: Invalidates _TOPO_CACHE on write.
    """
    global _TOPO_CACHE
    _TOPO_CACHE = {"data": None, "expiry": 0}
    try:
        if preserve_live:
            # Load current live data from DB (Source of Truth: Background Threads & Manual Updates)
            current_db = db.load_full_topology()
            # V3.3.9 FIX: Preserve ALL critical fields that might change during long loops
            live_fields = ['status', 'ping_ms', 'last_seen_ts', 'last_ping', 'paid_until']
            current_clients = {c['id']: c for c in current_db.get('clients', []) if 'id' in c}
            
            if 'clients' in incoming:
                for c in incoming['clients']:
                    cid = c.get('id')
                    if cid and cid in current_clients:
                        # 1. Base fields
                        for fld in live_fields:
                            if fld in current_clients[cid]:
                                c[fld] = current_clients[cid][fld]
                        
                        # 2. Nested billing fields
                        if 'billing' in current_clients[cid] and isinstance(current_clients[cid]['billing'], dict):
                            if 'billing' not in c: c['billing'] = {}
                            bill_lives = ['payment_status', 'original_profile', 'isolir_wa_sent']
                            for bf in bill_lives:
                                if bf in current_clients[cid]['billing']:
                                    c['billing'][bf] = current_clients[cid]['billing'][bf]

        # Save to SQLite
        return db.save_full_topology(incoming)
        
    except Exception as e:
        print(f"[DB ERROR] save_db critical failure: {e}")
        return False

def load_db(force_refresh=False):
    # Optimized Cache Logic
    global _TOPO_CACHE
    now = time.time()
    
    if not force_refresh and _TOPO_CACHE['data'] is not None and _TOPO_CACHE['expiry'] > now:
        return _TOPO_CACHE['data']

    try:
        data = db.load_full_topology()
        _TOPO_CACHE = {"data": data, "expiry": now + 5}
        return data
    except Exception as e:
        print(f"[DB ERROR] load_db failed: {e}")
        return {"server": {}, "odps": [], "clients": [], "extra_routers": []}

def load_finance():
    if not os.path.exists(FINANCE_FILE):
        return {"transactions": []}
    try:
        with open(FINANCE_FILE, 'r') as f:
            return json.load(f)
    except:
        return {"transactions": []}

def save_finance(data):
    return _safe_write_json(FINANCE_FILE, data)

def _parse_json_file_loose(path, default_obj):
    try:
        if not os.path.exists(path):
            return default_obj
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            raw = f.read().strip()
        try:
            return json.loads(raw)
        except Exception:
            start = raw.find('{')
            end = raw.rfind('}')
            if start != -1 and end != -1 and end > start:
                return json.loads(raw[start:end+1])
            return default_obj
    except Exception:
        return default_obj

# Load Config
cfg = load_config()
ADMIN_PASSWORD = cfg.get('admin_password', 'admin')
PASSWORD_VIEWER = cfg.get('viewer_password', 'tamu')
TELEGRAM_BOT_TOKEN = cfg.get('telegram_bot_token', '')
TELEGRAM_CHAT_ID = cfg.get('telegram_chat_id', '')
SERVICE_NAME = cfg.get('service_name', os.environ.get('NMS_SERVICE', 'monitoring-wifi.service'))
APP_PORT = int(cfg.get('app_port', 5002))

print(f"[SYSTEM] MAPS MONITORING V{CURRENT_VERSION} STARTED")

_CONFIG_CACHE = {"data": None, "expiry": 0}

def reload_config_globals():
    global cfg, ADMIN_PASSWORD, PASSWORD_VIEWER, TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID, SERVICE_NAME, APP_PORT, _CONFIG_CACHE
    
    now = time.time()
    if _CONFIG_CACHE['data'] is not None and _CONFIG_CACHE['expiry'] > now:
        cfg = _CONFIG_CACHE['data']
    else:
        new_cfg = load_config()
        # Prevent wiping out passwords if disk read failed momentarily (Windows Atomic Swap)
        if new_cfg.get('admin_password') == 'admin' and cfg and cfg.get('admin_password') != 'admin':
            # Trust the old one for a bit, don't update cache.
            pass
        else:
            cfg = new_cfg
            _CONFIG_CACHE = {"data": cfg, "expiry": now + 5} # 5 second cache to bridge atomic renames
            
    ADMIN_PASSWORD = cfg.get('admin_password', 'admin')
    PASSWORD_VIEWER = cfg.get('viewer_password', 'tamu')
    TELEGRAM_BOT_TOKEN = cfg.get('telegram_bot_token', '')
    TELEGRAM_CHAT_ID = cfg.get('telegram_chat_id', '')
    SERVICE_NAME = cfg.get('service_name', 'monitoring-wifi.service')
    APP_PORT = int(cfg.get('app_port', 5002))

# --- CACHE ---
MK_RES = {} 
MK_CACHE = {} 

# DATABASE FUNCTIONS
def init_default_db():
    return {
        "server": {
            "id": "server_utama", "name": "SERVER UTAMA", 
            "coordinates": [-6.1754, 106.8272], # Default Monas, Jakarta
            "login": {"host":"", "user":"", "pass":"", "port":8728}, 
            "status": "online", "manual_wan": "", "ping_target": "8.8.8.8",
            "port_config": {"lan": "5", "sfp": "0"}
        },
        "extra_routers": [], "odps": [], "clients": []
    }

# --- TIMEZONE HELPERS ---
def get_local_now():
    """Returns server local time"""
    return datetime.now()

def add_log(name, status, msg):
  with log_lock:
    try:
        logs = []
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r') as f: logs = json.load(f)
        
        # Anti-Flood: Check if last log is same status for same client
        if logs and logs[0].get('name') == name and logs[0].get('status') == status:
            return

        logs.insert(0, {"time": get_local_now().strftime("%Y-%m-%d %H:%M:%S"), "name": name, "status": status, "msg": msg})
        _safe_write_json(LOG_FILE, logs[:100])
    except: pass

def send_telegram_message(text):
    """Send text message to Telegram using settings.automation or environment variables as fallback"""
    # 1. Try Settings First
    settings = load_settings()
    tg_conf = settings.get('automation', {}).get('telegram', {})
    
    token = tg_conf.get('bot_token') or TELEGRAM_BOT_TOKEN
    chat_id = tg_conf.get('chat_id') or TELEGRAM_CHAT_ID
    
    if not token or not chat_id: return {"status":"skipped"}
    
    try:
        requests.post(f"https://api.telegram.org/bot{token}/sendMessage", 
                      data={'chat_id': chat_id, 'text': text}, timeout=10)
        return {"status": "ok"}
    except Exception as e: return {"status": "error", "msg": str(e)}

def dispatch_telegram_event(event_type, data):
    """
    Template-based notification dispatcher
    event_type: 'up', 'down', 'backup', 'startup'
    data: dict with placeholder values
    """
    settings = load_settings()
    tg_conf = settings.get('automation', {}).get('telegram', {})
    
    # Check if Globally Enabled
    if not tg_conf.get('enabled', False): return
    
    # Check Specific Trigger
    notifs = tg_conf.get('notifications', {})
    
    mapping = {
        'up': 'online',
        'down': 'offline',
        'backup': 'backup_report',
        'startup': 'startup_report'
    }
    
    key = mapping.get(event_type)
    if key and not notifs.get(key, False): return # Disabled by user preference
    
    # Prepare Template
    msg = ""
    dt = get_local_now()
    date_str = dt.strftime('%d-%m-%Y') # DD-MM-YYYY
    time_str = dt.strftime('%H:%M:%S')
    
    # --- AUTO CALCULATE TOTALS IF 0 ---
    if data.get('total_online', 0) == 0 and data.get('total_offline', 0) == 0:
        try:
            db_tmp = load_db()
            clients_tmp = db_tmp.get('clients', [])
            # Active = Online + Isolir
            t_on = sum(1 for c in clients_tmp if c.get('status') in ['online', 'isolir'])
            t_off = sum(1 for c in clients_tmp if c.get('status') == 'offline')
            data['total_online'] = t_on
            data['total_offline'] = t_off
        except Exception as e:
            print(f"[TG CALC ERROR] {e}")
    # ----------------------------------
    
    # --- CUSTOM TEMPLATE LOGIC ---
    tpl_up = tg_conf.get('template_up', '').strip()
    tpl_down = tg_conf.get('template_down', '').strip()
    
    # Prepare Template Data
    tpl_data = {
        'name': str(data.get('name', '-')),
        'ip': str(data.get('ip', '-')),
        'status': 'ONLINE' if event_type == 'up' else 'OFFLINE',
        'date': date_str,
        'time': time_str,
        'total_online': str(data.get('total_online', 0)),
        'total_offline': str(data.get('total_offline', 0)),
        'packet': str(data.get('packet', '-'))
    }
    
    def apply_tpl(tpl, defaults):
        for k, v in defaults.items():
            tpl = tpl.replace('{' + k + '}', v)
        return tpl

    try:
        if event_type == 'up':
            if tpl_up:
                msg = apply_tpl(tpl_up, tpl_data)
            else:
                msg = f"✅ PPPoE Connected\nTanggal: {date_str}\nJam: {time_str}\n"
                msg += f"User: {data.get('name', '-')}\n"
                msg += f"IP Client: {data.get('ip', '-')}\n"
                msg += f"Total Active: {data.get('total_online', 0)} Client\n"
                msg += f"Total Disconnected: {data.get('total_offline', 0)} Client\n"
                msg += f"Service: {data.get('packet', '-')}"
            
        elif event_type == 'down':
            if tpl_down:
                msg = apply_tpl(tpl_down, tpl_data)
            else:
                msg = f"❌ PPPoE Disconnected\nTanggal: {date_str}\nJam: {time_str}\n"
                msg += f"User: {data.get('name', '-')}\n"
                msg += f"Total Active: {data.get('total_online', 0)} Client\n"
                msg += f"Total Disconnected: {data.get('total_offline', 0)} Client"
            
        elif event_type == 'backup':
            msg = f"💾 Auto Backup Success\nTanggal: {date_str}\nJam: {time_str}\n"
            msg += f"File: {data.get('filename', '-')}\n"
            msg += f"Size: {data.get('size', '-')} KB"
            
        elif event_type == 'startup':
            msg = f"⚙️ System Startup\nNMS Service Started\nTanggal: {date_str}\nJam: {time_str}"
            
        if msg:
            send_telegram_message(msg)
            
    except Exception as e:
        print(f"[TELEGRAM] Dispatch Error: {e}")

def send_telegram_file(filepath, caption):
    settings = load_settings()
    tg_conf = settings.get('automation', {}).get('telegram', {})
    
    token = tg_conf.get('bot_token') or TELEGRAM_BOT_TOKEN
    chat_id = tg_conf.get('chat_id') or TELEGRAM_CHAT_ID

    if not token or not chat_id: 
        return {"status":"error", "msg": "Token atau Chat ID kosong"}
    try:
        with open(filepath, 'rb') as f:
            response = requests.post(
                f"https://api.telegram.org/bot{token}/sendDocument", 
                data={'chat_id': chat_id, 'caption': caption}, 
                files={'document': f}, 
                timeout=30
            )
        
        if response.status_code == 200:
            result = response.json()
            if result.get('ok'):
                return {"status": "ok", "msg": "File berhasil dikirim ke Telegram"}
            else:
                return {"status": "error", "msg": f"Telegram API Error: {result.get('description', 'Unknown')}"}
        else:
            return {"status": "error", "msg": f"HTTP {response.status_code}: {response.text[:200]}"}
    except Exception as e: 
        return {"status": "error", "msg": f"Exception: {str(e)}"}

# ==============================================================================
#  BILLING HELPER FUNCTIONS
# ==============================================================================
def get_router_data(router_id, db=None):
    """ Ambil kredensial router berdasarkan ID """
    if db is None:
        db = load_db()
    
    if router_id == 'server_utama':
        return load_db().get("server", {})
    else:
        return next((r for r in load_db().get("extra_routers", []) if r['id'] == router_id), None)

def ensure_isolir_profile(router_id):
    """ Buat profil ISOLIR jika belum ada """
    conn = None
    try:
        router_data = get_router_data(router_id)
        if not router_data or not router_data.get('login', {}).get('host'):
            return False
        
        settings = load_settings()
        isolir_prof_name = settings.get('billing', {}).get('isolir_profile', 'ISOLIR')
        
        login = router_data['login']
        conn = routeros_api.RouterOsApiPool(
            login['host'], 
            username=login['user'], 
            password=login['pass'],
            port=int(login.get('port', 8728)), 
            plaintext_login=True
        )
        api = conn.get_api()
        
        # Check if profile exists
        profiles = api.get_resource('/ppp/profile').get()
        isolir_exists = any(p.get('name') == isolir_prof_name for p in profiles)
        
        if not isolir_exists:
            # Create profile with 1kbps bandwidth
            api.get_resource('/ppp/profile').add(
                name=isolir_prof_name,
                **{'rate-limit': '1k/1k'}
            )
            print(f"[BILLING] Created profile {isolir_prof_name} on {router_id}")
        
        return True
    except Exception as e:
        print(f"[ERROR] Failed to ensure ISOLIR profile on {router_id}: {e}")
        return False
    finally:
        if conn: conn.disconnect()

def get_pppoe_current_profile(username, router_id):
    """ Cek profil user PPPoE saat ini dari Mikrotik """
    conn = None
    try:
        router_data = get_router_data(router_id)
        if not router_data or not router_data.get('login', {}).get('host'):
            return None
        
        login = router_data['login']
        conn = routeros_api.RouterOsApiPool(
            login['host'], 
            username=login['user'], 
            password=login['pass'],
            port=int(login.get('port', 8728)), 
            plaintext_login=True
        )
        api = conn.get_api()
        secrets = api.get_resource('/ppp/secret')
        user = secrets.get(name=username)
        
        if user and len(user) > 0:
            profile = user[0].get('profile', 'default')
            return profile
        return None
    except Exception as e:
        print(f"[ERROR] get_pppoe_current_profile: {e}")
        return None
    finally:
        if conn: conn.disconnect()

def change_pppoe_profile(username, new_profile, router_id):
    """ Ganti profil secret PPPoE """
    conn = None
    try:
        router_data = get_router_data(router_id)
        if not router_data or not router_data.get('login', {}).get('host'):
            return {"status": "error", "msg": "Router not found"}
        
        login = router_data['login']
        conn = routeros_api.RouterOsApiPool(
            login['host'], 
            username=login['user'], 
            password=login['pass'],
            port=int(login.get('port', 8728)), 
            plaintext_login=True
        )
        api = conn.get_api()
        
        secrets = api.get_resource('/ppp/secret')
        user = secrets.get(name=username)
        
        # Fallback: Case-insensitive search
        if not user:
            all_secrets = secrets.get()
            user = [s for s in all_secrets if s.get('name', '').lower() == username.lower()]

        if user:
            # Use .id or id for compatibility
            target_id = user[0].get('.id') or user[0].get('id')
            secrets.set(id=target_id, profile=new_profile)
            return {"status": "ok"}
        else:
            return {"status": "error", "msg": f"User '{username}' no match in MikroTik"}
    except Exception as e:
        return {"status": "error", "msg": str(e)}
    finally:
        if conn: conn.disconnect()

def kick_pppoe_user(username, router_id):
    """ Kick user PPPoE dari active connections """
    conn = None
    try:
        router_data = get_router_data(router_id)
        if not router_data or not router_data.get('login', {}).get('host'):
            return {"status": "error", "msg": "Router not found"}
        
        login = router_data['login']
        conn = routeros_api.RouterOsApiPool(
            login['host'], 
            username=login['user'], 
            password=login['pass'],
            port=int(login.get('port', 8728)), 
            plaintext_login=True
        )
        api = conn.get_api()
        
        actives = api.get_resource('/ppp/active')
        session = actives.get(name=username)
        
        if session:
            # Use .id or id for compatibility
            target_id = session[0].get('.id') or session[0].get('id')
            actives.remove(id=target_id)
        
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "msg": str(e)}
    finally:
        if conn: conn.disconnect()

def add_to_address_list(address, list_name, router_id, comment=""):
    """ Tambahkan entry ke IP Firewall Address-list """
    conn = None
    try:
        router_data = get_router_data(router_id)
        if not router_data or not router_data.get('login', {}).get('host'):
            return {"status": "error", "msg": "Router not found"}
        
        login = router_data['login']
        conn = routeros_api.RouterOsApiPool(
            login['host'], 
            username=login['user'], 
            password=login['pass'],
            port=int(login.get('port', 8728)), 
            plaintext_login=True
        )
        api = conn.get_api()
        
        resource = api.get_resource('/ip/firewall/address-list')
        
        # SMART CLEANUP: If comment is provided, remove old entries with SAME comment
        # This handles IP changes (removes old IP block, adds new IP block)
        if comment:
            old_entries = resource.get(list=list_name, comment=comment)
            for entry in old_entries:
                target_id = entry.get('.id') or entry.get('id')
                resource.remove(id=target_id)

        # Double check by address to avoid duplicates
        existing = resource.get(address=address, list=list_name)
        if not existing:
            resource.add(address=address, list=list_name, comment=comment)
        
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "msg": str(e)}
    finally:
        if conn: conn.disconnect()

def remove_from_address_list(address, list_name, router_id, comment=None):
    """ Hapus entry dari IP Firewall Address-list """
    conn = None
    try:
        router_data = get_router_data(router_id)
        if not router_data or not router_data.get('login', {}).get('host'):
            return {"status": "error", "msg": "Router not found"}
        
        login = router_data['login']
        conn = routeros_api.RouterOsApiPool(
            login['host'], 
            username=login['user'], 
            password=login['pass'],
            port=int(login.get('port', 8728)), 
            plaintext_login=True
        )
        api = conn.get_api()
        
        resource = api.get_resource('/ip/firewall/address-list')
        
        # Cleanup by Comment (Name) - Most reliable for IP changes
        if comment:
            entries = resource.get(list=list_name, comment=comment)
            for e in entries:
                target_id = e.get('.id') or e.get('id')
                resource.remove(id=target_id)
        
        # Cleanup by specific IP
        if address and address != '-':
            entries = resource.get(address=address, list=list_name)
            for e in entries:
                target_id = e.get('.id') or e.get('id')
                resource.remove(id=target_id)
        
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "msg": str(e)}
    finally:
        if conn: conn.disconnect()

def calculate_due_date(year, month, billing_day):
    """ Hitung tanggal jatuh tempo (dengan handle tanggal 31/kabisat) """
    import calendar
    
    # Get last day of the month
    last_day = calendar.monthrange(year, month)[1]
    
    # Use the smaller value (handles date 31 in Feb, Apr, etc)
    safe_day = min(billing_day, last_day)
    
    return datetime(year, month, safe_day)

# ==============================================================================
#  HELPER & MONITORING ENGINE
# ==============================================================================
def parse_size(size_str):
    try: return float(size_str) if size_str else 0.0
    except: return 0.0

def parse_hotspot_limit_bytes(datalimit):
    if datalimit is None: return None
    raw_data = str(datalimit).upper().strip()
    if raw_data == "" or raw_data == "0": return "0"
    try:
        num_str = ''.join([c for c in raw_data if c.isdigit() or c == '.'])
        num = float(num_str) if num_str else 0
        if 'G' in raw_data: bytes_total = int(num * 1024 * 1024 * 1024)
        elif 'M' in raw_data: bytes_total = int(num * 1024 * 1024)
        elif 'K' in raw_data: bytes_total = int(num * 1024)
        else: bytes_total = int(num)
        return str(bytes_total)
    except:
        return "0"

def format_speed(bps):
    try:
        val = float(bps)
        if val >= 1000000000: return f"{val/1000000000:.1f}G"
        if val >= 1000000: return f"{val/1000000:.1f}M"
        if val >= 1000: return f"{val/1000:.0f}k"
        return f"{val:.0f}"
    except: return "0"

def get_cpu_temp():
    try:
        with open("/sys/class/thermal/thermal_zone0/temp", "r") as f: return int(int(f.read().strip()) / 1000)
    except: return 0

def apply_bulk_updates(updates):
    """ Terapkan update status secara atomik lewat SQLite """
    if not updates: return
    try:
        # Invalidate Topology Cache V3.3.9 to ensure next load gets pings/status
        global _TOPO_CACHE
        _TOPO_CACHE['expiry'] = 0
        
        db.apply_bulk_updates(updates)
    except Exception as e:
        print(f"[DB ERROR] apply_bulk_updates failed: {e}")

def ping_ip_manual(ip):
    if not ip or ip in ["0.0.0.0", "Dynamic"]: return False
    try:
        param_c = '-n' if platform.system().lower() == 'windows' else '-c'
        param_w = '-w' if platform.system().lower() == 'windows' else '-W'
        wait_v = '1000' if platform.system().lower() == 'windows' else '1'
        subprocess.check_output(['ping', param_c, '1', param_w, wait_v, ip], stderr=subprocess.STDOUT)
        return True
    except: return False

def ping_ip_linux(ip):
    """Ping menggunakan /usr/bin/ping (Linux asli) untuk stabilitas"""
    if not ip or ip in ["0.0.0.0", "Dynamic"]: return False
    # Hanya untuk Linux
    if platform.system().lower() == 'windows': return ping_ip_manual(ip)
    try:
        subprocess.check_output(['/usr/bin/ping', '-c', '1', '-W', '1', ip], stderr=subprocess.STDOUT)
        return True
    except: return False

def fetch_single_router_data(router_id, login_data, router_config):
    if not routeros_api or not login_data.get("host"): 
        MK_RES[router_id] = {"error": True}
        return

    conn = None
    try:
        api_port = int(login_data.get("port", 8728))
        conn = routeros_api.RouterOsApiPool(
            login_data["host"], username=login_data["user"], password=login_data["pass"], 
            port=api_port, plaintext_login=True
        )
        api = conn.get_api()

        # Load settings and current DB for profile sync
        settings = load_settings()
        db_now = load_db()

        res = api.get_resource('/system/resource').get()
        ident = api.get_resource('/system/identity').get()
        
        det_lan = 0; det_sfp = 0
        try:
            all_ifaces = api.get_resource('/interface').get()
            for i in all_ifaces:
                itype = i.get('type', '').lower()
                idefault = i.get('default-name', '').lower()
                if (itype == 'ether') or ('ether' in idefault) or ('sfp' in idefault):
                    if 'sfp' in i.get('name', '').lower() or 'sfp' in idefault: det_sfp += 1
                    else: det_lan += 1
            if det_lan == 0 and det_sfp == 0: det_lan = 1
        except: det_lan = 1

        wan_rx_tot = 0; wan_tx_tot = 0; wan_name = "Scanning..."
        manual_wan = router_config.get('manual_wan', '')
        target_ifaces = []
        if manual_wan:
            target_ifaces = [x.strip() for x in manual_wan.split(',') if x.strip()]
            wan_name = manual_wan
        else:
            try:
                routes = api.get_resource('/ip/route').get(dst_address='0.0.0.0/0', active='true')
                if routes:
                    gw = routes[0].get('gateway')
                    arp = api.get_resource('/ip/arp').get(address=gw)
                    detected = arp[0].get('interface') if arp else gw
                    target_ifaces = [detected]; wan_name = detected
            except: pass
            
        for iface in target_ifaces:
            try:
                traf = api.get_resource('/interface').call('monitor-traffic', {'interface': iface, 'once': 'true'})
                if traf:
                    wan_rx_tot += parse_size(traf[0].get('rx-bits-per-second', '0'))
                    wan_tx_tot += parse_size(traf[0].get('tx-bits-per-second', '0'))
            except: pass

        ping_tgt = router_config.get('ping_target', '8.8.8.8')
        ext_ping_res = "Wait..."
        try:
            p_res = api.get_resource('/').call('ping', {'address': ping_tgt, 'count': '1'})
            ext_ping_res = p_res[0].get('time', 'RTO') if p_res else "RTO"
        except: ext_ping_res = "Error"

        try:
            arps = api.get_resource('/ip/arp').get()
            arp_map = {a.get('address'): True for a in arps if a.get('complete')=='true' or a.get('mac-address')}
        except: arp_map = {}

        try:
            leases = api.get_resource('/ip/dhcp-server/lease').get()
            lease_map = {l.get('address'): l.get('status') for l in leases}
        except: lease_map = {}

        try:
            nws = api.get_resource('/tool/netwatch').get()
            netwatch_map = {n.get('host'): n.get('status') for n in nws}
        except: netwatch_map = {}

        isolir_prof_name = settings.get('billing', {}).get('isolir_profile', 'ISOLIR')
        
        # ISOLIR MAP (V3.1.9)
        try:
            i_list = api.get_resource('/ip/firewall/address-list').get(list=isolir_prof_name)
            isolir_map = {i.get('address'): True for i in i_list}
        except: isolir_map = {}

        secrets = api.get_resource('/ppp/secret').get()
        actives = api.get_resource('/ppp/active').get()
        
        # HOTSPOT SUPPORT (V3)
        hotspot_actives = []
        try:
            hotspot_actives = api.get_resource('/ip/hotspot/active').get()
        except: pass

        active_map = {a.get('name'): a for a in actives}
        secret_map = {s.get('name'): s for s in secrets}
        
        # Local Secrets Cache
        MK_CACHE[router_id] = {
            "secrets": [{"name":s['name'], "profile":s.get('profile','')} for s in secrets],
            "radius_candidates": [] 
        }

        # Filter RADIUS Candidates (Active but not in Secret and not in NMS DB)
        # Note: NMS DB filtering happens at API level to avoid race conditions here
        # Here we just collect ALL actives that are NOT in secrets
        radius_candidates = []
        
        # 1. PPPoE RADIUS Candidates
        for a in actives:
            if a.get('name') not in secret_map:
                radius_candidates.append({
                    "name": a.get('name'), 
                    "service": a.get('service', 'pppoe'),
                    "ip": a.get('address'),
                    "uptime": a.get('uptime'),
                    "profile": a.get('profile', ''),
                    "type": "pppoe_radius"
                })
        
        # 2. Hotspot RADIUS Candidates
        for h in hotspot_actives:
            radius_candidates.append({
                "name": h.get('user'), 
                "service": "hotspot",
                "ip": h.get('address'),
                "uptime": h.get('uptime'),
                "profile": h.get('profile', ''),
                "type": "hotspot"
            })
            
        MK_CACHE[router_id]["radius_candidates"] = radius_candidates

        # Queue Traffic Map (Restored from V2.9)
        traffic_map = {}
        try:
            all_queues = api.get_resource('/queue/simple').get()
            for q in all_queues:
                tgt = q.get('target','').split('/')[0]
                rate = q.get('rate','0/0').split('/')
                traffic_map[tgt] = int(rate[0]) + int(rate[1])
        except: pass

        # MikroTik Ping Map (Restore V2.9 Feature)
        mikrotik_ping_map = {}
        try:
            db_now = load_db()
            mikrotik_ping_clients = [
                c for c in db_now.get('clients', [])
                if c.get('managed_by', 'server_utama') == router_id
                and not c.get('credentials', {}).get('pppoe_user')
                and c.get('monitor_mode') == 'mikrotik_ping'
                and c.get('ip') and c.get('ip') not in ["0.0.0.0", "Dynamic", "-", ""]
            ]
            
            for client in mikrotik_ping_clients:
                try:
                    # Ping count 2 to be sure
                    ping_res = api.get_resource('/').call('ping', {'address': client['ip'], 'count': '2'})
                    # Check if any packet returned (not timeout)
                    # Robust Ping Check (Bug Fix #3)
                    is_up = any(
                        (p.get('received') == '1') or 
                        (p.get('ttl') and p.get('size')) or 
                        (p.get('time') and p.get('time') != 'RTO')
                        for p in ping_res
                    )
                    mikrotik_ping_map[client['ip']] = is_up
                except:
                    mikrotik_ping_map[client['ip']] = False
        except Exception as e:
            print(f"[WARN] MK Ping Loop Error: {e}")

        
        # Simpan ke Global Cache untuk API
        MK_RES[router_id] = {
            "cpu": int(res[0].get('cpu-load', 0)) if res else 0,
            "uptime": res[0].get('uptime', '00:00:00') if res else "-",
            "board": res[0].get('board-name', '-') if res else "-",
            "version": res[0].get('version', '-') if res else "-",
            "identity": ident[0].get('name', '-') if ident else "-",
            "wan_name": wan_name,
            "wan_rx": format_speed(wan_rx_tot),
            "wan_tx": format_speed(wan_tx_tot),
            "wan_rx_raw": wan_rx_tot,
            "wan_tx_raw": wan_tx_tot,
            "port_lan": det_lan,
            "port_sfp": det_sfp,
            "ext_ping": ext_ping_res,
            "ping_target": ping_tgt
        }

        # Race Condition Fix: Don't load-modify-save the whole DB.
        # Instead, calculate updates locally and apply atomically.
        db_snap = load_db()
        updates = []
        pending_notifs = []
        
        for c in db_snap.get('clients', []):
            if c.get('managed_by', 'server_utama') == router_id:
                monitor_mode = c.get('monitor_mode', 'default')
                c_ip = c.get('ip')
                pppoe_u = c.get('credentials', {}).get('pppoe_user')
                c_id = c.get('id')
                
                upd_entry = {'id': c_id}
                has_upd = False

                if pppoe_u and pppoe_u.strip() != "":
                    is_active = pppoe_u in active_map
                    is_disabled = secret_map.get(pppoe_u,{}).get('disabled')=='true'
                    
                    # Logic Isolir V3.1.8: Periksa Profil (Penting agar yang masih nyambung tapi profil isolir tetap merah)
                    isolir_prof = settings.get('billing', {}).get('isolir_profile', 'ISOLIR')
                    cur_p = secret_map.get(pppoe_u, {}).get('profile') or active_map.get(pppoe_u, {}).get('profile')
                    
                    if cur_p == isolir_prof:
                        new_stat = 'isolir'
                    elif is_active:
                        new_stat = 'online'
                    elif is_disabled:
                        new_stat = 'isolir'
                    else:
                        new_stat = 'offline'
                    
                    if c.get('status') != new_stat:
                        upd_entry['status'] = new_stat; has_upd = True
                        
                        # SMART LOG MSG: Check mode (PPPoE or Radius)
                        cl_mode = c.get('mode', 'pppoe')
                        mode_label = "PPPoE" if cl_mode == 'pppoe' else ("Radius" if cl_mode == 'pppoe_radius' else "Client")
                        
                        msg_pppoe = f"{mode_label} Connected" if new_stat == 'online' else (f"{mode_label} Isolir" if new_stat == 'isolir' else f"{mode_label} Disconnected")
                        add_log(c['name'], new_stat, msg_pppoe)
                        
                        # Auto-Notification
                        evt = 'up' if new_stat == 'online' else ('down' if new_stat == 'offline' else None)
                        if evt:
                            d_alert = {
                                'name': c['name'], 
                                'ip': active_map[pppoe_u].get('address') if is_active else c.get('ip', '-'),
                                'packet': c.get('packet_name') or '-',
                                'total_online': 0, 'total_offline': 0
                            }
                            pending_notifs.append((evt, d_alert))
                    
                    # BUG FIX V3: AGGRESSIVE IP SYNC & CLEANUP
                    # Jika aktif, ambil IP terbaru. Jika tidak aktif, PAKSA hapus IP agar tidak zombie.
                    if is_active: 
                        new_ip = active_map[pppoe_u].get('address')
                        if c.get('ip') != new_ip: 
                            upd_entry['ip'] = new_ip
                            has_upd = True
                    else:
                        if c.get('ip') != '-':
                            upd_entry['ip'] = '-'; has_upd = True
                    
                    # Sync Profile / Packet Name (V3.1.7: Ignore ISOLIR profile to prevent data loss)
                    cur_profile = secret_map.get(pppoe_u, {}).get('profile')
                    isolir_prof = settings.get('billing', {}).get('isolir_profile', 'ISOLIR')
                    if cur_profile and cur_profile != isolir_prof and c.get('packet_name') != cur_profile:
                        upd_entry['packet_name'] = cur_profile; has_upd = True
                    
                    # Clear ping_ms for PPPoE clients
                    if c.get('ping_ms') != -1:
                        upd_entry['ping_ms'] = -1; has_upd = True
                
                elif c_ip and c_ip not in ["0.0.0.0", "Dynamic", "-", ""]:
                    is_online = False
                    
                    # Monitor Logic
                    if monitor_mode == 'netwatch':
                        if netwatch_map.get(c_ip) == 'up': is_online = True
                    elif monitor_mode == 'mikrotik_ping':
                        if mikrotik_ping_map.get(c_ip, False): is_online = True
                    elif monitor_mode == 'api':
                        # COMPOSITE CHECK (V2.9 Feature)
                        if c_ip in arp_map: is_online = True
                        elif lease_map.get(c_ip) == 'bound': is_online = True
                        elif traffic_map.get(c_ip, 0) > 10: is_online = True
                    else: continue

                    new_stat = 'online' if is_online else 'offline'
                    
                    # Logic Isolir Statik (V3.1.9)
                    if isolir_map.get(c_ip):
                        new_stat = 'isolir'
                    
                    if c.get('status') != new_stat:
                        upd_entry['status'] = new_stat; has_upd = True
                        
                        src_label = "Router API"
                        if monitor_mode == 'mikrotik_ping': src_label = "MikroTik Ping" 
                        elif monitor_mode == 'netwatch': src_label = "Netwatch"
                        
                        msg_api = f"Connected ({src_label})" if new_stat == 'online' else f"Disconnected ({src_label})"
                        # If msg is just "Connected (Router API)", let's make it cleaner for toast
                        add_log(c['name'], new_stat, msg_api)

                        # Auto-Notification
                        evt = 'up' if new_stat == 'online' else 'down'
                        d_alert = {
                            'name': c['name'],
                            'ip': c_ip,
                            'packet': c.get('packet_name') or '-',
                            'total_online': 0, 'total_offline': 0
                        }
                        pending_notifs.append((evt, d_alert))
                
                # RADIUS ENFORCEMENT LOOP (V3)
                # If client is ISOLIR, we must ensure they are blocked even if they change IP
                if c.get('status') == 'isolir' and c.get('mode') == 'pppoe_radius':
                    # Check if user is currently active (Online with new IP?)
                    curr_ip = None
                    if pppoe_u in active_map: curr_ip = active_map[pppoe_u].get('address')
                    
                    if curr_ip:
                        # User is active! Check if this IP is already blocked
                        # We use a cache or just do it blindly? Doing it blindly might spam API.
                        # Optimization: Check if IP matches what we think is isolated?
                        # Or better: check address-list presence.
                        try:
                            fw_list = api.get_resource('/ip/firewall/address-list')
                            # Check if current IP is in ISOLIR list
                            isolir_prof = settings.get('billing', {}).get('isolir_profile', 'ISOLIR')
                            is_blocked = fw_list.get(address=curr_ip, list=isolir_prof)
                            
                            if not is_blocked:
                                # Not in list, add it
                                # Remove anything with same comment first
                                old_entries = fw_list.get(list=isolir_prof, comment=f"{isolir_prof}_{c['name']}")
                                for o in old_entries: fw_list.remove(id=o['.id'])
                                
                                add_to_address_list(curr_ip, isolir_prof, router_id, comment=f"{isolir_prof}_{c['name']}")
                                
                                # 3. Kick
                                kick_pppoe_user(pppoe_u, router_id)
                        except: pass
                
                if has_upd: updates.append(upd_entry)

        if updates: 
            apply_bulk_updates(updates)
            # Dispatch Notifications AFTER updates applied to DB (for accurate counts)
            for p_evt, p_data in pending_notifs:
                dispatch_telegram_event(p_evt, p_data)



    except Exception as e:
        print(f"[ERROR] MikroTik Fetch ({router_id}): {e}")
        MK_RES[router_id] = {"error": True}
    finally:
        if conn: conn.disconnect()



# --- THREAD LOOPS ---
def monitor_mikrotik_loop():
    while True:
        try:
            data = load_db()
            srv = data.get("server", {})
            if srv.get("login", {}).get("host"): fetch_single_router_data("server_utama", srv["login"], srv)
            for rtr in data.get("extra_routers", []):
                if rtr.get("login", {}).get("host"): fetch_single_router_data(rtr["id"], rtr["login"], rtr)
        except: pass
        time.sleep(10)


def turbo_ping_loop():
    time.sleep(5)
    while True:
        try:
            data_db = load_db()
            updates = []
            now = time.time()
            
            # 1. Separate Clients by Mode
            linux_clients = []
            python_clients = []
            
            for c in data_db.get('clients', []):
                # Skip PPPoE or non-IP users
                if c.get('credentials', {}).get('pppoe_user'): continue
                if not c.get('ip') or c.get('ip') in ["0.0.0.0", "Dynamic", "-", ""]: continue
                
                mode = c.get('monitor_mode', 'python_ping') # Default to python_ping (V2.9 behavior)
                if mode == 'default': mode = 'python_ping'
                
                if mode == 'linux_ping':
                    linux_clients.append(c)
                elif mode == 'python_ping':
                    python_clients.append(c)
            
            pending_notifs = []

            # Create update helper
            def prepare_update(cl, is_alive, ms_val, src_name):
                # Helper to create atomic update dict
                upd = {'id': cl['id']}
                changed = False
                
                if cl.get('ping_ms') != ms_val:
                    upd['ping_ms'] = ms_val; changed = True
                
                last_ok = cl.get('last_seen_ts', 0)
                curr_stat = cl.get('status', 'offline')
                
                if is_alive:
                    upd['last_seen_ts'] = now; changed = True
                    if curr_stat != 'online':
                        upd['status'] = 'online'; changed = True
                        # SMART LOG MSG: For Static IP (Ping)
                        add_log(cl['name'], 'online', f"Connected ({src_name} OK)")
                        
                        # Automation Alert
                        data_alert = {
                            'name': cl['name'], 
                            'ip': cl.get('ip', '-'), 
                            'total_online': 0, # Will be calc in dispatch or backend
                            'total_offline': 0,
                            'packet': cl.get('packet_name') or cl.get('service_plan', '-')
                        }
                        pending_notifs.append(('up', data_alert))
                else:
                    # Debounce 20s (V3 Standard)
                    if (now - last_ok) > 20:
                        is_isolir_db = cl.get('billing', {}).get('payment_status') == 'overdue'
                        target_stat = 'isolir' if is_isolir_db else 'offline'
                        
                        if curr_stat != target_stat:
                            upd['status'] = target_stat; changed = True
                            add_log(cl['name'], target_stat, f"Disconnected ({src_name} RTO)")
                            # Automation
                            data_alert = {'name': cl['name'], 'total_online': 0, 'total_offline': 0} # Totals calc in dispatch
                            pending_notifs.append(('down', data_alert))
                
                if changed: updates.append(upd)

            # 2. Process LINUX Ping (Sequential System Ping)
            for cl in linux_clients:
                try:
                    alive = ping_ip_linux(cl['ip']) # Uses absolute path /usr/bin/ping
                    ms = 10 if alive else -1
                    prepare_update(cl, alive, ms, "Linux Ping")
                except: pass

            # 3. Process PYTHON Ping (Turbo / Batch)
            if python_clients:
                if ICMP_READY:
                    targets = [c['ip'] for c in python_clients]
                    if targets:
                        try:
                            results = turbo_ping(targets, count=3, interval=0.5, timeout=1.5, privileged=True)
                            results_map = {r.address: r for r in results}
                            
                            for cl in python_clients:
                                res = results_map.get(cl['ip'])
                                if not res: continue
                                
                                alive = res.is_alive
                                val = int(res.min_rtt)
                                ms = 1 if val < 1 else val
                                if not alive: ms = -1
                                
                                # Fallback if permission error
                                if not alive and ping_ip_manual(cl['ip']):
                                    alive = True; ms = 10
                                    
                                prepare_update(cl, alive, ms, "Turbo Ping")
                        except Exception as e:
                            print(f"[TURBO ERROR] {e}")
                else:
                    # Fallback to manual if ICMP lib not installed
                    for cl in python_clients:
                        alive = ping_ip_manual(cl['ip'])
                        ms = 10 if alive else -1
                        prepare_update(cl, alive, ms, "System Ping")

            # 4. Apply Atomic Updates
            if updates: 
                apply_bulk_updates(updates)
                # Dispatch Notifications AFTER updates applied to DB (for accurate counts)
                for p_evt, p_data in pending_notifs:
                    dispatch_telegram_event(p_evt, p_data)
            
        except Exception as e:
            print(f"[PING LOOP ERROR] {e}")
        time.sleep(5)

def auto_backup_logic(force=False):
    """Logic Backup Utama"""
    try:
        settings = load_settings()
        bk_conf = settings.get('automation', {}).get('backup', {})
        
        if not force and not bk_conf.get('enabled', False): return
        
        # Check Time (HH:MM)
        now = get_local_now()
        sched_time = bk_conf.get('schedule_time', '02:00')
        hour = int(sched_time.split(':')[0])
        minute = int(sched_time.split(':')[1])
        
        if force or (now.hour == hour and now.minute == minute):
            global LAST_BACKUP_DATE
            today_str = now.strftime('%d-%m-%y') # Format request: 30-01-26
            if LAST_BACKUP_DATE == today_str: return
            
            # 1. Prepare Zip
            BACKUP_DIR = os.path.join(SCRIPT_DIR, 'backups')
            if not os.path.exists(BACKUP_DIR): os.makedirs(BACKUP_DIR)
            
            # Filename for internal storage
            filename = f"BACKUP_{now.strftime('%d-%m-%y_%H%M')}.zip"
            zip_path = os.path.join(BACKUP_DIR, filename)
            
            includes = bk_conf.get('include_files', ['topology.db', 'settings.json', 'finance.json', 'config.json', 'license.key', 'app.py', 'db_manager.py', 'license_utils.py', 'wa-bridge.js', 'package.json', 'keygen.py', 'CARA_PAKAI_KEYGEN.md', 'DISTRIBUSI_KLIEN.md', 'templates', 'static'])
            
            # Safety: Ensure app.py is ALWAYS included even if removed from settings
            if 'app.py' not in includes:
                includes.append('app.py')
            if 'static' not in includes:
                includes.append('static')
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for f in includes:
                    # Use absolute path for robustness
                    f_path = os.path.abspath(os.path.join(SCRIPT_DIR, f))
                    if not os.path.exists(f_path): continue
                    
                    if os.path.isfile(f_path):
                        zipf.write(f_path, f)
                    elif os.path.isdir(f_path):
                        for root, _, files in os.walk(f_path):
                            for file in files:
                                full_p = os.path.join(root, file)
                                rel_p = os.path.relpath(full_p, SCRIPT_DIR)
                                zipf.write(full_p, rel_p)
            
            LAST_BACKUP_DATE = today_str # Mark as done
            
            # 2. Send Telegram
            tg_conf = settings.get('automation', {}).get('telegram', {})
            if tg_conf.get('notifications', {}).get('backup_report', True):
                if tg_conf.get('enabled', False):
                    caption = f"💾 AUTO BACKUPS NMS V3 {now.strftime('%d-%m-%y')}"
                    send_telegram_file(zip_path, caption)
            
            # 3. Clean Old Backups
            keep_days = int(bk_conf.get('keep_days', 7))
            cutoff = time.time() - (keep_days * 86400)
            
            for f in os.listdir(BACKUP_DIR):
                fp = os.path.join(BACKUP_DIR, f)
                if os.path.getmtime(fp) < cutoff:
                    try: os.remove(fp)
                    except: pass
            
            time.sleep(70) # Prevent multiple execution in same minute

    except Exception as e:
        print(f"[BACKUP ERROR] {e}")

def auto_backup_loop():
    while True:
        try:
            auto_backup_logic()
        except: pass
        time.sleep(10)

def sync_billing_from_finance(client=None, db_data=None):
    """
    Proactively syncs a client's billing status/paid_until from finance history.
    Useful for legacy data or when finance entries were added without automatic triggers.
    """
    if not client: return False
    
    fin = load_finance()
    client_id = client.get('id')
    if not client_id: return False
    
    # Filter transactions for this client, sorted by date DESC
    txs = [t for t in fin.get('transactions', []) if t.get('client_id') == client_id 
           and t.get('category') in ['wifi_payment', 'Pembayaran WiFi']]
    if not txs: return False
    
    txs.sort(key=lambda x: x.get('date', ''), reverse=True)
    latest_tx = txs[0]
    
    changed = False
    import re
    note = latest_tx.get('note', '')
    date_ = latest_tx.get('date', '')
    
    # 1. Look for explicit "Lunas s/d YYYY-MM-DD" in note
    match = re.search(r'Lunas s/d (\d{4}-\d{2}-\d{2})', note)
    if match:
        new_expiry = match.group(1)
        if client.get('paid_until') != new_expiry:
            client['paid_until'] = new_expiry
            changed = True
    
    # 2. Update status based on paid_until vs NOW
    now = get_local_now()
    paid_until_str = client.get('paid_until')
    
    if 'billing' not in client: client['billing'] = {}
    old_status = client['billing'].get('payment_status')
    
    new_status = 'unpaid' # Default if no date or expired
    if paid_until_str:
        try:
            expiry_date = datetime.strptime(paid_until_str, '%Y-%m-%d').date()
            # If active until tomorrow or further, it's PAID
            # But if it ends TODAY or is already passed, it's UNPAID (Time to pay for NEXT period)
            if expiry_date > now.date():
                new_status = 'paid'
            else:
                new_status = 'unpaid'
        except:
            new_status = 'unpaid'

    if old_status != new_status:
        client['billing']['payment_status'] = new_status
        changed = True
            
    return changed

def run_billing_check(notify_only=False, target_user=None, force=False, template_mode='auto'):
    global GLOBAL_BILLING_HEARTBEAT
    GLOBAL_BILLING_HEARTBEAT = get_local_now().strftime('%d %b %Y, %H:%M:%S')
    try:
        # Load directly from billing.json
        billing_config = load_billing_config()
        
        if not force:
            # Check if isolation is disabled globally
            auto_isolir = billing_config.get('auto_isolir_enabled', True)
        else:
            auto_isolir = True
        
        db = load_db()
        now = get_local_now()
        grace_period = billing_config.get('grace_period_days', 3)
        isolir_profile = billing_config.get('isolir_profile', 'ISOLIR')
        
        changed = False
        processed_count = 0
        isolir_count = 0
        reactivate_count = 0
        wa_queue = [] # Queue for WhatsApp messages
        deferred_isolations = [] # NEW: Queue for MikroTik isolation commands
        
        # Localized month for templates
        settings = load_settings()
        pref_lang = billing_config.get('language') or settings.get('language', 'id')
        curr_month_id = get_month_name(now.month, pref_lang)
        
        db_data = load_db()
        target_found = False
        target_skip_reason = None
        
        if target_user:
            pass

        # Iterating through clients
        for idx, client in enumerate(db_data.get('clients', [])):
            try:
                # Get PPPoE user
                pppoe_user = client.get('credentials', {}).get('pppoe_user')
                
                # FILTERING LOGIC
                if target_user:
                    # Match against PPPoE User OR IP address OR Client Name
                    pp_user = (pppoe_user and pppoe_user.strip().lower() == target_user.strip().lower())
                    c_ip_match = (client.get('ip') and client.get('ip').strip() == target_user.strip())
                    c_name_match = (client.get('name') and client.get('name').strip().lower() == target_user.strip().lower())
                    
                    if not (pp_user or c_ip_match or c_name_match):
                        continue
                    target_found = True

                # [URGENT BUG FIX] FORCE MODE must be BEFORE Bypass/Enabled checks for manual testing
                if force and target_user and target_found:
                    router_id = client.get('managed_by', 'server_utama')
                    success = False
                    
                    # [BUG FIX V3.1.9] Save original_profile BEFORE change
                    current_profile_real = client.get('packet_name', 'default')
                    if pppoe_user:
                        try:
                            # Try to get live profile from Mikrotik
                            live_prof = get_pppoe_current_profile(pppoe_user, router_id)
                            if live_prof: current_profile_real = live_prof
                        except: pass
                    
                    if isolir_profile.upper() not in current_profile_real.upper():
                        if 'billing' not in client: client['billing'] = {}
                        client['billing']['original_profile'] = current_profile_real

                    # Execute Isolation
                    if pppoe_user:
                        is_radius = client.get('mode') == 'pppoe_radius'
                        if is_radius:
                            # PPPoE RADIUS Isolation (Address List method)
                            conn_pool = None
                            try:
                                conn_pool = get_router_connection(router_id)
                                if conn_pool:
                                    api = conn_pool.get_api()
                                    target_ip = None
                                    try:
                                        ppp_act = api.get_resource('/ppp/active').get(name=pppoe_user)
                                        if ppp_act: target_ip = ppp_act[0].get('address')
                                    except: pass
                                    
                                    if not target_ip: target_ip = client.get('ip')
                                    
                                    if target_ip and target_ip != '-':
                                        ensure_isolir_profile(router_id)
                                        res_add = add_to_address_list(target_ip, isolir_profile, router_id, comment=f"{isolir_profile}_{client['name']}")
                                        if res_add.get('status') == 'ok':
                                            kick_pppoe_user(pppoe_user, router_id)
                                            success = True
                                        else:
                                            target_skip_reason = f"Failed to add to address-list: {res_add.get('msg', 'Unknown')}"
                                    else:
                                        target_skip_reason = "User offline and no static IP found in DB"
                                else:
                                    target_skip_reason = "Could not connect to router"
                            except Exception as e:
                                target_skip_reason = f"Radius isolation error: {str(e)}"
                            finally:
                                if conn_pool: conn_pool.disconnect()
                        else:
                            # Local Secret Mode
                            ensure_isolir_profile(router_id)
                            res1 = change_pppoe_profile(pppoe_user, isolir_profile, router_id)
                            if res1.get('status') == 'ok':
                                time.sleep(1)
                                kick_pppoe_user(pppoe_user, router_id)
                                success = True
                            else:
                                # [BUG FIX] Use 'msg' instead of 'error' to match helper return
                                target_skip_reason = f"Failed to change profile: {res1.get('msg', 'Unknown')}"
                    elif client.get('ip') and client.get('ip') != '-':
                        # Static IP Mode
                        res1 = add_to_address_list(client['ip'], isolir_profile, router_id, comment=f"{isolir_profile}_{client['name']}")
                        if res1.get('status') == 'ok':
                            success = True
                        else:
                            target_skip_reason = f"Failed to add to address-list: {res1.get('msg', 'Unknown')}"
                    else:
                        target_skip_reason = "No PPPoE user or valid IP for isolation"

                    if success:
                        client['status'] = 'isolir'
                        if 'billing' not in client or not isinstance(client.get('billing'), dict):
                            client['billing'] = {}
                        client['billing']['payment_status'] = 'overdue'
                        client['billing']['isolir_date'] = now.strftime('%Y-%m-%d')
                        changed = True
                        isolir_count += 1
                        add_log(client['name'], 'isolir', f'Manual test isolation (FORCE) to profile: {isolir_profile}')
                        
                        # [NEW] Manual Force WA Notification
                        wa_enabled = billing_config.get('send_wa_notification', False)
                        phone_number = client.get('phone') or client.get('whatsapp_number') or client.get('wa_number') or client['billing'].get('wa_number')
                        if wa_enabled and phone_number:
                            msg_tpl = settings.get('wa_template_isolir', "Yth. {name}, layanan internet Anda diisolir karena tunggakan. Silakan lakukan pembayaran segera.")
                            # Dynamic Expired Date for Manual Isolir
                            manual_exp = now.strftime('%d-%m-%Y')
                            if client.get('paid_until'):
                                try:
                                    dt_me = datetime.strptime(client['paid_until'], '%Y-%m-%d')
                                    m_n_me = get_month_name(dt_me.month, pref_lang)
                                    manual_exp = f"{dt_me.day} {m_n_me} {dt_me.year}"
                                except: pass

                            wa_msg = msg_tpl.replace("{name}", client.get('name', 'Pelanggan'))\
                                           .replace("{id}", str(client.get('id', '')).replace('client_', ''))\
                                           .replace("{price}", "0")\
                                           .replace("{expired}", manual_exp)
                            wa_queue.append({"to": phone_number, "msg": wa_msg})
                            client['billing']['isolir_wa_sent'] = True

                        save_db(db_data, preserve_live=False)
                        break
                    else:
                        # Log failure to console for admin
                        add_log(client['name'], 'error', f"Gagal Isolir Manual: {target_skip_reason}")
                        break

                # BYPASS CHECK
                if client.get('bypass_billing', False):
                    if target_found: target_skip_reason = "Client has BILLING BYPASS enabled"
                    continue
                
                # NORMAL MODE: Check business rules
                # Skip if billing not enabled for client
                if 'billing' not in client or not isinstance(client['billing'], dict):
                    client['billing'] = {}
                billing = client['billing']
                
                if not billing.get('enabled', False):
                    if target_found: target_skip_reason = "Billing feature disabled for this client"
                    continue
                
                # [SELF-HEALING V3.4.2] Reset notification flag if NOT isolated
                if client.get('status') != 'isolir' and billing.get('isolir_wa_sent'):
                    billing['isolir_wa_sent'] = False
                    changed = True
                
                # NEW: Bypass Billing (Gratis Selamanya) check (V3.1.6)
                bypass_list = billing_config.get('bypass_list', [])
                if client.get('name') in bypass_list or str(client.get('id')) in bypass_list:
                    if target_found: target_skip_reason = "Client is in Bypass List (Gratis Selamanya)"
                    continue
                
                # Proactive Sync from Finance (V3.1.4)
                if sync_billing_from_finance(client, db_data):
                    changed = True
                    billing = client.get('billing', {}) # Re-fetch updated object
                
                if (not pppoe_user or pppoe_user.strip() == "") and (not client.get('ip') or client.get('ip') == '-'):
                    if target_found: target_skip_reason = "No PPPoE username or valid IP"
                    continue
                
                # NEW LOGIC: Use paid_until as primary source of truth
                paid_until_str = client.get('paid_until')
                
                due_date = None # Initialize due_date
                if paid_until_str:
                    try:
                        due_date = datetime.strptime(paid_until_str, '%Y-%m-%d')
                    except:
                        # Fallback: Pakai billing_day (di level client atau global)
                        # If paid_until_str was invalid, use billing_day logic
                        pass # Let the next block handle it if due_date is still None
                
                if not due_date: # If paid_until was not present or invalid
                    # Legacy: Pakai billing_day (di level client atau global)
                    b_day = billing.get('billing_day') or billing_config.get('default_billing_day', 5)
                    # Jika hari ini belum sampai tanggal tagihan bulan ini, berarti yang dicek adalah jatuh tempo bulan lalu
                    if now.day < b_day:
                        prev_m = now.month - 1
                        prev_y = now.year
                        if prev_m == 0: prev_m = 12; prev_y -= 1
                        due_date = calculate_due_date(prev_y, prev_m, b_day)
                    else:
                        due_date = calculate_due_date(now.year, now.month, b_day)
                
                days_overdue = (now - due_date).days
                payment_status = billing.get('payment_status', 'unpaid')
                
                processed_count += 1

                # --- PRE-CALCULATE VARIABLES FOR NOTIFICATIONS (V3.4.1) ---
                phone_number = client.get('phone') or client.get('whatsapp_number') or client.get('wa_number') or billing.get('wa_number')
                
                # Variabel Dasar (Price & Expired)
                billing_profiles = settings.get('billing_profiles', {})
                packet_name = (client.get('packet_name') or "").strip()
                
                price_val = 0
                for prof_name, prof_price in billing_profiles.items():
                    if prof_name.strip().lower() == packet_name.lower():
                        price_val = prof_price; break
                
                manual_arr_val = 0
                manual_arrears_list = settings.get('manual_arrears', [])
                for ma in manual_arrears_list:
                    if ma.get('client_name') == client['name']:
                        manual_arr_val += int(ma.get('amount', 0))

                unpaid_months_wa = 1
                if paid_until_str:
                    try:
                        dt_exp = datetime.strptime(paid_until_str, '%Y-%m-%d')
                        unpaid_months_wa = (now.year - dt_exp.year) * 12 + (now.month - dt_exp.month)
                        if unpaid_months_wa < 1: unpaid_months_wa = 1
                    except:
                        if days_overdue > 0: unpaid_months_wa = math.ceil(days_overdue / 30)
                elif days_overdue > 0:
                    unpaid_months_wa = math.ceil(days_overdue / 30)
                
                actual_package_debt = float(price_val) * unpaid_months_wa if payment_status != 'paid' else 0
                total_debt_num = actual_package_debt + manual_arr_val
                price_str = "{:,.0f}".format(total_debt_num).replace(",", ".")
                
                paid_until_val = client.get('paid_until')
                if paid_until_val and '-' in paid_until_val:
                    try:
                        dt_exp = datetime.strptime(paid_until_val, '%Y-%m-%d')
                        exp_m_name = get_month_name(dt_exp.month, pref_lang)
                        expired_date = f"{dt_exp.day} {exp_m_name} {dt_exp.year}"
                    except: expired_date = paid_until_val
                else:
                    last_day_val = calendar.monthrange(now.year, now.month)[1]
                    expired_date = f"{last_day_val} {curr_month_id} {now.year}"

                qris_path = os.path.join(SCRIPT_DIR, 'static', 'photos', 'qris.jpg')
                has_qris = os.path.exists(qris_path)
                wa_auto_qris = settings.get('wa_auto_qris', True)
                wa_manual_qris = settings.get('wa_manual_qris', True)
                wa_isolir_qris = settings.get('wa_isolir_qris', True)

                # --- NEW: WHATSAPP NOTIFICATION LOGIC ---
                wa_enabled = billing_config.get('send_wa_notification', False)

                if notify_only:
                     # Check extraction logic preview
                     _append_wa_log(f"[DEBUG] Client: {client.get('name')} | Overdue: {days_overdue} | Status: {payment_status} | Phone: {phone_number} | WA Enabled: {wa_enabled}")
                
                # Allow if enabled OR if manually triggered
                wa_auto_enabled = settings.get('wa_auto_enabled', True)
                wa_manual_enabled = settings.get('wa_manual_enabled', True)
                
                # Logic: If manual mode (button click), check wa_manual_enabled. If auto loop, check wa_auto_enabled.
                is_manual = template_mode == 'manual'
                can_send = (wa_enabled and not is_manual and wa_auto_enabled) or (is_manual and wa_manual_enabled)

                if can_send and phone_number:
                    # Logic Flags
                    wa_end_month_active = billing_config.get('wa_end_month_enabled', False)
                    wa_pre_isolir_active = billing_config.get('wa_pre_isolir_enabled', False)
                    wa_pre_isolir_days = int(billing_config.get('wa_pre_isolir_days', 2))

                    # One-time notification tracking (Month-Year period)
                    curr_period = now.strftime('%m-%Y')
                    if 'wa_sent_track' not in billing or not isinstance(billing.get('wa_sent_track'), dict):
                        billing['wa_sent_track'] = {}

                    # Helper untuk replace semua variabel
                    def format_wa_msg(tpl):
                        # Clean ID: hapus awalan 'client_' jika ada
                        raw_id = str(client.get('id', ''))
                        clean_id = raw_id.replace('client_', '')
                        
                        return tpl.replace("{name}", client.get('name', 'Pelanggan'))\
                                  .replace("{id}", clean_id)\
                                  .replace("{month}", curr_month_id)\
                                  .replace("{price}", price_str)\
                                  .replace("{amount}", price_str)\
                                  .replace("{expired}", expired_date)\
                                  .replace("{x}", str(wa_pre_isolir_days))

                    # 1. REMOVED: H-3 REMINDER (Per User Request)
                    
                    # 2. END-OF-MONTH / DUE-DATE REMINDER
                    # Allow run if active OR manual mode
                    # V3.4.1 FIX: Prevent scheduled messages from being sent during routine isolir checks (off-schedule).
                    # Only send if notify_only is True (from 9 AM scheduler) or it's triggered manually.
                    if (notify_only or template_mode == 'manual') and (wa_end_month_active or template_mode == 'manual') and (payment_status == 'unpaid' or manual_arr_val > 0) and not client.get('bypass_billing'):
                        b_mode = billing_config.get('billing_mode', 'monthly')
                        should_trigger_wa = False
                        
                        if b_mode == 'cyclic':
                            # Mode 30 Hari: Trigger pas HARI-H Jatuh Tempo
                            if days_overdue == 0:
                                should_trigger_wa = True
                        else:
                            # Mode Global: Trigger pas HARI TERAKHIR BULAN
                            last_day = calendar.monthrange(now.year, now.month)[1]
                            if now.day == last_day:
                                should_trigger_wa = True
                                
                        # FORCE TRIGGER FOR MANUAL MODE
                        if template_mode == 'manual':
                            should_trigger_wa = True

                        if should_trigger_wa:
                            # Check track STRICTLY (No manual bypass)
                            # User Request: "proteksi pasti hanya bisa kirim 1x"
                            if billing['wa_sent_track'].get('eom') != curr_period:
                                msg_tpl = settings.get('wa_template', "Halo {name}, tagihan anda belum terbayar. Mohon segera diselesaikan. Terima kasih.")
                                
                                wa_msg = format_wa_msg(msg_tpl)
                                attach_qris = has_qris and wa_manual_qris
                                wa_queue.append({"to": phone_number, "msg": wa_msg, "image": qris_path if attach_qris else None})
                                billing['wa_sent_track']['eom'] = curr_period
                                changed = True
                                if notify_only: _append_wa_log(f"[TRACE] Queued Manual/EOM for {client['name']}")
                            else:
                                if notify_only: _append_wa_log(f"[INFO] Skip EOM {client['name']}: Already sent for {curr_period}")
                        else:
                             if notify_only: _append_wa_log(f"[DEBUG] Skip EOM {client['name']}: Trigger condition not met (EOM/Cyclic/Manual)")
                    
                    # Log if skip due to global can_send
                    elif notify_only:
                        _append_wa_log(f"[DEBUG] Skip Notifications for {client['name']}: can_send=False (WA Enabled: {wa_enabled}, Manual Mode: {is_manual})")

                    # 3. PRE-ISOLATION WARNING (H-X before Grace ends)
                    # Trigger only on AUTO mode
                    # V3.4.1 FIX: Only trigger during the scheduled 9 AM notification cycle (notify_only).
                    if (notify_only) and wa_pre_isolir_active and payment_status == 'unpaid' and template_mode != 'manual':
                        trigger_day = (grace_period + 1) - wa_pre_isolir_days
                        if days_overdue == trigger_day:
                             # Check track
                            if billing['wa_sent_track'].get('pre_isolir') != curr_period:
                                msg_tpl = settings.get('wa_template_auto', "Yth. {name}, layanan internet Anda akan dinonaktifkan dalam {x} hari karena tunggakan. Mohon segera melakukan pembayaran.")
                                wa_msg = format_wa_msg(msg_tpl)
                                attach_qris = has_qris and wa_auto_qris
                                wa_queue.append({"to": phone_number, "msg": wa_msg, "image": qris_path if attach_qris else None})
                                billing['wa_sent_track']['pre_isolir'] = curr_period
                                changed = True
                
                # ISOLIR & REACTIVATION: Skip if notify_only is True
                if not notify_only:
                    # ISOLIR CONDITION: Priority to paid_until
                    should_isolate = False
                    
                    # V3.1.8 SMART ISOLIR: Also isolate if there is manual debt and it's past cycle/eom
                    manual_arr_val = 0 # Re-calculate just in case for isolation logic
                    manual_arrears_list = settings.get('manual_arrears', [])
                    for ma in manual_arrears_list:
                        if ma.get('client_name') == client['name']:
                            manual_arr_val += int(ma.get('amount', 0))
                    has_manual_debt = manual_arr_val > 0
                    
                    if paid_until_str:
                        # If paid_until exists, check if date is expired OR if it has manual debt past due
                        # V3.3.9 FIX: Don't isolate if ALREADY PAID manually
                        if payment_status == 'paid':
                            should_isolate = False
                        elif days_overdue > grace_period:
                            should_isolate = True
                        elif has_manual_debt:
                            # V3.3.4 FIX: Gunakan perhitungan kalender asli (timedelta) untuk menghindari bug "Tanggal 33"
                            try:
                                b_day = billing.get('billing_day') or billing_config.get('default_billing_day', 5)
                                # Target isolir bulan ini = tgl tagihan bulan ini + grace
                                isolir_deadline = datetime(now.year, now.month, 1) + timedelta(days=b_day - 1 + grace_period)
                                if now > isolir_deadline:
                                    # Still check status for manual debt case
                                    if payment_status != 'paid':
                                        should_isolate = True
                            except:
                                # Fallback simple addition if somehow date constructor fails
                                if now.day > (billing.get('billing_day') or 5) + grace_period and payment_status != 'paid':
                                    should_isolate = True
                    else:
                        # Fallback for legacy clients without paid_until
                        if (days_overdue > grace_period and payment_status == 'unpaid') or has_manual_debt:
                            should_isolate = True

                    if should_isolate and auto_isolir:
                        # Only isolir if not already isolated
                        if client.get('status') != 'isolir':
                            # DEFER ISOLATION: Add to deferred list to execute after WA notifications
                            deferred_isolations.append({
                                'client_idx': idx,
                                'client_name': client['name'],
                                'router_id': client.get('managed_by', 'server_utama'),
                                'pppoe_user': pppoe_user,
                                'ip_addr': client.get('ip'),
                                'days_overdue': days_overdue,
                                'packet_name': client.get('packet_name', 'default'),
                                'isolir_profile': isolir_profile  # <--- FIX: Simpan profil spesifik router saat dimasukkan
                            })

                            # QUEUE ISOLATION NOTIFICATION NOW
                            wa_isolir_enabled = settings.get('wa_isolir_enabled', True)
                            if wa_enabled and wa_isolir_enabled and phone_number:
                                # Only send isolation WA once per arrear period
                                if not billing.get('isolir_wa_sent'):
                                    msg_tpl = settings.get('wa_template_isolir', "Yth. {name}, layanan internet Anda diisolir karena tunggakan sebesar Rp {price}. Silakan lakukan pembayaran segera.")
                                    
                                    # Use pre-calculated price_str and expired_date
                                    wa_msg = msg_tpl.replace("{name}", client.get('name', 'Pelanggan'))\
                                                   .replace("{id}", str(client.get('id', '')).replace('client_', ''))\
                                                   .replace("{price}", price_str)\
                                                   .replace("{amount}", price_str)\
                                                   .replace("{expired}", expired_date)
                                    
                                    attach_qris = has_qris and wa_isolir_qris
                                    wa_queue.append({"to": phone_number, "msg": wa_msg, "image": qris_path if attach_qris else None})
                                    billing['wa_sent_track']['isolir'] = curr_period # Use 'isolir' key for tracking
                                    billing['isolir_wa_sent'] = True
                                    changed = True
                                else:
                                    if notify_only: _append_wa_log(f"[INFO] Skip Isolir WA {client['name']}: Already sent (isolir_wa_sent=True)")
                            elif notify_only:
                                _append_wa_log(f"[DEBUG] Skip Isolir WA {client['name']}: wa_isolir_enabled={settings.get('wa_isolir_enabled', True)}, wa_enabled={wa_enabled}, phone={phone_number}")
                    
                    # REACTIVATION CONDITION: Previously isolated but now paid
                    elif client.get('status') == 'isolir' and payment_status == 'paid':
                        router_id = client.get('managed_by', 'server_utama')
                        success = False
                        
                        if pppoe_user:
                            is_radius = client.get('mode') == 'pppoe_radius'
                            if is_radius:
                                # PPPoE RADIUS Reactivation (Remove from Address List)
                                conn_ra = None
                                try:
                                    conn_ra = get_router_connection(router_id)
                                    if conn_ra:
                                        api = conn_ra.get_api()
                                        fw_list = api.get_resource('/ip/firewall/address-list')
                                        entries = fw_list.get(list=isolir_profile, comment=f"{isolir_profile}_{client['name']}")
                                        for e in entries: fw_list.remove(id=e['id'])
                                except: pass
                                finally:
                                    if conn_ra: conn_ra.disconnect()
                                kick_pppoe_user(pppoe_user, router_id)
                                success = True
                            else:
                                # Local Secret Reactivation
                                current_profile = billing.get('original_profile') or client.get('packet_name', 'default')
                                res = change_pppoe_profile(pppoe_user, current_profile, router_id)
                                if res.get('status') == 'ok':
                                    time.sleep(1)
                                    kick_pppoe_user(pppoe_user, router_id)
                                    success = True
                        elif client.get('ip') and client.get('ip') != '-':
                            res = remove_from_address_list(client['ip'], isolir_profile, router_id, comment=f"{isolir_profile}_{client['name']}")
                            if res.get('status') == 'ok':
                                success = True
                        
                        if success:
                            client['status'] = 'online'
                            billing['isolir_wa_sent'] = False # Reset flag for next arrear period (V3.1.8)
                            changed = True
                            reactivate_count += 1
                            add_log(client['name'], 'online', 'Auto-reactivation after payment')

                            # 3. REACTIVATION NOTIFICATION (AUTO)
                            wa_react_enabled = settings.get('wa_reactivate_enabled', True)
                            if wa_enabled and wa_react_enabled and phone_number:
                                msg_tpl = settings.get('wa_template_reactivate', "Halo {name}, pembayaran telah diterima dan layanan internet Anda telah diaktifkan kembali. Terima kasih.")
                                wa_msg = format_wa_msg(msg_tpl)
                                wa_queue.append({"to": phone_number, "msg": wa_msg, "image": None})
                    else:
                        if target_found: 
                            if payment_status == 'paid': target_skip_reason = "Status is PAID"
                            elif days_overdue <= grace_period: target_skip_reason = f"Not overdue yet ({days_overdue} days overdue, grace {grace_period})"
                            else: target_skip_reason = "Conditions not met"

            except Exception as e:
                print(f"[BILLING] Error processing client {client.get('name')}: {e}")
                if target_found: target_skip_reason = f"Error: {str(e)}"
                continue

        # --- AFTER MAIN LOOP: SEND NOTIFICATIONS & EXECUTE DEFERRED ISOLATIONS ---

        # 1. Trigger WhatsApp Batch first
        if wa_queue:
            try:
                import tempfile, json
                with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, dir=SCRIPT_DIR) as tf:
                    json.dump(wa_queue, tf); tf.flush()
                    temp_path = tf.name
                # Ensure file is closed before spawning worker
                tf.close() 
                spawn_wa_worker(mode="batch", task_file=temp_path)
                
                # [CORE FIX]: Delay isolation by 10 seconds to allow WA delivery while internet IS STILL ON
                if deferred_isolations:
                    time.sleep(10)
            except Exception as e:
                pass

        # 2. Process Deferred MikroTik Isolation Commands
        if deferred_isolations:
            for item in deferred_isolations:
                client = db_data['clients'][item['client_idx']]
                pppoe_user = item['pppoe_user']
                router_id = item['router_id']
                ip_addr = item['ip_addr']
                client_name = item['client_name']
                days_overdue = item['days_overdue']
                item_isolir_profile = item.get('isolir_profile', 'ISOLIR')
                
                success = False
                if pppoe_user:
                    # PPPoE / RADIUS Isolation
                    is_radius = client.get('mode') == 'pppoe_radius'
                    if is_radius:
                        # Address List method for Radius
                        conn = None
                        try:
                            conn = get_router_connection(router_id)
                            if conn:
                                api = conn.get_api()
                                # Cleanup & Re-add logic
                                try:
                                    fw_list = api.get_resource('/ip/firewall/address-list')
                                    old = fw_list.get(list=item_isolir_profile, comment=f"{item_isolir_profile}_{client_name}")
                                    for o in old: fw_list.remove(id=o['id'])
                                except: pass
                                
                                target_ip = None
                                ppp_act = api.get_resource('/ppp/active').get(name=pppoe_user)
                                if ppp_act: target_ip = ppp_act[0].get('address')
                                
                                if not target_ip: target_ip = ip_addr
                                if target_ip and target_ip != '-':
                                    res_add = add_to_address_list(target_ip, item_isolir_profile, router_id, comment=f"{item_isolir_profile}_{client_name}")
                                    if res_add.get('status') == 'ok':
                                        kick_pppoe_user(pppoe_user, router_id)
                                        success = True
                                else:
                                    # User Offline? Mark success anyway so db updates status to ISOLIR (Parity with Original)
                                    success = True
                        except: pass
                        finally:
                            if conn: conn.disconnect()
                    else:
                        # Local Secret Mode
                        ensure_isolir_profile(router_id)
                        
                        # [BUG FIX]: Get real current profile from Mikrotik, fallback to DB packet_name, then 'default'
                        real_prof_mk = get_pppoe_current_profile(pppoe_user, router_id)
                        cur_prof = real_prof_mk if (real_prof_mk and real_prof_mk != "") else (item.get('packet_name') or 'default')
                        
                        res1 = change_pppoe_profile(pppoe_user, item_isolir_profile, router_id)
                        if res1.get('status') == 'ok':
                            time.sleep(1)
                            kick_pppoe_user(pppoe_user, router_id)
                            if 'billing' not in client: client['billing'] = {}
                            
                            # Only overwrite original_profile if the current one isn't ALREADY isolir
                            if item_isolir_profile.upper() not in cur_prof.upper():
                                client['billing']['original_profile'] = cur_prof
                                
                            success = True
                elif ip_addr and ip_addr != '-':
                    # Static IP Mode
                    res1 = add_to_address_list(ip_addr, item_isolir_profile, router_id, comment=f"{item_isolir_profile}_{client_name}")
                    if res1.get('status') == 'ok': success = True
                
                if success:
                    client['status'] = 'isolir'
                    if 'billing' not in client: client['billing'] = {}
                    client['billing']['payment_status'] = 'overdue'
                    client['billing']['isolir_date'] = now.strftime('%Y-%m-%d')
                    changed = True
                    isolir_count += 1
                    add_log(client_name, 'isolir', f"Auto-isolir (Scheduled): Tunggakan {days_overdue} hari")
                    add_log("SYSTEM", "system", f"Isolir Otomatis: {client_name} (Tunggakan {days_overdue} hari)")

        # Summarize results
        if changed:
            # FIX: preserve_live MUST be False so we can push our isolir modifications into the DB!
            save_db(db_data, preserve_live=False)
            
        # HEARTBEAT LOG
        wa_sent_count = len(wa_queue) if wa_queue else 0
        # Summary log (Heartbeat) - Always show in V3.4.1
        g_wa_en = billing_config.get('send_wa_notification', False)
        summary_msg = f"Billing Heartbeat: {processed_count} checked, {isolir_count} isolated, {reactivate_count} active, {len(wa_queue)} notifications (WA: {g_wa_en})."
        add_log("SYSTEM_BILLING", "system", summary_msg)

        if target_user:
            if target_found:
                return f"Target {target_user} scan finished. Result: {target_skip_reason or 'Processed'}"
            else:
                return f"Target user '{target_user}' not found in DB"
        
        return f"Routine check finished. {processed_count} checked, {isolir_count} isolated, {reactivate_count} active."
    except Exception as e:
        print(f"[ERROR] Billing check error: {e}")
        return f"Error: {e}"

# --- WHATSAPP BRIDGE LOGIC ---
WA_LOG_FILE = os.path.join(SCRIPT_DIR, 'wa_logs.json')
active_wa_workers = [] # Track running processes

def spawn_wa_worker(mode="test", target="", message="", task_file=None, image=None):
    """
    Menjalankan Node.js bridge untuk mengirim pesan WhatsApp.
    Mode: 'test' (single) atau 'batch' (multiple via file).
    """
    try:
        bridge_path = os.path.join(SCRIPT_DIR, 'wa-bridge.js')
        nm_path = os.path.join(SCRIPT_DIR, 'node_modules')
        if not os.path.exists(bridge_path):
            return False, "Bridge file missing"

        # Log Aktivitas ke WA_LOG_FILE
        now_ts = get_local_now().strftime('%Y-%m-%d %H:%M:%S')
        
        # --- DIAGNOSTIK LINUX/SERVER ---
        import shutil
        node_exe = shutil.which("node") or shutil.which("nodejs")
        
        if not node_exe:
            # Coba cari di path umum linux/armbian jika tidak ada di lingkungan PATH
            common_paths = [
                "/usr/bin/node", "/usr/local/bin/node", "/bin/node", 
                "/usr/bin/nodejs", "/usr/local/bin/nodejs",
                "/opt/node/bin/node", "/snap/bin/node"
            ]
            for p in common_paths:
                if os.path.exists(p):
                    node_exe = p
                    break
        
        if not node_exe:
            # Debug: Catat PATH yang terlihat oleh Python saat ini
            env_path = os.environ.get('PATH', 'Tidak Ditemukan')
            _append_wa_log(f"[{now_ts}] ERROR: Node.js tidak ditemukan. Sistem PATH: {env_path}")
            return False, "Node.js not found"
            
        if not os.path.exists(nm_path):
            _append_wa_log(f"[{now_ts}] ERROR: Folder 'node_modules' tidak ada. Jalankan 'npm install' di folder app.")
            return False, "node_modules missing. Run npm install."
        # -------------------------------

        log_entry = f"[{now_ts}] START: Mode={mode}"
        if task_file: log_entry += f", File={os.path.basename(task_file)}"
        else: log_entry += f", To={target}"
        if image: log_entry += f", Image={os.path.basename(image)}"
        _append_wa_log(log_entry)

        # Jalankan Command: [node] wa-bridge.js --mode x --to y --msg z [--file f] [--image i]
        cmd = [node_exe, bridge_path, "--mode", mode]
        if task_file:
            cmd.extend(["--file", task_file])
        else:
            cmd.extend(["--to", target, "--msg", message])
            if image:
                cmd.extend(["--image", image])
        
        # Jalankan di background
        import subprocess
        try:
            process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, cwd=SCRIPT_DIR)
            active_wa_workers.append(process)
            
            # Cek status dalam 2 detik pertama (Deteksi Crash Startup)
            time.sleep(2)
            if process.poll() is not None:
                stdout, stderr = process.communicate()
                _append_wa_log(f"[{now_ts}] CRITICAL: Script berhenti mendadak ({process.returncode}). Error: {stderr or stdout}")
                return False, f"Startup Error: {stderr or stdout}"
            
            # Jika masih jalan, biarkan dia jalan di background
            def monitor_output(p):
                # Tangkap STDOUT
                for line in p.stdout:
                    if not line: break
                    msg = line.strip()
                    # Capture all standard bridge tags
                    tags = ["[STATUS]", "[SUCCESS]", "[ERROR]", "[FAILED]", "[SEND]", "[QR]", "[WAIT]"]
                    if any(t in msg for t in tags):
                        _append_wa_log(f"[{get_local_now().strftime('%H:%M:%S')}] {msg}")
                
                # Tangkap STDERR
                for line in p.stderr:
                    if not line: break
                    _append_wa_log(f"[{get_local_now().strftime('%H:%M:%S')}] node-err: {line.strip()}")
                
                return_code = p.wait()
                _append_wa_log(f"[{get_local_now().strftime('%H:%M:%S')}] SHUTDOWN: Service dimatikan (Code: {return_code})")
            
            threading.Thread(target=monitor_output, args=(process,), daemon=True).start()
            
            _append_wa_log(f"[{now_ts}] INFO: Service berhasil dibangunkan (PID: {process.pid})")
            return True, "Worker started"
            
        except Exception as e:
            _append_wa_log(f"[{now_ts}] EXCEPTION: {str(e)}")
            return False, str(e)
    except Exception as e:
        return False, str(e)

def _append_wa_log(text):
    """Simpan log khusus WhatsApp ke wa_logs.json"""
    try:
        logs = []
        if os.path.exists(WA_LOG_FILE):
            with open(WA_LOG_FILE, 'r') as f: logs = json.load(f)
        
        logs.append(text)
        if len(logs) > 50: logs = logs[-50:] # Simpan 50 log terakhir
        
        with open(WA_LOG_FILE, 'w') as f: json.dump(logs, f, indent=2)
    except: pass

@app.route('/api/logs/wa')
def get_wa_logs_route():
    if not check_auth(request): return jsonify({"error":"Unauthorized"}), 401
    try:
        if os.path.exists(WA_LOG_FILE):
            with open(WA_LOG_FILE, 'r') as f: return jsonify(json.load(f))
    except: pass
    return jsonify([])

@app.route('/api/whatsapp/test', methods=['POST'])
def api_wa_test():
    """Trigger tes kirim WA instan"""
    data = request.json or {}
    target = data.get('target', '').strip()
    client_name = data.get('name', 'Pelanggan')
    t_type = data.get('type', 'manual') # 'manual' or 'auto'
    
    settings = load_settings()
    now = get_local_now()
    pref_lang = settings.get('language', 'id')
    curr_month_id = get_month_name(now.month, pref_lang)

    if t_type == 'auto':
        template = settings.get('wa_template_auto', "Halo {name}, (Auto Response)")
    else:
        template = settings.get('wa_template', "Halo {name}, ini adalah pesan tagihan/tes dari NMS.")
        
    # Tambahkan Tunggakan Manual jika ada
    manual_arr_val = 0
    manual_arrears_list = settings.get('manual_arrears', [])
    for ma in manual_arrears_list:
        if ma.get('client_name') == client_name:
            manual_arr_val += int(ma.get('amount', 0))

    # Calculate price based on bill passed or fallback
    try:
        bill_val = data.get('bill')
        if bill_val is not None:
            # Use bill directly if passed from frontend
            price_num = float(bill_val)
        else:
            # Fallback for generic test
            price_num = 0 + manual_arr_val
        
        price_str = "{:,.0f}".format(price_num).replace(",", ".")
    except:
        price_str = "0"

    # Clean ID
    raw_id = str(data.get('id', ''))
    if not raw_id:
        # Cari ID dari DB jika tidak dikirim dari FE
        db = load_db()
        for c in db.get('clients', []):
            if c.get('name') == client_name:
                raw_id = str(c.get('id', ''))
                break
    clean_id = raw_id.replace('client_', '')

    # Smart Expired Date for Test
    test_expired = f"28 {curr_month_id} {now.year}" # Default EOM Feb
    db_test = load_db()
    for c in db_test.get('clients', []):
        if c.get('name') == client_name:
            p_until = c.get('paid_until')
            if p_until and '-' in p_until:
                try:
                    dt_exp = datetime.strptime(p_until, '%Y-%m-%d')
                    m_name = get_month_name(dt_exp.month, pref_lang)
                    test_expired = f"{dt_exp.day} {m_name} {dt_exp.year}"
                except: pass
            break

    message = template.replace("{name}", client_name)\
                      .replace("{id}", clean_id)\
                      .replace("{month}", curr_month_id)\
                      .replace("{price}", price_str)\
                      .replace("{amount}", price_str)\
                      .replace("{expired}", test_expired)\
                      .replace("{x}", "2")
    
    # Path QRIS (Jika ada)
    qris_path = os.path.join(SCRIPT_DIR, 'static', 'photos', 'qris.jpg')
    has_qris = os.path.exists(qris_path)
    
    # Check if QRIS should be attached for this type
    wa_auto_qris = settings.get('wa_auto_qris', True)
    wa_manual_qris = settings.get('wa_manual_qris', True)
    attach_qris = False
    if t_type == 'auto':
        attach_qris = has_qris and wa_auto_qris
    else:
        attach_qris = has_qris and wa_manual_qris

    if not target: return jsonify({"status": "error", "msg": "Nomor tujuan harus diisi"}), 400
    
    ok, msg = spawn_wa_worker(mode="test", target=target, message=message, image=qris_path if attach_qris else None)
    if ok: return jsonify({"status": "ok", "msg": "Worker dijalankan. Pantau log di halaman Settings -> WhatsApp (Log Aktivitas di bagian bawah)."})
    else: return jsonify({"status": "error", "msg": msg})

@app.route('/api/whatsapp/init', methods=['POST'])
def api_wa_init():
    """Bangunkan worker WA untuk pairing atau sinkronisasi status"""
    ok, msg = spawn_wa_worker(mode="test", target="", message="") # Empty target = init mode
    if ok: return jsonify({"status": "ok", "msg": "Menghubungkan ke WhatsApp Service... Mohon tunggu beberapa detik."})
    else: return jsonify({"status": "error", "msg": msg})

@app.route('/api/whatsapp/logs')
def api_wa_logs():
    """Ambil logs khusus WhatsApp"""
    if os.path.exists(WA_LOG_FILE):
        with open(WA_LOG_FILE, 'r') as f: return jsonify(json.load(f))
    return jsonify([])

@app.route('/api/whatsapp/logs/clear', methods=['POST'])
def api_wa_logs_clear():
    """Hapus log aktivitas WhatsApp"""
    try:
        if os.path.exists(WA_LOG_FILE):
            os.remove(WA_LOG_FILE)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)})

@app.route('/api/whatsapp/reset', methods=['POST'])
def api_wa_reset():
    """Hapus sesi WhatsApp"""
    global active_wa_workers
    try:
        # 1. Matikan semua worker yang masih jalan (biar tidak ada zombie)
        for p in active_wa_workers:
            if p.poll() is None:
                try:
                    if platform.system().lower() == 'windows':
                        subprocess.run(['taskkill', '/F', '/T', '/PID', str(p.pid)], capture_output=True)
                    else:
                        p.terminate()
                        p.wait(timeout=2)
                except:
                    try: p.kill()
                    except: pass
        active_wa_workers = []
        
        # 2. Hapus sisa file QR (biar UI bersih)
        qr_file = os.path.join(SCRIPT_DIR, 'wa_qr.txt')
        if os.path.exists(qr_file): os.remove(qr_file)

        # 3. Hapus folder sesi
        session_dir = os.path.join(SCRIPT_DIR, 'wa_session')
        import shutil
        if os.path.exists(session_dir): shutil.rmtree(session_dir)
        
        _append_wa_log(f"[{get_local_now().strftime('%H:%M:%S')}] SYSTEM: Sesi dan proses direset total. Silakan scan ulang.")
        return jsonify({"status": "ok", "msg": "Sesi WA berhasil dihapus dan proses dimatikan."})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)})

@app.route('/api/whatsapp/qr')
def api_wa_qr_status():
    """Cek status koneksi dan ambil QR jika ada"""
    qr_file = os.path.join(SCRIPT_DIR, 'wa_qr.txt')
    session_file = os.path.join(SCRIPT_DIR, 'wa_session', 'creds.json')
    
    # 1. Cek Sesi Utama
    if os.path.exists(session_file):
        return jsonify({"status": "connected", "qr": None})

    # 2. Cek apakah ada worker yang sedang 'pairing'
    is_pairing = False
    global active_wa_workers
    active_wa_workers = [p for p in active_wa_workers if p.poll() is None]
    if active_wa_workers:
        is_pairing = True

    # 3. Ambil data QR jika ada file-nya
    qr_data = None
    if os.path.exists(qr_file):
        try:
            with open(qr_file, 'r') as f:
                qr_data = f.read().strip()
                # Jika ada QR tapi worker mati, berarti worker itu zombie/lama
                if not is_pairing:
                     try: os.remove(qr_file)
                     except: pass
                     qr_data = None
        except: pass
    
    status = "pairing" if (is_pairing or qr_data) else "disconnected"
    return jsonify({"status": status, "qr": qr_data})


LAST_BILLING_NOTIF_DATE = None
GLOBAL_BILLING_HEARTBEAT = "Menunggu Sinkronisasi..."
LAST_BILLING_CHECK_TS = 0  # unix timestamp
LAST_WA_CHECK_TS = 0       # unix timestamp

def auto_billing_loop():
    """Auto-isolir clients based on configurable interval or specific time"""
    global GLOBAL_BILLING_HEARTBEAT
    time.sleep(15)  # Wait for system init
    add_log("SYSTEM", "system", "Background Thread: Auto-Billing Loop Started")
    
    last_check_time = None
    while True:
        try:
            now = get_local_now()
            GLOBAL_BILLING_HEARTBEAT = now.strftime('%d %b %Y, %H:%M:%S')
            
            # Load interval configuration (in hours)
            b_cfg = load_billing_config()
            interval_hours = int(b_cfg.get('billing_check_interval_hours', 24))
            
            # 1. Logic Interval (Routine Isolation)
            should_run_interval = False
            if last_check_time is None:
                should_run_interval = True
            else:
                elapsed_seconds = (now - last_check_time).total_seconds()
                interval_seconds = interval_hours * 3600
                if elapsed_seconds >= interval_seconds:
                    should_run_interval = True
            
            # 2. Logic Specific Time (Notifications) - V3.3.4 PROFESSIONAL LOGIC
            wa_time_str = b_cfg.get('wa_notif_time', '09:00')
            try:
                wa_time_clean = wa_time_str.lower().replace('am', '').replace('pm', '').strip()
                t_hour, t_min = map(int, wa_time_clean.split(':'))
                if 'pm' in wa_time_str.lower() and t_hour < 12: t_hour += 12
                if 'am' in wa_time_str.lower() and t_hour == 12: t_hour = 0
                
                sched_today = now.replace(hour=t_hour, minute=t_min, second=0, microsecond=0)
                today_str = now.strftime('%Y-%m-%d')
                
                if now >= sched_today:
                    _sys_sett = _load_settings_raw()
                    _last_run_date = _sys_sett.get('last_auto_billing_date')
                    
                    if _last_run_date != today_str:
                        now_ts = now.strftime('%H:%M:%S')
                        _append_wa_log(f"[{now_ts}] INFO: Mengecek jadwal kirim hari ini ({today_str})...")
                        _append_wa_log(f"[{now_ts}] INFO: Menemukan jadwal {wa_time_str}. Membangunkan gateway WA...")
                        
                        # V3.3.4: Lock state BEFORE heavy work to strictly prevent re-entry/looping
                        # V3.4.1 FIX: Always fetch FRESH settings before stamping to avoid Lost Update
                        _fresh_sett = _load_settings_raw()
                        _fresh_sett['last_auto_billing_date'] = today_str
                        _save_settings_raw(_fresh_sett)
                        
                        # Executing
                        global LAST_WA_CHECK_TS
                        LAST_WA_CHECK_TS = int(now.timestamp())
                        run_billing_check(notify_only=True)
                        
                        now_ts = get_local_now().strftime('%H:%M:%S')
                        _append_wa_log(f"[{now_ts}] INFO: Tugas selesai. Mematikan gateway WA (Standby).")
                    else:
                        # Already done today. Silently standby.
                        pass
            except Exception as e:
                print(f"[BILLING AUTO] Time parse error '{wa_time_str}': {e}")

            if should_run_interval:
                global LAST_BILLING_CHECK_TS
                LAST_BILLING_CHECK_TS = int(now.timestamp())
                run_billing_check(notify_only=False) # Routine Isolation (No WA log)
                last_check_time = now

        except Exception as e:
            add_log("SYSTEM", "error", f"Billing Loop Error: {str(e)}")
            print(f"[BILLING LOOP ERROR] {e}")
            
        time.sleep(60)  # Check every 1 minute

# --- START BACKGROUND THREADS (GLOBAL SCOPE FOR V3.3.8 PRODUCTION) ---
t1 = threading.Thread(target=turbo_ping_loop, daemon=True); t1.start()
t2 = threading.Thread(target=monitor_mikrotik_loop, daemon=True); t2.start()
t3 = threading.Thread(target=auto_backup_loop, daemon=True); t3.start()
t4 = threading.Thread(target=auto_billing_loop, daemon=True); t4.start()

# --- AUTH CHECK ---
def check_auth(req):
    reload_config_globals()
    t = req.headers.get('X-Auth-Token')
    if not t: t = req.args.get('token') # Support query param for downloads
    if not t: t = req.values.get('token') # Fallback to form values
    if t: t = t.strip() # Remove whitespace
    
    if t == ADMIN_PASSWORD: return "admin"
    if t == PASSWORD_VIEWER: return "viewer"
    return None

@app.route('/api/debug/run_billing', methods=['POST'])
def api_debug_run_billing():
    auth = check_auth(request)
    if not auth: return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.json or {}
    target_user = data.get('user') or data.get('target_user')
    force = data.get('force', False)
    notify_only = data.get('notify_only', False)
    
    # [NEW] Pre-validation: Check if user exists in DB
    if target_user:
        db_v = load_db()
        found_in_db = False
        t_user_low = target_user.strip().lower()
        for c in db_v.get('clients', []):
            pu = c.get('credentials', {}).get('pppoe_user')
            pi = c.get('ip')
            cn = c.get('name')
            
            match_pu = (pu and pu.strip().lower() == t_user_low)
            match_pi = (pi and pi.strip() == target_user.strip())
            match_cn = (cn and cn.strip().lower() == t_user_low)
            
            if match_pu or match_pi or match_cn:
                found_in_db = True; break
                
        if not found_in_db:
            return jsonify({'status': 'error', 'message': f"User/IP/Nama '{target_user}' tidak ditemukan di database."})
    
    # Run in background to avoid timeout
    threading.Thread(target=run_billing_check, kwargs={
        'target_user': target_user,
        'force': force,
        'notify_only': notify_only
    }).start()
    
    msg = "Proses penagihan (Notifikasi & Cek Isolir) berjalan di background."
    if notify_only:
        msg = "Proses Notifikasi WA berjalan di background (Tanpa Isolir)."
        
    return jsonify({'status': 'ok', 'message': msg})
@app.route('/api/version_check')
def api_version_check():
    import hashlib
    try:
        with open(__file__, 'rb') as f:
            md5 = hashlib.md5(f.read()).hexdigest()
    except: md5 = "error"
    return jsonify({
        "status": "active",
        "version": "V3.1 STABLE CLEAN",
        "file_path": os.path.abspath(__file__),
        "is_licensed": is_licensed(),
        "file_md5": md5
    })

@app.route('/')
def dashboard(): return render_template('dashboard.html')

@app.route('/maps')
def maps(): return render_template('index.html')

@app.route('/client')
def client_page(): return render_template('client.html')

@app.route('/hotspot')
def page_hotspot(): return render_template('hotspot.html')

@app.route('/pppoe')
def page_pppoe(): return render_template('pppoe.html')

@app.route('/billing')
def page_billing(): return render_template('billing.html')

@app.route('/network')
def page_network(): return render_template('network.html')

@app.route('/monitor')
def page_monitor(): return render_template('monitor.html')

@app.route('/finance')
def finance_page(): return render_template('finance.html')

@app.route('/about')
def page_about(): return render_template('about.html')

@app.route('/settings')
def settings_page(): return render_template('settings.html')

@app.route('/api/login', methods=['POST'])
def login():
    p = request.json.get('password')
    if p == ADMIN_PASSWORD: return jsonify({"status":"ok", "role":"admin"})
    if p == PASSWORD_VIEWER: return jsonify({"status":"ok", "role":"viewer"})
    return jsonify({"error":"Wrong password"}), 401

def get_system_uptime():
    try:
        boot_time = psutil.boot_time()
        uptime_seconds = time.time() - boot_time
        days = int(uptime_seconds // (24 * 3600))
        hours = int((uptime_seconds % (24 * 3600)) // 3600)
        minutes = int((uptime_seconds % 3600) // 60)
        if days > 0:
            return f"{days}d {hours}h {minutes}m"
        return f"{hours}h {minutes}m"
    except:
        return "-"

@app.route('/api/data')
def get_data_route():
    role = check_auth(request)
    if not role: return jsonify({"error":"Unauthorized"}), 401
    topo = load_db()
    
    # Inject Live Data for ALL Routers (V3.3.8 FIX)
    def _inject_live(node, r_id):
        if not node: return
        if r_id in MK_RES:
            res = MK_RES[r_id]
            if not res.get('error'):
                node['status'] = 'online'
                node['_detected_wan'] = {'rx': res.get('wan_rx','0'), 'tx': res.get('wan_tx','0'), 'name': res.get('wan_name')}
                node['_detected_ports'] = {'lan': res.get('port_lan',0), 'sfp': res.get('port_sfp',0)}
            else:
                node['status'] = 'offline'
        else:
            # If no sync yet, default to offline for better UI feedback
            node['status'] = 'offline'

    _inject_live(topo.get('server'), "server_utama")
    for rtr in topo.get('extra_routers', []):
        _inject_live(rtr, rtr.get('id'))
    # License Info
    lic_info = {"status": "Unlicensed", "client": "-", "type": "FREE", "active": False}
    if is_licensed():
        try:
            with open(LICENSE_FILE, 'r') as f:
                k = f.read().strip()
            v, info = verify_license(k)
            if v:
                lic_info = {"status": "Active", "client": info.get('cli','-'), "type": "LICENSED (PRO)", "active": True}
        except: pass

    settings_data = load_settings()
    
    return jsonify({
        "topology": topo,
        "settings": settings_data,
        "system": get_system_stats_cached(), 
        "mikrotik_data": MK_RES, 
        "role": role,
        "license": lic_info
    })

def get_system_stats_cached():
    """Optimasi V3.3.9: Cache psutil stats for 5 seconds to reduce CPU load."""
    global _SYSTEM_STATS_CACHE
    now = time.time()
    if _SYSTEM_STATS_CACHE['expiry'] > now:
        return _SYSTEM_STATS_CACHE['data']
    
    try:
        stats = {
            "cpu": psutil.cpu_percent(), 
            "ram": psutil.virtual_memory().percent, 
            "disk": psutil.disk_usage('/').percent, 
            "temp": get_cpu_temp(),
            "uptime": get_system_uptime()
        }
        _SYSTEM_STATS_CACHE = {"data": stats, "expiry": now + 5}
        return stats
    except:
        return {}

@app.route('/api/save', methods=['POST'])
def save_route():
    if check_auth(request) != "admin": 
        return jsonify({"error":"Forbidden"}), 403
        
    incoming = request.json
    if not incoming: return jsonify({"status": "error", "msg": "No data received"}), 400
    
    # NEW: Auto-enable billing for new clients in incoming data if not specified
    if isinstance(incoming, dict) and 'clients' in incoming:
        for c in incoming['clients']:
            if 'billing' not in c:
                # Default for new clients: Enabled
                c['billing'] = {'enabled': True}
    
    # save_db now handles merging status internally
    save_db(incoming, preserve_live=True)

    # [FIX] Auto-Refresh: Trigger immediate MikroTik status fetch in background
    # so the UI shows login status right away without waiting for the 5s loop.
    def _trigger_immediate_fetch():
        try:
            db_snap = load_db()
            srv = db_snap.get("server", {})
            if srv.get("login", {}).get("host"):
                fetch_single_router_data("server_utama", srv["login"], srv)
            for rtr in db_snap.get("extra_routers", []):
                if rtr.get("login", {}).get("host"):
                    fetch_single_router_data(rtr["id"], rtr["login"], rtr)
        except Exception as e:
            print(f"[SAVE] Immediate fetch error: {e}")

    t = threading.Thread(target=_trigger_immediate_fetch, daemon=True)
    t.start()

    return jsonify({"status":"ok", "msg": "Data saved (Live status preserved)"})

@app.route('/api/clients/import/template')
def api_clients_import_template():
    if not check_auth(request): return jsonify({"error":"Forbidden"}), 403
    if not OPENPYXL_READY:
        return jsonify({"status": "error", "msg": "openpyxl not installed on server"}), 500
    
    tpl_path = os.path.join(SCRIPT_DIR, 'template_import_client.xlsx')
    
    # Always regenerate to include new instructions
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Import Client'
    
    # Add Instructions
    instructions = [
        ['💡 PETUNJUK PENGISIAN TEMPLATE IMPORT CLIENT'],
        ['1. Tipe Koneksi: Isi dengan "PPPoE", "PPPoE Radius", atau "Statik"'],
        ['2. Username PPPoE: Wajib diisi untuk tipe PPPoE & PPPoE Radius (Samakan dengan secret di Mikrotik)'],
        ['3. IP Address: Wajib diisi untuk tipe Statik'],
        ['4. Nama Paket / Profile: Untuk PPPoE Radius & Statik isi manual, untuk PPPoE standar akan otomatis sync dari Mikrotik jika kosong'],
        ['5. Koordinat: Format harus "lat, lng" (Contoh: -7.123, 110.456) agar muncul di peta'],
        ['6. Billing: Isi "On" untuk aktifkan isolir otomatis, atau "Off" untuk nonaktifkan'],
        ['7. Induk: Isi dengan ID atau Nama ODP (Contoh: ODP-A-01). Kosongkan jika ingin menempel ke Router'],
        [''], # Blank row before headers
    ]
    for row in instructions:
        ws.append(row)
        
    # Headers
    headers = [
        'Pengelolah Mikrotik', 'Nama Client', 'Tipe Koneksi', 
        'Username PPPoE', 'IP Address', 'Nama Paket / Profile', 
        'WhatsApp', 'Koordinat', 'Billing', 'Induk'
    ]
    ws.append(headers)
    
    # Sample Rows
    ws.append(['server_utama', 'Contoh PPPoE', 'PPPoE', 'user_contoh', '-', 'Paket_10M', '628123456789', '-7.123, 110.456', 'On', ''])
    ws.append(['server_utama', 'Contoh Radius', 'PPPoE Radius', 'user_radius', '-', 'Paket_20M', '628123456789', '-7.124, 110.457', 'On', ''])
    ws.append(['server_utama', 'Contoh Statik', 'Statik', '-', '192.168.1.100', 'Paket_Statik', '628123456789', '-7.125, 110.458', 'Off', ''])
    
    # Simple Styling for Headers (Internal row index is len(instructions) + 1)
    header_row_idx = len(instructions) + 1
    for cell in ws[header_row_idx]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    wb.save(tpl_path)
    return send_file(tpl_path, as_attachment=True, download_name='template_import_client.xlsx')

@app.route('/api/clients/import', methods=['POST'])
def api_clients_import():
    if check_auth(request) != "admin": 
        return jsonify({"error":"Forbidden"}), 403
        
    if 'file' not in request.files:
        return jsonify({"status": "error", "msg": "No file uploaded"}), 400
        
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({"status": "error", "msg": "No selected file"}), 400

    if not openpyxl:
        return jsonify({"status": "error", "msg": "openpyxl not installed on server"}), 500

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        # Load Existing Data
        data = load_db()
        topology = data # in app.py load_db() returns the dict
        clients = topology.get('clients', [])
        routers = [{'id': 'server_utama', 'name': topology.get('server', {}).get('name', ''), 'identity': topology.get('server', {}).get('identity', '')}]
        for r in topology.get('extra_routers', []):
            routers.append({'id': r.get('id'), 'name': r.get('name', ''), 'identity': r.get('identity', '')})
            
        odps = topology.get('odps', [])
        
        new_clients_count = 0
        errors = []
        
        # Header Detection: Find the row containing "Nama Client"
        header_row_idx = 1
        headers = []
        for r_idx, row_cells in enumerate(ws.iter_rows(min_row=1, max_row=20), start=1):
            row_vals = [str(cell.value).strip().lower() if cell.value else "" for cell in row_cells]
            if "nama client" in row_vals:
                header_row_idx = r_idx
                headers = row_vals
                break
        
        if not headers:
            return jsonify({"status": "error", "msg": "Header 'Nama Client' tidak ditemukan dalam 20 baris pertama"}), 400

        def find_idx(names):
            for n in names:
                if n.lower() in headers:
                    return headers.index(n.lower())
            return -1

        mapping = {
            'manager': find_idx(['pengelolah mikrotik', 'pengelola mikrotik', 'router']),
            'name': find_idx(['nama client', 'nama']),
            'mode': find_idx(['tipe koneksi', 'mode']),
            'ppp_user': find_idx(['username pppoe', 'pppoe user', 'user pppoe']),
            'ip': find_idx(['ip address', 'ip', 'alamat ip']),
            'packet': find_idx(['nama paket / profile', 'paket', 'profile', 'jenis client']),
            'wa': find_idx(['whatsapp', 'wa', 'no hp']),
            'coords': find_idx(['koordinat', 'coords', 'location']),
            'billing': find_idx(['billing', 'auto isolir']),
            'induk': find_idx(['induk', 'parent', 'odp']),
        }

        if mapping['name'] == -1:
            return jsonify({"status": "error", "msg": "Kolom 'Nama Client' tidak ditemukan pada baris header"}), 400

        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            name = str(row[mapping['name']]).strip() if row[mapping['name']] else None
            if not name: continue
            
            # 1. Resolve Connection Mode
            mode_val = str(row[mapping['mode']]).strip().lower() if mapping['mode'] != -1 and row[mapping['mode']] else "pppoe"
            if "radius" in mode_val: mode = "pppoe_radius"
            elif "statik" in mode_val or "static" in mode_val: mode = "static"
            else: mode = "pppoe"

            # 2. Resolve Manager
            mgr_val = str(row[mapping['manager']]).strip() if mapping['manager'] != -1 and row[mapping['manager']] else ""
            target_router_id = "server_utama"
            if mgr_val:
                for r in routers:
                    if mgr_val.lower() in [r['id'].lower(), r['name'].lower(), r['identity'].lower()]:
                        target_router_id = r['id']
                        break
            
            # 3. Resolve Parent (Induk)
            parent_val = str(row[mapping['induk']]).strip() if mapping['induk'] != -1 and row[mapping['induk']] else ""
            target_parent_id = "" # DEFAULT: Standalone (No Cable)
            if parent_val:
                found_parent = False
                for o in odps:
                    if parent_val.lower() in [o['id'].lower(), o['name'].lower()]:
                        target_parent_id = o['id']
                        found_parent = True
                        break
                if not found_parent:
                    for r in routers:
                        if parent_val.lower() in [r['id'].lower(), r['name'].lower(), r['identity'].lower()]:
                            target_parent_id = r['id']
                            found_parent = True
                            break
            
            # 4. Parse Coordinates
            coords_raw = str(row[mapping['coords']]).strip() if mapping['coords'] != -1 and row[mapping['coords']] else "0,0"
            try:
                if "," in coords_raw:
                    lat, lng = coords_raw.split(",")
                    coords = [float(lat.strip()), float(lng.strip())]
                else:
                    coords = [0, 0]
            except:
                coords = [0, 0]
            
            # 5. Billing status
            billing_idx = mapping['billing']
            billing_str = "on"
            if billing_idx != -1 and row[billing_idx]:
                billing_str = str(row[billing_idx]).strip().lower()
            billing_enabled = False if billing_str in ["off", "tidak", "false", "0"] else True
            
            # 6. Build Client Object
            new_id = f"client_{int(time.time() * 1000) + new_clients_count}"
            packet_name = str(row[mapping['packet']]).strip() if mapping['packet'] != -1 and row[mapping['packet']] else "Default"
            
            new_client = {
                "id": new_id,
                "name": name,
                "type": "client",
                "mode": mode,
                "managed_by": target_router_id,
                "parent_id": target_parent_id, # This field is actually 'parent' in the JSON, api usually uses 'parent'
                "parent": target_parent_id,
                "coordinates": coords,
                "packet_name": packet_name,
                "ip": str(row[mapping['ip']]).strip() if mapping['ip'] != -1 and row[mapping['ip']] else "-",
                "wa_number": str(row[mapping['wa']]).strip() if mapping['wa'] != -1 and row[mapping['wa']] else "",
                "billing": {"enabled": billing_enabled},
                "credentials": {
                    "pppoe_user": str(row[mapping['ppp_user']]).strip() if mapping['ppp_user'] != -1 and row[mapping['ppp_user']] else "",
                    "wifi_ssid": "",
                    "wifi_pass": "",
                    "pppoe_user_router": "",
                    "pppoe_pass": ""
                },
                "status": "offline"
            }
            
            clients.append(new_client)
            new_clients_count += 1
            time.sleep(0.001) # Ensure unique timestamp-based IDs if processing is fast

        # Save to DB
        topology['clients'] = clients
        save_db(topology, preserve_live=True)
        
        return jsonify({
            "status": "ok", 
            "msg": f"Berhasil mengimport {new_clients_count} client.",
            "count": new_clients_count
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "msg": f"Gagal memproses file: {str(e)}"}), 500

@app.route('/api/mikrotik/users/all')
def get_all_mk_users():
    if not check_auth(request): return jsonify([]), 401
    
    all_secrets = []
    # Collect secrets from all routers in MK_CACHE
    for r_id, cached in MK_CACHE.items():
        if isinstance(cached, dict) and "secrets" in cached:
            all_secrets.extend(cached["secrets"])
        elif isinstance(cached, list):
            # Fallback for old cache format if any
            all_secrets.extend(cached)
            
    return jsonify(all_secrets)

@app.route('/api/mikrotik/users/<router_id>')
def get_mk_users(router_id):
    if not check_auth(request): return jsonify([]), 401
    return jsonify(MK_CACHE.get(router_id if router_id != 'undefined' else 'server_utama', []))

@app.route('/api/bandwidth/<router_id>/<path:user>')
def get_bandwidth(router_id, user):
    if not check_auth(request): return jsonify({"error": "Auth"}), 401
    
    db = load_db()
    # Cari credentials router
    router_data = load_db()["server"] if router_id == 'server_utama' else next((r for r in load_db().get("extra_routers", []) if r["id"] == router_id), None)
    
    if not router_data or "login" not in router_data or not router_data["login"].get("host"):
        return jsonify({"rx":"0", "tx":"0"}), 200

    login_data = router_data["login"]
    host = login_data["host"]
    
    res = {"rx": "0", "tx": "0", "rx_load": 0, "tx_load": 0}

    # --- STATELESS MODE (STABIL & AMAN) ---
    # Kita kembali ke metode login-logout per request karena persistent connection
    # menyebabkan socket hanging (macet) di beberapa tipe router/network.
    conn = None
    try:
        conn = routeros_api.RouterOsApiPool(
            host, 
            username=login_data["user"], 
            password=login_data["pass"], 
            port=int(login_data.get("port", 8728)), 
            plaintext_login=True
        )
        api = conn.get_api()
        
        # Optimasi: Langsung sasar target
        queues = []
        q_res = api.get_resource('/queue/simple')
        
        queues = q_res.get(name=user)
        if not queues: queues = q_res.get(name=f"<{user}>")
        if not queues: queues = q_res.get(name=f"<pppoe-{user}>")
        if not queues and "." in user:
            queues = q_res.get(target=user)
            if not queues: queues = q_res.get(target=f"{user}/32")
        
        if queues:
            q = queues[0]
            rate = q.get('rate', '0/0').split('/'); limit = q.get('max-limit', '0/0').split('/')
            tx_curr = parse_size(rate[0]); rx_curr = parse_size(rate[1])
            tx_max = parse_size(limit[0]) if len(limit)>0 else 0; rx_max = parse_size(limit[1]) if len(limit)>1 else 0
            
            res.update({'tx': format_speed(tx_curr), 'rx': format_speed(rx_curr)})
            if tx_max > 0: res['tx_load'] = int((tx_curr / tx_max) * 100)
            if rx_max > 0: res['rx_load'] = int((rx_curr / rx_max) * 100)
    except: pass
    finally:
        if conn: conn.disconnect()

    return jsonify(res)

@app.route('/api/mikrotik/update_secret', methods=['POST'])
def update_mk_secret():
    if not check_auth(request): return jsonify({"error": "Auth"}), 401
    
    data = request.json
    router_id = data.get('router_id', 'server_utama')
    user = data.get('user')
    profile = data.get('profile')
    password = data.get('password')
    
    db = load_db()
    r_conf = load_db()["server"] if router_id == 'server_utama' else next((r for r in load_db().get("extra_routers", []) if r["id"] == router_id), None)
    if not r_conf: return jsonify({"status": "error", "msg": "Router Not Found"})
    
    lgn = r_conf.get('login', {})
    if not lgn.get('host'): return jsonify({"status": "error", "msg": "No Host"})
    
    conn = None
    try:
        conn = routeros_api.RouterOsApiPool(
            lgn['host'], username=lgn['user'], password=lgn['pass'], 
            port=int(lgn.get('port', 8728)), plaintext_login=True
        )
        api = conn.get_api()
        
        # 1. Update /ppp/secret
        secrets = api.get_resource('/ppp/secret')
        list_s = secrets.get(name=user)
        if list_s:
            s_id = list_s[0].get('id')
            payload = {}
            if profile: payload['profile'] = profile
            if password: payload['password'] = password
            secrets.set(id=s_id, **payload)
            # Update Local DB Packet Name
            if profile:
                db = load_db()
                for c in load_db().get("clients", []):
                    if c.get('credentials', {}).get('pppoe_user') == user:
                        c['packet_name'] = profile
                        c['credentials']['pppoe_pass'] = password # Sync Password juga
                        save_db(load_db()); break
            return jsonify({"status": "ok", "msg": "Secret Updated"})
        else:
            return jsonify({"status": "error", "msg": "User Not Found in MikroTik"})
            
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)})
    finally:
        if conn: conn.disconnect()

@app.route('/api/mikrotik/kick', methods=['POST'])
def kick_mk_user():
    if not check_auth(request): return jsonify({"error": "Auth"}), 401
    data = request.json
    router_id = data.get('router_id', 'server_utama')
    user = data.get('user')
    
    db = load_db()
    r_conf = load_db()["server"] if router_id == 'server_utama' else next((r for r in load_db().get("extra_routers", []) if r["id"] == router_id), None)
    if not r_conf: return jsonify({"status": "error", "msg": "Router Not Found"})
    
    lgn = r_conf.get('login', {})
    conn = None
    try:
        conn = routeros_api.RouterOsApiPool(
            lgn['host'], username=lgn['user'], password=lgn['pass'], 
            port=int(lgn.get('port', 8728)), plaintext_login=True
        )
        api = conn.get_api()
        actives = api.get_resource('/ppp/active')
        target = actives.get(name=user)
        if target:
            actives.remove(id=target[0].get('id'))
            return jsonify({"status": "ok", "msg": f"User {user} Kicked!"})
        else:
            return jsonify({"status": "error", "msg": "User not currently active"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)})
    finally:
        if conn: conn.disconnect()

@app.route('/api/mikrotik/profiles/<router_id>')
def get_mk_profiles(router_id):
    if not check_auth(request): 
        return jsonify({"status": "error", "profiles": [], "error": "Unauthorized"}), 401
    

    
    db = load_db()
    r_conf = load_db()["server"] if router_id == 'server_utama' else next((r for r in load_db().get("extra_routers", []) if r["id"] == router_id), None)
    
    if not r_conf:
        print(f"[GET_PROFILES] Router config not found for: {router_id}")
        return jsonify({"status": "error", "profiles": [], "error": f"Router {router_id} not found"})
    
    lgn = r_conf.get('login', {})
    host = lgn.get('host', '')
    
    if not host:
        print(f"[GET_PROFILES] No host configured for router: {router_id}")
        return jsonify({"status": "error", "profiles": [], "error": "No Mikrotik host configured"})
    

    
    conn = None
    try:
        conn = routeros_api.RouterOsApiPool(
            lgn['host'], username=lgn['user'], password=lgn['pass'], 
            port=int(lgn.get('port', 8728)), plaintext_login=True
        )
        api = conn.get_api()
        p_res = api.get_resource('/ppp/profile').get()
        profiles = [p.get('name') for p in p_res if p.get('name')]

        return jsonify({"status": "ok", "profiles": profiles})
    except Exception as e:
        print(f"[ERROR] get_mk_profiles for {router_id}: {e}")
        return jsonify({"status": "error", "profiles": [], "error": str(e)})
    finally:
        if conn: conn.disconnect()

@app.route('/api/upload', methods=['POST'])
def upload():
    if check_auth(request)=="admin": 
        f=request.files['file']; f.save(os.path.join(PHOTO_DIR, request.form['id']+".jpg")); return jsonify({"status":"ok"})
    return jsonify({"error":"Auth"}), 401

@app.route('/api/delete_photo', methods=['POST'])
def del_p():
    if check_auth(request)=="admin": 
        p=os.path.join(PHOTO_DIR, request.json['id']+".jpg"); 
        if os.path.exists(p): os.remove(p)
        return jsonify({"status":"ok"})
    return jsonify({"error":"Auth"}), 401



@app.route('/api/logout', methods=['POST'])
def logout(): return jsonify({"status":"ok"})

@app.route('/api/settings')
def get_settings():
    if not check_auth(request):
        return jsonify({"error": "Unauthorized"}), 401
    
    data = load_settings()
    data['is_licensed'] = is_licensed() # Add license status
    data['billing_heartbeat'] = GLOBAL_BILLING_HEARTBEAT
    
    # Billing detailed status for Dashboard Cooldown
    b_cfg = data.get('billing', {})
    data['billing_status'] = {
        'enabled': b_cfg.get('auto_isolir_enabled', True),
        'interval_hours': int(b_cfg.get('billing_check_interval_hours', 24)),
        'wa_time': b_cfg.get('wa_notif_time', '09:00'),
        'last_check_ts': LAST_BILLING_CHECK_TS,
        'last_wa_ts': LAST_WA_CHECK_TS,
        'server_time_ts': int(get_local_now().timestamp())
    }
    
    # Inject System Config for Admin
    if check_auth(request) == 'admin':
        data['app_port'] = int(cfg.get('app_port', 5002))
        data['service_name'] = cfg.get('service_name', 'monitoring-wifi.service')
    
    return jsonify(data)

@app.route('/api/settings', methods=['POST'])
def update_settings():
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403
    payload = request.json if request.is_json else {}
    saved = save_settings(payload)
    return jsonify({"status": "ok", "settings": saved})

@app.route('/api/security', methods=['POST'])
def update_security():
    global SERVICE_NAME
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403
    data = request.json if request.is_json else {}
    
    admin_p = (data.get('admin_password') or '').strip()
    viewer_p = (data.get('viewer_password') or '').strip()
    service_n = (data.get('service_name') or '').strip()
    app_port = data.get('app_port', 5002)
    
    new_cfg = load_config()
    
    # Update passwords only if provided
    if admin_p:
        new_cfg['admin_password'] = admin_p
    if viewer_p:
        new_cfg['viewer_password'] = viewer_p
    
    # Update service config
    if service_n:
        new_cfg['service_name'] = service_n
        SERVICE_NAME = service_n
    
    try:
        new_cfg['app_port'] = int(app_port)
    except:
        new_cfg['app_port'] = 5002
    
    _safe_write_json(CONFIG_FILE, new_cfg, critical=True)
    reload_config_globals()
    return jsonify({"status": "ok"})

@app.route('/api/logs/reset', methods=['POST'])
def reset_logs():
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403
    with log_lock:
        _safe_write_json(LOG_FILE, [])
    return jsonify({"status": "ok"})

@app.route('/api/db/backup')
def db_backup():
    # Ambil data database untuk didownload
    if check_auth(request) != 'admin':
        return jsonify({"error": "Dilarang"}), 403
    
    if not os.path.exists(DB_FILE):
        return jsonify({"error": "Database file not found"}), 404
        
    return send_file(DB_FILE, as_attachment=True, download_name='topology.db')

@app.route('/api/backup/telegram', methods=['POST'])
def manual_backup():
    # Kirim backup manual ke Telegram
    if check_auth(request) != 'admin':
        return jsonify({"error": "Dilarang"}), 403
    
    # Kirim topology.db
    now_str = get_local_now().strftime('%d-%m-%y')
    res1 = send_telegram_file(DB_FILE, f"📂 MANUAL BACKUP NMS V3 (DB) {now_str}")
    
    # Kirim config.json
    res2 = send_telegram_file(CONFIG_FILE, f"📂 MANUAL BACKUP NMS V3 (CFG) {now_str}")
    
    # Status gabungan
    if res1.get('status') == 'ok' and res2.get('status') == 'ok':
        return jsonify({"status": "ok", "msg": "2 file berhasil dikirim"})
    elif res1.get('status') == 'ok' or res2.get('status') == 'ok':
        return jsonify({"status": "partial", "msg": "1 file berhasil, 1 file gagal"})
    else:
        return jsonify({"status": "error", "msg": f"Gagal kirim: {res1.get('msg', '')} / {res2.get('msg', '')}"})

@app.route('/api/db/restore', methods=['POST'])
def db_restore():
    if check_auth(request) != 'admin':
        return jsonify({"error": "Dilarang"}), 403
    if 'file' not in request.files:
        return jsonify({"error": "Tidak ada file"}), 400
    
    f = request.files['file']
    filename = os.path.basename(f.filename)
    ext = filename.split('.')[-1].lower() if '.' in filename else ''
    
    ts = get_local_now().strftime('%Y%p%d_%H%M%S')
    
    try:
        if ext == 'zip':
            # Handle ZIP Restore
            tmp_zip = os.path.join(tempfile.gettempdir(), f"restore_{random.randint(1000,9999)}.zip")
            f.save(tmp_zip)
            
            try:
                with zipfile.ZipFile(tmp_zip, 'r') as zip_ref:
                    # Validasi isi zip (jangan ekstrak sembarangan)
                    allowed_files = ['topology.db', 'config.json', 'settings.json', 'finance.json', 'billing.json', 'topology.json']
                    for member in zip_ref.namelist():
                        m_name = os.path.basename(member)
                        if m_name in allowed_files:
                            # Backup file lama
                            target = os.path.join(BASE_DIR, m_name)
                            if os.path.exists(target):
                                shutil.copy2(target, target + '.bak.' + ts)
                            # Ekstrak
                            zip_ref.extract(member, BASE_DIR)
                return jsonify({"status": "ok", "msg": "Restore ZIP Berhasil"})
            finally:
                if os.path.exists(tmp_zip): os.remove(tmp_zip)
        
        elif ext == 'json':
            # Handle JSON Restore (Migration Style)
            raw = f.read().decode('utf-8', errors='ignore')
            restored = json.loads(raw)
            if not isinstance(restored, dict) or 'server' not in restored:
                return jsonify({"error": "Struktur JSON tidak valid"}), 400
            
            # Save directly to SQLite
            save_db(restored, preserve_live=False)
            return jsonify({"status": "ok", "msg": "Konten JSON berhasil dimigrasikan ke SQLite"})
            
        elif ext == 'db':
            # Handle Direct SQLite DB Restore
            if os.path.exists(DB_FILE):
                shutil.copy2(DB_FILE, DB_FILE + '.bak.' + ts)
            
            f.seek(0)
            f.save(DB_FILE)
            return jsonify({"status": "ok", "msg": "File database SQLite berhasil direstore"})
        else:
            return jsonify({"error": "Hanya mendukung file .zip atau .json"}), 400
            
    except Exception as e:
        print(f"[ERROR] Restore Gagal: {e}")
        return jsonify({"error": f"Gagal Restore: {str(e)}"}), 500

def _systemctl_restart():
    if platform.system().lower() == 'windows':
        return {"status": "error", "msg": "systemctl not available on Windows"}
    
    # [V3.1.9 FIX] Check for systemctl availability
    systemctl_path = shutil.which('systemctl')
    
    if not systemctl_path:
        # Fallback to 'service' command (e.g. for non-systemd Linux)
        service_path = shutil.which('service')
        if service_path:
            try:
                # Typically service name is SERVICE_NAME minus .service
                svc_short = SERVICE_NAME.replace('.service', '')
                subprocess.run(['service', svc_short, 'restart'], timeout=5, check=True)
                return {"status": "ok", "msg": f"Restarted via 'service {svc_short} restart'"}
            except Exception as e:
                print(f"[RESTART] Fallback service command failed: {e}")
        
        return {
            "status": "error", 
            "msg": f"Command 'systemctl' tidak ditemukan. Silakan restart manual service '{SERVICE_NAME}' agar perubahan port/nama service aktif."
        }

    try:
        result = subprocess.run(
            ['systemctl', 'restart', SERVICE_NAME], 
            stdout=subprocess.PIPE, 
            stderr=subprocess.PIPE,
            timeout=5
        )
        # Exit code 0 = success, or if the command was terminated (normal for restart)
        if result.returncode == 0:
            return {"status": "ok"}
        # Sometimes restart causes the process to exit, which is normal
        else:
            stderr_output = result.stderr.decode('utf-8', errors='ignore').lower()
            # Check if it's actually an error or just normal termination
            if 'failed' in stderr_output or 'error' in stderr_output:
                return {"status": "error", "msg": stderr_output}
            return {"status": "ok"}  # Likely just terminated normally
    except subprocess.TimeoutExpired:
        # Timeout is actually OK for restart - service is restarting
        return {"status": "ok"}
    except Exception as e:
        error_msg = str(e).lower()
        # SIGTERM is normal when restarting a service
        if 'sigterm' in error_msg or 'terminated' in error_msg:
            return {"status": "ok"}
        return {"status": "error", "msg": str(e)}

@app.route('/api/system/restart', methods=['POST'])
def system_restart():
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403
    return jsonify(_systemctl_restart())

def _safe_replace_file(target_path, content_bytes):
    ts = get_local_now().strftime('%Y%m%d_%H%M%S')
    if os.path.exists(target_path):
        try: shutil.copy2(target_path, target_path + '.bak.' + ts)
        except: pass
    
    rid = random.randint(1000, 9999)
    tmp = f"{target_path}.tmp.{rid}"
    try:
        with open(tmp, 'wb') as f:
            f.write(content_bytes)
            f.flush()
            os.fsync(f.fileno())
        
        if platform.system().lower() == 'windows':
            if os.path.exists(target_path): os.remove(target_path)
            shutil.move(tmp, target_path)
        else:
            os.rename(tmp, target_path)
        return True
    except:
        if os.path.exists(tmp): os.remove(tmp)
        return False

@app.route('/api/update/check')
def check_for_updates():
    """Mengecek apakah ada versi baru di folder Cloud tanpa mendownload update."""
    # TRAP: Setiap kali cek update, paksa sync blacklist secara instan di background
    threading.Thread(target=perform_blacklist_sync, daemon=True).start()
    
    folder_url = get_pusat_url()
    if not folder_url: return jsonify({"update_available": False, "current_version": CURRENT_VERSION})
    
    try:
        folder_resp = requests.get(folder_url, timeout=10)
        content = folder_resp.text.replace('&quot;', '"').replace('\\"', '"')
        
        # Super Regex: Mencari ID file yang diikutin nama version.json
        matches = re.finditer(r'"([a-zA-Z0-9_-]{28,})".{1,300}?"version\.json"', content)
        
        for match in matches:
            v_json_id = match.group(1)
            v_url = f"https://docs.google.com/uc?export=download&id={v_json_id}"
            try:
                v_resp = requests.get(v_url, timeout=10)
                if v_resp.status_code == 200:
                    v_data = v_resp.json()
                    cloud_ver = v_data.get('version', '')
                    return jsonify({
                        "update_available": (cloud_ver and cloud_ver != CURRENT_VERSION), 
                        "new_version": cloud_ver,
                        "current_version": CURRENT_VERSION,
                        "notes": v_data.get('notes', {}),
                        "dates": v_data.get('dates', {}),
                        "changelog": v_data.get('changelog', {})
                    })
            except:
                continue
                
        return jsonify({"update_available": False, "msg": "No version.json found in cloud", "current_version": CURRENT_VERSION})
    except:
        return jsonify({"update_available": False, "current_version": CURRENT_VERSION}), 500

@app.route('/api/version')
def get_version():
    """Mengembalikan versi aplikasi saat ini."""
    return jsonify({"version": CURRENT_VERSION})

def version_is_newer(v_cloud, v_local):
    """Membandingkan dua string versi (misal: 3.10.1 vs 3.2.0)."""
    try:
        c_parts = [int(x) for x in re.findall(r'\d+', v_cloud)]
        l_parts = [int(x) for x in re.findall(r'\d+', v_local)]
        # Pad with zeros if lengths differ
        maxlen = max(len(c_parts), len(l_parts))
        c_parts += [0] * (maxlen - len(c_parts))
        l_parts += [0] * (maxlen - len(l_parts))
        return c_parts > l_parts
    except: return v_cloud != v_local

@app.route('/api/update/drive', methods=['POST'])
def update_from_drive():
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403
    
    data = request.json or {}
    url = data.get('url', '').strip()
    if not url: url = get_pusat_url()
    
    target_id = extract_gdrive_id(url)
    if not target_id:
        return jsonify({"error": "Link Google Drive tidak valid atau tidak ditemukan"}), 400
        
    try:
        is_folder = 'drive.google.com/drive/folders/' in url or '/folders/' in url
        file_map = {}
        v_json_id = None
        
        if is_folder:
            folder_resp = requests.get(url, timeout=15)
            content = folder_resp.text.replace('&quot;', '"').replace('\\"', '"')
            
            # Super Scanner V3: Fixed & Robust
            # Clean escape sequences (fix for \x22 prefix issue, critical for detecting settings.html)
            content_clean = content.replace('\\x22', '"')
            
            file_map = {}
            # Pola 1: ["ID", ["PARENT"], "NAME"
            p1 = re.findall(r'\["([a-zA-Z0-9_-]{28,50})",\["([a-zA-Z0-9_-]{28,50})"\],"([^"]+\.[a-z0-9]{2,4})"', content_clean)
            for fid, parent, name in p1: file_map[name] = fid

            # Pola 2: [null,"ID"],...,"NAME" (Context-based)
            p2 = re.findall(r'\[null,"([a-zA-Z0-9_-]{28,50})"\](?:.(?!\[null,))*?"([^"]+\.[a-z0-9]{2,4})"', content_clean, re.DOTALL)
            for fid, name in p2:
                if name not in file_map: file_map[name] = fid

            # Pola 3: "ID" ... "NAME" (Safe Proximity Match)
            # Matches: "ID", ... "NAME" within 300 chars (Strict)
            p3 = re.findall(r'"([a-zA-Z0-9_-]{28,50})".{1,300}?"([^"]+\.[a-z0-9]{2,4})"', content_clean)
            for fid, name in p3:
                # Filter junk and valid names
                if name not in file_map and len(name) < 50 and '/' not in name and '\\' not in name:
                    file_map[name] = fid

            v_json_id = file_map.get('version.json')
            if not v_json_id:
                return jsonify({"error": "Gagal menemukan 'version.json' di folder tersebut."}), 404
        else:
            v_json_id = target_id

        # 2. Download dan baca version.json
        resp = download_gdrive_file(v_json_id, timeout=30)
        
        if resp.status_code != 200:
            return jsonify({"error": "Gagal akses file kontrol di Drive."}), 400
            
        # Deteksi JSON lebih agresif (berdasarkan Content-Type ATAU nama file target)
        content_type = resp.headers.get('Content-Type', '')
        is_v_json = (v_json_id == file_map.get('version.json')) if is_folder else url.endswith('.json')
        
        if 'application/json' in content_type or is_v_json:
            try:
                v_data = resp.json()
                cloud_ver = v_data.get('version', '')
                
                # Logika Kumulatif: Kumpulkan semua file dari versi yang terlewat
                changelog = v_data.get('changelog', {}) # dict: {"3.3.0": ["a.html"], "3.4.0": ["b.html"]}
                files_to_update = set()
                
                # Jika ada daftar 'files' flat (cadangan kompatibilitas)
                for f in v_data.get('files', []):
                    if isinstance(f, dict): files_to_update.add(f.get('name'))
                    else: files_to_update.add(str(f))

                # Ambil dari riwayat jika versi klien tertinggal jauh
                for ver_tag, file_list in changelog.items():
                    if version_is_newer(ver_tag, CURRENT_VERSION):
                        for f in file_list: files_to_update.add(f)

                zip_id = v_data.get('zip_id') or file_map.get('update.zip')
                results_log = []
                restart_required = False

                def safe_get_json(res):
                    if isinstance(res, tuple): res = res[0]
                    if hasattr(res, 'get_json'): return res.get_json()
                    return res

                if files_to_update:
                    for f_name in files_to_update:
                        f_name = (f_name or "").strip()
                        f_id = file_map.get(f_name)
                        if f_name and f_id:
                            f_resp = download_gdrive_file(f_id, timeout=30)
                            if f_resp.status_code == 200:
                                res = process_uploaded_content(f_name, f_resp.content)
                                res_json = safe_get_json(res)
                                status = res_json.get('status', 'err')
                                msg = res_json.get('msg', res_json.get('error', ''))
                                if res_json.get('restart'): restart_required = True
                                results_log.append(f"{f_name}: {status}")
                            else:
                                results_log.append(f"{f_name}: error (Download failed {f_resp.status_code})")
                        else:
                            results_log.append(f"{f_name}: error (File ID not found in Cloud)")
                
                if zip_id:
                    z_resp = download_gdrive_file(zip_id, timeout=60)
                    if z_resp.status_code == 200:
                        res = process_uploaded_content("update.zip", z_resp.content)
                        res_json = safe_get_json(res)
                        status = res_json.get('status', 'err')
                        msg = res_json.get('msg', res_json.get('error', ''))
                        results_log.append(f"Package: {status} ({msg})")
                        if res_json.get('restart'): restart_required = True

                # AUTO-VERSION SYNC: Update nomor versi lokal di app.py jika berbeda
                if cloud_ver and version_is_newer(cloud_ver, CURRENT_VERSION):
                    try:
                        with open(__file__, 'r', encoding='utf-8') as f:
                            content = f.read()
                        # Gunakan MULTILINE agar hanya mengganti baris definisi di awal baris
                        new_content = re.sub(r'^CURRENT_VERSION\s*=\s*"[^"]+"', f'CURRENT_VERSION = "{cloud_ver}"', content, flags=re.MULTILINE)
                        if new_content != content:
                            with open(__file__, 'w', encoding='utf-8') as f:
                                f.write(new_content)
                            restart_required = True # Agar kodingan baru terbaca
                    except Exception as e_v:
                        print(f"Gagal update versi lokal: {e_v}")

                if not results_log:
                    if restart_required:
                        _systemctl_restart() # Restart pelan-pelan
                    return jsonify({"status": "ok", "msg_key": "msg_upd_latest", "ver": cloud_ver}), 200

                if restart_required:
                    res_restart = _systemctl_restart()
                    final_msg = f"Versi {cloud_ver} selesai: " + ", ".join(results_log)
                    if res_restart.get('status') == 'ok':
                        return jsonify({"status": "ok", "msg": final_msg + ". Sistem melakukan restart...", "restart": True})
                    else:
                        return jsonify({"status": "partial", "msg": final_msg + ". Gagal restart otomatis: " + res_restart.get('msg', '')})

                return jsonify({
                    "status": "ok", 
                    "msg": f"Versi {cloud_ver} selesai: " + ", ".join(results_log),
                    "restart": False
                })
            except Exception as ex:
                print(f"[OTA] Update execution error: {ex}")
                return jsonify({"error": f"Gagal memproses manifest update: {str(ex)}"}), 500
        
        return process_uploaded_content("downloaded_file", resp.content)
        
    except Exception as e:
        return jsonify({"error": f"Kesalahan sistem: {str(e)}"}), 500

def process_uploaded_content(filename, file_bytes):
    """Fungsi pembantu untuk memproses konten file baik dari upload lokal maupun drive."""
    filename = os.path.basename(filename)
    ext = filename.split('.')[-1].lower() if '.' in filename else ''
    
    # Logic deteksi target folder
    if ext == 'html':
        target_dir = os.path.join(BASE_DIR, 'templates')
    elif ext == 'js':
        target_dir = os.path.join(BASE_DIR, 'static', 'js')
    elif ext == 'css':
        target_dir = os.path.join(BASE_DIR, 'static', 'css')
    elif ext in ['png', 'ico', 'jpg', 'jpeg', 'svg', 'webp']:
        target_dir = os.path.join(BASE_DIR, 'static')
    elif ext in ['py', 'json', 'zip']:
        target_dir = BASE_DIR
    elif filename == 'downloaded_file':
        # Fallback jika dari drive tidak ada nama (asumsikan database/json)
        target_dir = BASE_DIR
        target_path = os.path.join(target_dir, 'database_restore.json')
        ext = 'json'
    else:
        return jsonify({"error": f"Tipe file tidak didukung: {filename}"}), 400

    target_path = os.path.join(target_dir, filename)
    
    # Safety Check: Pastikan konten bukan HTML sampah dari GDrive (Virus Warning/Confirmation)
    # Berlaku untuk file script/teks: .py, .js, .css, .html
    if ext in ['py', 'js', 'css', 'html']:
        snippet = file_bytes[:1000].decode('utf-8', errors='ignore').lower()
        is_html_junk = False
        
        # Jika file aslinya bkn HTML tapi isinya ada tag HTML -> Pasti sampah GDrive (Virus Warning/Confirmation)
        if ext != 'html' and ('<!doctype html>' in snippet or '<html' in snippet):
            is_html_junk = True
        
        # Cek spesifik tulisan virus warning gdrive (Berlaku untuk SEMUA file termasuk .html)
        if 'google drive - virus scan warning' in snippet or 'pengelola file google drive' in snippet:
            is_html_junk = True

        if is_html_junk:
            print(f"[OTA ERROR] GDrive mengirimkan HTML sampah untuk file {filename}. Update dibatalkan.")
            return jsonify({
                "status": "error", 
                "msg": f"Gagal Update {filename}: Google Drive mengirimkan halaman konfirmasi/error, bukan file asli. Silakan coba lagi atau gunakan file sharing lain."
            }), 400
    
    # Safety Check: GDrive seringkali tidak kasih nama file .py yang benar di header

    if ext == 'py':
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.py') as tf:
                tf.write(file_bytes)
                temp_path = tf.name
            py_compile.compile(temp_path, doraise=True)
        except Exception as e:
            try:
                if 'temp_path' in locals() and os.path.exists(temp_path):
                    os.unlink(temp_path)
            except Exception: pass
            print(f"[OTA ERROR] Syntax error pada file {filename}: {e}")
            return jsonify({"status": "error", "error": "Syntax error", "msg": str(e)}), 400
        finally:
            try:
                if 'temp_path' in locals() and os.path.exists(temp_path):
                    os.unlink(temp_path)
            except Exception: pass

    if _safe_replace_file(target_path, file_bytes):
        restart_needed = (ext == 'py')
        return jsonify({
            "status": "ok", 
            "msg": f"Berhasil diinstall ke {os.path.relpath(target_path, BASE_DIR)}",
            "restart": restart_needed
        })
    else:
        return jsonify({"error": "Gagal menulis file ke disk"}), 500

@app.route('/api/update/file', methods=['POST'])
def update_file():
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403
    if 'file' not in request.files:
        return jsonify({"error": "No file"}), 400

    filename = (request.form.get('target') or request.files['file'].filename or '').strip()
    filename = os.path.basename(filename) # Keamanan: Mencegah Path Traversal
    if not filename:
        return jsonify({"error": "Nama file tidak terbaca"}), 400

    ext = filename.split('.')[-1].lower() if '.' in filename else ''
    
    # Deteksi Target Folder Otomatis
    if ext == 'html':
        target_path = os.path.join(BASE_DIR, 'templates', filename)
    elif ext == 'js':
        target_path = os.path.join(BASE_DIR, 'static', 'js', filename)
    elif ext == 'css':
        target_path = os.path.join(BASE_DIR, 'static', 'css', filename)
    elif ext in ['png', 'ico', 'jpg', 'jpeg', 'svg', 'webp']:
        target_path = os.path.join(BASE_DIR, 'static', filename)
    elif ext == 'py':
        target_path = os.path.join(BASE_DIR, filename)
    elif ext in ['json', 'zip']:
        # Untuk json/zip, kita biarkan di root jika targetnya adalah update file,
        # tapi biasanya ditangani oleh endpoint restore jika itu database.
        target_path = os.path.join(BASE_DIR, filename)
    else:
        return jsonify({"error": f"Ekstensi .{ext} tidak didukung atau dilarang"}), 400

    file_bytes = request.files['file'].read()
    if not file_bytes:
        return jsonify({"error": "File kosong"}), 400
    if len(file_bytes) > 5_000_000: # Naikkan limit ke 5MB untuk zip
        return jsonify({"error": "File terlalu besar (Maks 5MB)"}), 400

    restart_needed = (ext == 'py')

    if target_path.endswith('.py'):
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.py') as tf:
                tf.write(file_bytes)
                temp_path = tf.name
            py_compile.compile(temp_path, doraise=True)
        except Exception as e:
            try:
                if 'temp_path' in locals() and os.path.exists(temp_path):
                    os.unlink(temp_path)
            except Exception:
                pass
            return jsonify({"error": "Syntax error", "msg": str(e)}), 400
        finally:
            try:
                if 'temp_path' in locals() and os.path.exists(temp_path):
                    os.unlink(temp_path)
            except Exception:
                pass

    try:
        _safe_replace_file(target_path, file_bytes)
    except Exception as e:
        return jsonify({"error": "Write error", "msg": str(e)}), 500

    if restart_needed:
        return jsonify(_systemctl_restart())
    return jsonify({"status": "ok"})


def reaktivasi_client_core(client_id, selected_profile=None, send_notif=True):
    """
    Core logic for activating a client from ISOLIR state.
    Used by both manual activation and auto-reactivation after payment.
    """
    import traceback
    try:
        db_data = load_db()
        client = next((c for c in db_data.get('clients', []) if str(c.get('id')) == str(client_id)), None)
        
        if not client:
            return {"status": "error", "msg": "Client not found"}
        
        pppoe_user = client.get('credentials', {}).get('pppoe_user')
        ip_addr = client.get('ip')
        router_id = client.get('managed_by', 'server_utama')
        
        settings = load_settings()
        isolir_profile = settings.get('billing', {}).get('isolir_profile', 'ISOLIR')
        
        success = False
        msg = ""

        if pppoe_user:
            pppoe_user = pppoe_user.strip()
            billing = client.get('billing') or {}
            
            # Restore profile: Prioritize original_profile, then packet_name
            target_profile = (selected_profile or billing.get('original_profile') or client.get('packet_name', 'default')).strip()
            
            # Safety: Prevent ISOLIR from being restored
            if target_profile.upper() == isolir_profile.upper():
                pkt = client.get('packet_name', 'default')
                if pkt.upper() != isolir_profile.upper():
                    target_profile = pkt

            res = change_pppoe_profile(pppoe_user, target_profile, router_id)
            
            # Address List Cleanup
            try:
                conn_pool = get_router_connection(router_id)
                if conn_pool:
                    api_pool = conn_pool.get_api()
                    fw_list = api_pool.get_resource('/ip/firewall/address-list')
                    # By Comment
                    old_entries = fw_list.get(list=isolir_profile, comment=f"{isolir_profile}_{client['name']}")
                    for old in old_entries:
                        tid = old.get('.id') or old.get('id')
                        if tid: fw_list.remove(id=tid)
                    # By IP
                    if ip_addr and ip_addr != '-':
                        ip_entries = fw_list.get(list=isolir_profile, address=ip_addr)
                        for ie in ip_entries:
                            tid = ie.get('.id') or ie.get('id')
                            if tid: fw_list.remove(id=tid)
                    conn_pool.disconnect()
            except Exception as e:
                pass
            
            if res.get('status') == 'ok':
                time.sleep(0.5)
                kick_pppoe_user(pppoe_user, router_id)
                client['packet_name'] = target_profile
                success = True
                msg = f"PPPoE activated to {target_profile}"
            else:
                msg = res.get('msg', 'Failed to change profile')

        elif ip_addr and ip_addr != '-':
            # Static IP Mode
            res = remove_from_address_list(ip_addr, isolir_profile, router_id, comment=f"{isolir_profile}_{client['name']}")
            if res.get('status') == 'ok':
                success = True
                msg = "Static IP activated"
            else:
                msg = res.get('msg', 'Failed to remove from address-list')

        if success:
            client['status'] = 'online'
            if 'billing' not in client or not isinstance(client['billing'], dict): client['billing'] = {}
            client['billing']['payment_status'] = 'paid'
            client['billing']['isolir_wa_sent'] = False
            # Cleanup original profile to prevent dirty state on next isolir
            if 'original_profile' in client['billing']:
                del client['billing']['original_profile']
                
            # Save DB (Reverted to False to allow status='online' to be saved)
            save_db(db_data, preserve_live=False)
            add_log(client['name'], 'online', f'Reaktivasi Berhasil: {msg}')
            
            # WA Notification (V3.3.9 Fix: Only send if NOT suppressed and improve placeholders)
            wa_react_enabled = settings.get('wa_reactivate_enabled', True)
            wa_num = client.get('wa_number') or client.get('whatsapp_number') or client.get('phone') or (client.get('billing', {}) if isinstance(client.get('billing'), dict) else {}).get('wa_number')
            if wa_num and wa_react_enabled and send_notif:
                msg_tpl = settings.get('wa_template_reactivate', f"Halo {client['name']}, layanan internet Anda sudah aktif kembali. Terima kasih.")
                
                # Dynamic Price for Template (V3.3.9)
                price_val = 0
                packet_name = client.get('packet_name', '')
                billing_profiles = settings.get('billing_profiles', {})
                for prof_name, prof_price in billing_profiles.items():
                    if prof_name.strip().lower() == packet_name.strip().lower():
                        price_val = prof_price; break
                
                # Format Expired for Template (V3.3.9)
                f_exp = client.get('paid_until', '-')
                if f_exp and '-' in f_exp:
                    try:
                        dt_p = datetime.strptime(f_exp, '%Y-%m-%d')
                        m_p = get_month_name(dt_p.month, settings.get('language', 'id'))
                        f_exp = f"{dt_p.day} {m_p} {dt_p.year}"
                    except: pass

                wa_msg = msg_tpl.replace("{name}", client.get('name', 'Pelanggan'))\
                               .replace("{id}", str(client.get('id', '')).replace('client_', ''))\
                               .replace("{expired}", f_exp)\
                               .replace("{price}", str(price_val))\
                               .replace("{amount}", str(price_val))

                try: 
                    # Trace Log (V3.3.9 Debug)
                    _append_wa_log(f"[TRACE] Manual Reactivate WA: {client.get('name')} to {wa_num}")
                    
                    # V3.3.9: Use mode="batch" and list format as wa-bridge requires
                    temp_wa = f"wa_react_{int(time.time())}_{random.randint(100,999)}.json"
                    temp_path = os.path.join(TEMP_FOLDER, temp_wa)
                    with open(temp_path, "w") as f:
                        json.dump([{"to": wa_num, "msg": wa_msg}], f)
                    spawn_wa_worker(mode="batch", task_file=temp_path)
                except Exception as e: 
                    print(f"Failed to queue WA reactivate: {e}")
                
            return {"status": "ok", "msg": msg}
        else:
            add_log(client['name'], 'error', f'Gagal Reaktivasi: {msg}')
            return {"status": "error", "msg": msg}

    except Exception as e:
        err_info = traceback.format_exc()
        with open("reaktivasi_error.txt", "w") as f:
            f.write(err_info)
        return {"status": "error", "msg": str(e)}

@app.route('/api/billing/client/<client_id>/activate', methods=['POST'])
def billing_activate_client(client_id):
    """Manual activation - restore from ISOLIR state"""
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json or {}
    selected_profile = data.get('profile')
    
    res = reaktivasi_client_core(client_id, selected_profile)
    return jsonify(res)


@app.route('/api/billing/client/<client_id>/isolir', methods=['POST'])
def billing_isolir_client(client_id):
    """Manual isolir (emergency)"""
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    db_data = load_db()
    client = next((c for c in db_data.get('clients', []) if str(c.get('id')) == str(client_id)), None)
    
    if not client:
        return jsonify({"error": "Client not found"}), 404
    
    pppoe_user = client.get('credentials', {}).get('pppoe_user')
    ip_addr = client.get('ip')
    router_id = client.get('managed_by', 'server_utama')
    
    success = False
    msg = ""
    
    # Persiapan Profil Isolir
    settings = load_settings()
    isolir_profile = settings.get('billing', {}).get('isolir_profile', 'ISOLIR')
    ensure_isolir_profile(router_id)

    if pppoe_user:
        # PPPoE Mode: Check if Radius or Local
        is_radius = client.get('mode') == 'pppoe_radius'
        
        if is_radius:
            # PPPoE RADIUS Mode: Use Address List method (IP-based)
            target_ip = None
            try:
                conn = get_router_connection(router_id)
                if conn:
                    api = conn.get_api()
                    ppp_act = api.get_resource('/ppp/active').get(name=pppoe_user)
                    if ppp_act: target_ip = ppp_act[0].get('address')
                    conn.disconnect()
            except: pass
            
            if not target_ip: target_ip = client.get('ip')

            if target_ip and target_ip != '-':
                res = add_to_address_list(target_ip, isolir_profile, router_id, comment=f"{isolir_profile}_{client['name']}")
                if res.get('status') == 'ok':
                    time.sleep(1)
                    kick_pppoe_user(pppoe_user, router_id)
                    success = True
                    msg = "PPPoE RADIUS isolated (added to address-list)"
                else:
                    msg = res.get('msg', 'Failed to add RADIUS IP to address-list')
            else:
                msg = "No active session or valid IP found for RADIUS client"
        else:
            # Local PPPoE: Change profile
            current_profile_real = get_pppoe_current_profile(pppoe_user, router_id) or client.get('packet_name', 'default')
            res = change_pppoe_profile(pppoe_user, isolir_profile, router_id)
            if res.get('status') == 'ok':
                time.sleep(1)
                kick_pppoe_user(pppoe_user, router_id)
                if 'billing' not in client: client['billing'] = {}
                # [BUG FIX V3.1.9] Hanya simpan profil asli jika profil saat ini BUKAN isolir
                if isolir_profile.upper() not in current_profile_real.upper():
                    client['billing']['original_profile'] = current_profile_real
                success = True
                msg = f"PPPoE Client isolated"
            else:
                msg = res.get('msg', 'Failed to change profile')
    elif ip_addr and ip_addr != '-':
        # Static IP Mode: Add to ISOLIR list
        res = add_to_address_list(ip_addr, isolir_profile, router_id, comment=f"{isolir_profile}_{client['name']}")
        if res.get('status') == 'ok':
            success = True
            msg = "Static IP Client isolated (added to ISOLIR list)"
        else:
            msg = res.get('msg', 'Failed to add to address-list')
    else:
        return jsonify({"error": "Client has no PPPoE user or valid IP"}), 400

    if success:
        if 'billing' not in client: client['billing'] = {}
        client['billing']['payment_status'] = 'overdue'
        client['billing']['isolir_date'] = get_local_now().strftime('%Y-%m-%d')
        client['status'] = 'isolir'
        save_db(db_data, preserve_live=False)
        add_log(client['name'], 'isolir', f'Manual isolir: {msg}')

        # SEND WHATSAPP NOTIFICATION (V3.3.9 Fix: Mode="batch" & Correct Keys)
        wa_isolir_enabled = settings.get('wa_isolir_enabled', True)
        wa_num = client.get('wa_number') or client.get('whatsapp_number') or client.get('phone') or (client.get('billing', {}) if isinstance(client.get('billing'), dict) else {}).get('wa_number')
        if wa_num and wa_isolir_enabled:
            try:
                # Use Template from Settings (V3.3.9)
                msg_tpl = settings.get('wa_template_isolir', "Yth. {name}, layanan internet Anda diisolir sementara karena keterlambatan pembayaran. Silakan lakukan pembayaran agar layanan kembali normal.")
                
                # Dynamic Price for Template
                price_val = 0
                packet_name = client.get('packet_name', '')
                billing_profiles = settings.get('billing_profiles', {})
                for prof_name, prof_price in billing_profiles.items():
                    if prof_name.strip().lower() == packet_name.strip().lower():
                        price_val = prof_price; break
                
                # Replace placeholders
                msg_wa = msg_tpl.replace("{name}", client.get('name', 'Pelanggan'))\
                               .replace("{id}", str(client.get('id', '')).replace('client_', ''))\
                               .replace("{price}", str(price_val))\
                               .replace("{amount}", str(price_val))\
                               .replace("{expired}", get_local_now().strftime('%d-%m-%Y'))
                
                # Trace Log (V3.3.9 Debug)
                _append_wa_log(f"[TRACE] Manual Isolir WA: {client.get('name')} to {wa_num}")
                
                # V3.3.9: Use mode="batch" and list format as wa-bridge requires
                temp_wa = f"wa_isolir_{int(time.time())}_{random.randint(100,999)}.json"
                temp_path = os.path.join(TEMP_FOLDER, temp_wa)
                with open(temp_path, "w") as f:
                    # JSON keys MUST be "to" and "msg"
                    json.dump([{"to": wa_num, "msg": msg_wa}], f)
                spawn_wa_worker(mode="batch", task_file=temp_path)
            except Exception as e: 
                print(f"[MANUAL_ISOLIR_WA_ERR] {e}")
                _append_wa_log(f"[ERROR] WA Isolir Fail: {str(e)}")

        return jsonify({"status": "ok", "msg": msg})
    else:
        return jsonify({"status": "error", "msg": msg})


@app.route('/api/finance/summary')
def fin_summary():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    # Ensuring carryover is generated if we entered a new month
    process_monthly_carryover()

    data = load_finance()
    txs = data.get('transactions', [])
    
    # Filter for current month using local time
    now = get_local_now()
    cur_month = now.strftime('%Y-%m')
    
    income_pure = 0    # Only real income this month (excluding carryover)
    expense_pure = 0   # Real expenses this month
    monthly_carry = 0  # Net balance carried from ALL previous months
    paid_clients = []
    
    for t in txs:
        t_date = t.get('date', '')
        
        # We only need to check current month transactions because 
        # the 'balance_carryover' for this month already summarizes everything before it.
        if t_date.startswith(cur_month):
            amt = int(t.get('amount', 0))
            t_type = t.get('type')
            t_cat = t.get('category', '')

            if t_cat == 'balance_carryover':
                # This entry IS the total balance from January to Last Month
                if t_type == 'income': monthly_carry += amt
                else: monthly_carry -= amt
            else:
                # Real transaction this month
                if t_type == 'income':
                    income_pure += amt
                    if t_cat == 'wifi_payment' and t.get('client_id'):
                        paid_clients.append(t.get('client_id'))
                elif t_type == 'expense':
                    expense_pure += amt
            
    # Total Global Cash = (Balance from the past) + (Activity this month)
    total_balance = monthly_carry + income_pure - expense_pure
    
    return jsonify({
        "income": income_pure,              # Card 1: Pure income this month
        "expense": expense_pure,            # Card 2: Total expense this month
        "balance": income_pure - expense_pure,   # Card 3: Monthly surplus/deficit
        "total_balance": total_balance,      # Card 4: Global Cash in hand (Physical Wallet)
        "month": cur_month,
        "paid_clients": list(set(paid_clients))
    })

@app.route('/api/finance/data')
def api_finance_data():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    return jsonify(load_finance())

def process_monthly_carryover():
    """
    Checks if current month has a carryover transaction.
    If not, calculate ALL previous balance and insert it.
    """
    with db_lock: # Reuse DB lock to prevent race conditions roughly
        data = load_finance()
        txs = data.get('transactions', [])
        
        now = get_local_now()
        cur_month_str = now.strftime('%Y-%m') # e.g. 2026-02
        
        # 1. Check if carryover exists for this month
        exists = any(t.get('category') == 'balance_carryover' and t.get('date', '').startswith(cur_month_str) for t in txs)
        if exists: return
        
        # 2. If valid previous data exists (don't run on brand new empty system unless needed?)
        # Actually run it.
        
        # Perhitungan Saldo Awal menggunakan logika "Waterflow" (Akuntansi Bertahap)
        # Mencari saldo pindahan terakhir sebelum bulan ini
        latest_carry = None
        for t in txs:
            if t.get('category') == 'balance_carryover':
                t_date = t.get('date', '')
                if t_date < cur_month_str + "-01":
                    if not latest_carry or t_date > latest_carry['date'] or (t_date == latest_carry['date'] and t['id'] > latest_carry['id']):
                        latest_carry = t

        prev_balance = 0
        has_history = False
        cutoff_date = ""
        cutoff_id = ""

        if latest_carry:
            has_history = True
            cutoff_date = latest_carry['date']
            cutoff_id = latest_carry['id']
            amt = int(latest_carry.get('amount', 0))
            if latest_carry.get('type') == 'income': prev_balance = amt
            else: prev_balance = -amt

        for t in txs:
            if t.get('category') == 'balance_carryover': continue
            
            t_date = t.get('date', '')
            if t_date < cur_month_str + "-01":
                # Tambah transaksi yang terjadi setelah atau pada hari yang sama dengan saldo pindahan terakhir
                if not latest_carry or t_date > cutoff_date or (t_date == cutoff_date and t['id'] > cutoff_id):
                    has_history = True
                    amt = int(t.get('amount', 0))
                    if t.get('type') == 'income': prev_balance += amt
                    else: prev_balance -= amt
            
        # 3. Create Carryover Transaction
        # Get Previous Month Name (Indonesian)
        # Python date math
        first_of_this_month = now.replace(day=1)
        last_month_obj = first_of_this_month - timedelta(days=1)
        months_id = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
        prev_month_name = months_id[last_month_obj.month - 1]
        
        note = f"Sisa Saldo Bulan {prev_month_name} {last_month_obj.year}"
        
        # Only carry over if positive? Or negative too? 
        # User said "keuangan masuk" (Income). If negative, it's technically expense/debt?
        # Let's map it: Positive -> Income. Negative -> Expense.
        
        tx_type = 'income'
        final_amt = prev_balance
        if prev_balance < 0:
            tx_type = 'expense'
            final_amt = abs(prev_balance)
            note = f"Minus Saldo Bulan {prev_month_name} {last_month_obj.year}"
            
        if final_amt == 0: return # Nothing to carry over
            
        new_tx = {
            "id": str(int(time.time() * 1000)),
            "date": cur_month_str + "-01", # 1st of current month
            "type": tx_type,
            "category": "balance_carryover",
            "amount": final_amt,
            "note": note,
            "user": "system"
        }
        
        data.setdefault('transactions', []).append(new_tx)
        
        # 4. CLEANUP (Keep Max 5 Months History strictly)
        # Current Month is kept. We keep 5 months back.
        # Logic: Go back 6 months to find cutoff.
        d = first_of_this_month
        for _ in range(6):
            d = d - timedelta(days=1)
            d = d.replace(day=1)
        cutoff_str = d.strftime('%Y-%m') # '%Y-%m-01' is safer comparison if string
        
        # Filter: Keep if date >= cutoff_str
        # Note: '2025-08' > '2025-07' works for strings
        original_len = len(data['transactions'])
        data['transactions'] = [t for t in data['transactions'] if t.get('date', '') >= cutoff_str]
        removed_count = original_len - len(data['transactions'])
        
        save_finance(data)
        print(f"[FINANCE] Generated Auto-Carryover for {cur_month_str}. Cleaned {removed_count} old transactions (Cutoff: {cutoff_str})")

def prettify_money(v): return "Rp " + "{:,}".format(v)

@app.route('/api/finance/history')
def fin_history():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    data = load_finance()
    # Return reverse chronological order (newest first)
    hist = sorted(data.get('transactions', []), key=lambda x: x.get('id', ''), reverse=True)
    return jsonify(hist)

@app.route('/api/finance/last_transaction/<client_id>')
def fin_last_tx(client_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    data = load_finance()
    # Find newest transaction for this client
    txs = [t for t in data.get('transactions', []) if str(t.get('client_id')) == str(client_id)]
    if not txs:
        return jsonify({"status": "empty", "msg": "Belum ada riwayat pembayaran."})
    
    # Sort descending by ID (timestamp)
    last = sorted(txs, key=lambda x: x.get('id', ''), reverse=True)[0]
    return jsonify({"status": "found", "data": last})

@app.route('/api/finance/add', methods=['POST'])
@app.route('/api/finance/transaction', methods=['POST'])
def fin_add_tx():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    payload = request.json or {}
    type_ = payload.get('type', 'income') # Default income for legacy
    cat = payload.get('category', 'wifi_payment') # Default wifi for payment modal
    amt = int(payload.get('amount', 0))
    note = payload.get('note', '')
    date_ = payload.get('date') or get_local_now().strftime('%Y-%m-%d')
    client_id = payload.get('client_id')
    duration = int(payload.get('duration', 1)) 
    auto_reactivate = payload.get('auto_reactivate', True)
    
    if amt <= 0:
        return jsonify({"error": "Invalid amount"}), 400
        
    data = load_finance()
    tx_id = payload.get('id')
    is_update = False
    
    if tx_id:
        # Check if transaction exists
        for t in data.get('transactions', []):
            if str(t.get('id')) == str(tx_id):
                t.update({
                    "date": date_,
                    "type": type_,
                    "category": cat,
                    "amount": amt,
                    "note": note,
                    "client_id": client_id,
                    "user": check_auth(request)
                })
                new_tx = t
                is_update = True
                break
    
    if not is_update:
        new_tx = {
            "id": str(tx_id) if tx_id else str(int(time.time() * 1000)),
            "date": date_,
            "type": type_,
            "category": cat,
            "amount": amt,
            "note": note,
            "client_id": client_id,
            "user": check_auth(request)
        }
        data.setdefault('transactions', []).append(new_tx)
    
    save_finance(data)
    
    # AUTO-REACTIVATION & PAID_UNTIL UPDATE (ONLY FOR NEW TRANSACTIONS)
    if is_update:
        return jsonify({"status": "ok", "id": new_tx['id']})

    try:
        if (cat == 'wifi_payment' or cat == 'Pembayaran WiFi') and client_id:
            topology = db.load_full_topology()
            client = next((c for c in topology['clients'] if c['id'] == client_id), None)
            
            if client:
                # 1. Update paid_until (Top Level)
                now = get_local_now()
                try:
                    current_expiry = datetime.strptime(client.get('paid_until', ''), '%Y-%m-%d')
                except:
                    current_expiry = now

                # --- PREMIUM PRICE VERIFICATION ---
                billing_config = load_billing_config()
                profiles = billing_config.get('billing_profiles', {})
                pkt_name = client.get('packet_name')
                if pkt_name and pkt_name in profiles:
                    price_per_month = int(profiles[pkt_name])
                    expected = price_per_month * duration
                    if amt != expected:
                        diff = amt - expected
                        status_label = "Lebih" if diff > 0 else "Kurang"
                        sign = "+" if diff > 0 else "-"
                        # Format as IDR (100.000)
                        fmt_diff = "{:,.0f}".format(abs(diff)).replace(',', '.')
                        mismatch_tag = f" (!! MISMATCH: {status_label} {sign}Rp {fmt_diff})"
                        new_tx['note'] = new_tx.get('note', '') + mismatch_tag
                        # Note: This mismatch tag is for internal Finance logs and will be filtered in printReceipt.
                
                # --- AUTO ARREARS / SMART BILLING LOGIC (V3.6) Pre-check ---
                # Load settings early to determine effective duration vs debt payoff
                settings_tmp = load_settings()
                m_arrs_tmp = settings_tmp.get('manual_arrears', [])
                client_ma_tmp = next((ma for ma in m_arrs_tmp if str(ma.get('client_name')).strip().lower() == str(client['name']).strip().lower()), None)
                
                effective_duration = duration
                remaining_amt_tmp = amt
                
                if client_ma_tmp:
                    old_debt_tmp = int(client_ma_tmp.get('amount', 0))
                    if old_debt_tmp > 0:
                        payoff_tmp = min(remaining_amt_tmp, old_debt_tmp)
                        client_ma_tmp['amount'] = old_debt_tmp - payoff_tmp
                        remaining_amt_tmp -= payoff_tmp
                        if payoff_tmp > 0:
                            client_ma_tmp['desc'] = f"Hutang dibayar ({date_})"
                        
                        # If payment was fully used for debt payoff, don't extend service period
                        if remaining_amt_tmp <= 0:
                            effective_duration = 0

                # LOGIC BARU (V3.1.2): Mode Global vs 30 Hari
                billing_config = load_billing_config()
                b_mode = billing_config.get('billing_mode', 'monthly') # monthly=Global, cyclic=30 Hari
                
                if effective_duration <= 0:
                    new_expiry = current_expiry
                elif b_mode == 'cyclic':
                    # MODE 30 HARI - OPSI B (AKUMULASI FLEXIBLE)
                    # Jika sudah telat (expired), mulai dari hari ini (now)
                    # Jika masih aktif, akumulasi (tambah dari expiry lama)
                    base_date = max(now, current_expiry)
                    
                    # Tambah per 30 hari
                    new_expiry = base_date + timedelta(days=30 * effective_duration)
                else:
                    # MODE GLOBAL (BULANAN)
                    # Logic: Mundur/Maju berdasarkan tgl jatuh tempo setting
                    b_day = billing_config.get('default_billing_day', 20)
                    
                    # 1. Tentukan "Base Month" (Bulan Dasar Mulai)
                    if current_expiry and current_expiry > datetime(2000,1,1):
                        # Jika sudah ada tanggal expire, kita lihat apakah dia menunggak (sudah lewat tgl bayar)
                        # V3.1.8 FIX: Jika dia menunggak, start dari expiry lama agar melunasi bulan yang bolong.
                        # Tapi kita beri toleransi, jika expiry lama sudaaah sangat jauh (> 6 bulan), 
                        # mungkin reset dari 'now' saja agar tidak kaget bayar banyak bulan sekaligus.
                        if (now - current_expiry).days > 180:
                             start_from_year = now.year
                             start_from_month = now.month
                        else:
                             start_from_year = current_expiry.year
                             start_from_month = current_expiry.month
                    else:
                        # Klien baru atau reset data
                        if now.day > b_day:
                            # Jika bayar > tgl tagihan, berarti bayar untuk bulan DEPAN (karena bulan ini telat/sudah lewat)
                            base_month_idx = now.year * 12 + now.month
                            start_from_year = (base_month_idx // 12)
                            start_from_month = (base_month_idx % 12) + 1
                            if start_from_month > 12: 
                                start_from_month = 1; start_from_year += 1
                        else:
                            # Bayar untuk bulan berjalan
                            start_from_year = now.year
                            start_from_month = now.month

                    # 2. Hitung Target Akhir berdasarkan durasi
                    # Monthly mode: 1 bulan bayar = lunas sampai deadline bulan berikutnya.
                    # Contoh: Bayar Januari -> Lunas s/d 28 Februari (aman selama Februari, isolir awal Maret)
                    target_month = start_from_month + effective_duration
                    
                    target_year = start_from_year + (target_month - 1) // 12
                    target_month = (target_month - 1) % 12 + 1
                    
                    # 3. Tentukan Tanggal Akhir (Jatuh Tempo)
                    import calendar
                    last_day_in_target = calendar.monthrange(target_year, target_month)[1]
                    actual_day = min(b_day, last_day_in_target)
                    new_expiry = datetime(target_year, target_month, actual_day)

                client['paid_until'] = new_expiry.strftime('%Y-%m-%d')
                
                # Update Note with new expiry if not verbose
                if "Lunas s/d" not in new_tx['note']:
                    new_tx['note'] += f" (Lunas s/d {client['paid_until']})"
                    save_finance(data) # Resave with updated note



                # 2. Billing sub-object update
                billing = client.get('billing', {})
                if not billing: 
                    billing = {}
                    client['billing'] = billing

                billing['last_payment_date'] = date_

                # V3.6.1 FIX: Don't unconditionally reset status to 'paid' if they still have manual debt
                total_ma = 0
                for ma in settings_tmp.get('manual_arrears', []):
                    if (ma.get('client_name') or "").strip().upper() == (client.get('name') or "").strip().upper():
                        total_ma += int(ma.get('amount') or 0)
                
                if total_ma <= 0:
                    billing['payment_status'] = 'paid'
                    billing['overdue_months'] = 0
                else:
                    # If they still have debt, Keep status or set to 'partial'? 
                    # For now just don't reset it to 'paid' if it was already debt/isolir
                    if billing.get('payment_status') == 'paid':
                         # If it was paid but now they have manual debt (e.g. from adjustment), keep it paid but logically they have debt.
                         # But usually if they have debt, status should be 'debt'.
                         billing['payment_status'] = 'debt'
                
                # 3. Auto Activate if isolated
                was_reactivated = False
                if auto_reactivate and client.get('status') == 'isolir':
                    res_core = reaktivasi_client_core(client_id, send_notif=False)
                    if res_core.get('status') == 'ok':
                        was_reactivated = True
                        # No need for manual add_log or client status update here as it's handled by core
                
                # --- AUTO ARREARS / SMART BILLING LOGIC (V3.6) Final ---
                try:
                    # Use existing settings_tmp, client_ma_tmp, effective_duration, and remaining_amt_tmp
                    ps = settings_tmp 
                    b_profs = ps.get('billing_profiles', {})
                    p_name = (client.get('packet_name') or "").strip().lower()
                    p_price = 0
                    for pn, pv in b_profs.items():
                        if pn.strip().lower() == p_name:
                            p_price = int(pv); break
                    
                    if p_price > 0:
                        # Logic: Only calculate new arrears for the extra months we are buying
                        expected = p_price * effective_duration
                        diff = expected - remaining_amt_tmp
                        
                        if diff != 0:
                            if client_ma_tmp:
                                old_val = int(client_ma_tmp.get('amount', 0))
                                client_ma_tmp['amount'] = old_val + diff
                                client_ma_tmp['desc'] = f"Updated via Smart Billing ({date_})"
                            else:
                                ps.setdefault('manual_arrears', []).append({
                                    "id": str(int(time.time() * 1000)),
                                    "client_name": client['name'],
                                    "amount": diff,
                                    "desc": f"Smart Billing Adjustment ({date_})"
                                })
                            
                            # V3.4.1 FIX: Load FRESH settings for arrears update to avoid Lost Update bug
                            # (The user might have changed Web Title while this payment was processing)
                            _fresh_s = _load_settings_raw()
                            _fresh_s['manual_arrears'] = ps['manual_arrears']
                            save_settings(_fresh_s)
                            
                            # Finance Note Update for transparency
                            if diff < 0: # Overpaid
                                new_tx['note'] += f" (Kelebihan: {abs(diff)})"
                            else: # Underpaid
                                new_tx['note'] += f" (Kekurangan: {diff})"
                            save_finance(data)
                except Exception as e:
                    print(f"[SMART BILLING ERROR] {e}")

                # Save changes to SQLite Database (Targeted Atomic Update V3.3.9)
                apply_bulk_updates([{
                    'id': client['id'],
                    'paid_until': client.get('paid_until'),
                    'billing': client.get('billing'),
                    'status': client.get('status'),
                    'packet_name': client.get('packet_name')
                }])

                # 4. SEND WHATSAPP PAYMENT NOTIFICATION (V3.3)
                try:
                    s_wa = load_settings()
                    wa_pay_enabled = s_wa.get('wa_payment_notif_enabled')
                    wa_react_enabled = s_wa.get('wa_reactivate_enabled', True)

                    if (was_reactivated and wa_react_enabled) or (not was_reactivated and wa_pay_enabled):
                        # Check phone
                        p_check = client.get('phone') or client.get('whatsapp_number') or client.get('wa_number') or billing.get('wa_number')
                        if p_check:
                            # Basic cleanup
                            import re
                            p_clean = re.sub(r'\D', '', str(p_check))
                            if p_clean.startswith('0'): p_clean = "62" + p_clean[1:]
                            
                            if p_clean:
                                # Format price and expiry for template
                                f_amt = "{:,}".format(int(amt)).replace(",", ".")
                                f_exp = client.get('paid_until', '')

                                if was_reactivated:
                                    tpl = s_wa.get('wa_template_reactivate')
                                    if not tpl:
                                        tpl = "Halo {name}, pembayaran Rp {amount} telah diterima dan layanan Anda telah diaktifkan kembali hingga {expired}. Terima kasih."
                                else:
                                    tpl = s_wa.get('wa_template_payment')
                                    if not tpl:
                                        tpl = "Terima kasih, pembayaran wifi a.n {name} sebesar Rp {amount} pada {date} telah diterima."
                                    
                                # Smart Expired Date for Payment Notif
                                f_exp = client.get('paid_until', '')
                                if f_exp and '-' in f_exp:
                                    try:
                                        dt_p = datetime.strptime(f_exp, '%Y-%m-%d')
                                        m_p = get_month_name(dt_p.month, s_wa.get('language', 'id'))
                                        f_exp = f"{dt_p.day} {m_p} {dt_p.year}"
                                    except: pass

                                f_id = str(client.get('id', '')).replace('client_', '')
                                final_msg = tpl.replace('{name}', client.get('name', 'Pelanggan'))\
                                               .replace('{id}', f_id)\
                                               .replace('{amount}', f_amt)\
                                               .replace('{price}', f_amt)\
                                               .replace('{date}', date_)\
                                               .replace('{expired}', f_exp)
                                
                                # Spawn worker in "test" mode (single message)
                                spawn_wa_worker(mode="test", target=p_clean, message=final_msg)
                            # Log removed
                            pass
                except Exception as wa_e:
                    print(f"[PAYMENT_NOTIF_ERR] {wa_e}")

    except Exception as e:
        print(f"[FINANCE_HOOK] Error: {e}")

    # V3.3.9 Logic: Return newest expiry for immediate UI refresh
    db_u = load_db(force_refresh=True)
    cl_upd = next((c for c in db_u.get('clients', []) if str(c.get('id')) == str(client_id)), {})
    return jsonify({"status": "ok", "id": new_tx['id'], "paid_until": cl_upd.get('paid_until')})

@app.route('/api/billing/client/<client_id>/update_expiry', methods=['POST'])
def update_client_expiry(client_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    data = request.json
    new_expiry = data.get('paid_until') # Expected format YYYY-MM-DD (converted from DD-MM-YYYY in frontend)
    
    if not new_expiry:
        return jsonify({"error": "Expiry date is required"}), 400
        
    db_data = load_db()
    found = False
    for c in db_data.get('clients', []):
        if str(c.get('id')) == str(client_id):
            # Targeted Update (V3.3.9 FIX): Use apply_bulk_updates to bypass save_db protection
            apply_bulk_updates([{
                'id': client_id,
                'paid_until': new_expiry
            }])
            found = True; break
            
    if found:
        add_log("SYSTEM", "system", f"Manual Expiry Update: {client_id} -> {new_expiry}")
        return jsonify({"status": "ok", "msg": f"Masa aktif berhasil diupdate ke {new_expiry}", "paid_until": new_expiry})
    return jsonify({"error": "Client not found"}), 404

@app.route('/api/finance/transaction/<tx_id>', methods=['DELETE'])
def fin_delete_tx(tx_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    data = load_finance()
    txs = data.get('transactions', [])
    
    # Find transaction to log it
    target_tx = next((t for t in txs if str(t.get('id')) == str(tx_id)), None)
    if not target_tx:
        return jsonify({"error": "Transaction not found"}), 404
        
    # Filter out the transaction
    new_txs = [t for t in txs if str(t.get('id')) != str(tx_id)]
    
    data['transactions'] = new_txs
    save_finance(data)
    
    # Log deletion
    desc = target_tx.get('note', '') or target_tx.get('category', 'tx')
    amt = target_tx.get('amount', 0)
    client_id = target_tx.get('client_id')
    add_log("SYSTEM", "system", f"Pembayaran Dibatalkan: {desc} (Rp {amt})")
    
    # --- SMART EXPIRY CHECK AFTER DELETE (Bug Fix: Phantom PAID status) ---
    if client_id:
        try:
            db_data = load_db()
            found_client = False
            for c in db_data.get('clients', []):
                if str(c.get('id')) == str(client_id):
                    # Kita dapatkan klien yang transaksinya barusan dihapus
                    now_str = get_local_now().strftime('%Y-%m-%d')
                    expiry = c.get('paid_until', '')
                    
                    if 'billing' not in c:
                        c['billing'] = {}
                        
                    # 1. Jika tanggal expiry sudah lewat ATAU kosong, set status jadi nunggak
                    if not expiry or expiry < now_str:
                        c['billing']['payment_status'] = 'overdue'
                    # 2. Jika tanggal expiry hari ini atau ke depan, belum telat. Beri status belum terbayar 'unpaid' / biarkan sesuai rules.
                    else:
                        c['billing']['payment_status'] = 'unpaid'
                    
                    found_client = True
                    break
            
            if found_client:
                save_db(db_data, preserve_live=False)
                # Note: preserve_live=False assumes topology is mostly static or handled by thread
        except Exception as e:
            print(f"[DELETE TX HOOK ERR] {e}")
            
    return jsonify({"status": "ok", "msg": "Transaksi berhasil dihapus"})


@app.route('/api/telegram/test', methods=['POST'])
def test_telegram_connection():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json or {}
    token = data.get('token')
    chat_id = data.get('chat_id')
    
    if not token or not chat_id:
        return jsonify({"status": "error", "msg": "Token & Chat ID required"})
        
    try:
        msg = f"📢 NMS Connectivity Test\n\n✅ Bot Linked Successfully!\n✅ System: {SERVICE_NAME}"
        res = requests.post(f"https://api.telegram.org/bot{token}/sendMessage", 
                     data={'chat_id': chat_id, 'text': msg}, timeout=10)
        
        if res.status_code == 200:
            return jsonify({"status": "ok", "msg": "Test Message Sent!"})
        else:
            return jsonify({"status": "error", "msg": f"Telegram Error: {res.text}"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)})

@app.route('/api/backup/download/<filename>')
def download_backup_file(filename):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    # Security: basic sanitization
    filename = os.path.basename(filename) 
    backup_dir = os.path.join(SCRIPT_DIR, 'backups')
    return send_from_directory(backup_dir, filename, as_attachment=True)

@app.route('/api/backup/list')
def list_backups_api():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    backup_dir = os.path.join(SCRIPT_DIR, 'backups')
    if not os.path.exists(backup_dir): return jsonify([])
    
    files = []
    for f in os.listdir(backup_dir):
        if f.endswith('.zip'):
            fp = os.path.join(backup_dir, f)
            files.append({
                "name": f,
                "size": round(os.path.getsize(fp) / 1024, 1), # KB
                "date": datetime.fromtimestamp(os.path.getmtime(fp)).strftime('%Y-%m-%d %H:%M')
            })
    
    # Sort new first
    files.sort(key=lambda x: x['date'], reverse=True)
    return jsonify(files)

@app.route('/api/backup/trigger', methods=['POST'])
def trigger_manual_backup():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    # Run in thread with force=True
    threading.Thread(target=auto_backup_logic, kwargs={'force': True}).start()
    return jsonify({"status": "ok", "msg": "Backup process started in background"})

@app.route('/api/backup/delete/<filename>', methods=['DELETE'])
def delete_backup_file(filename):
    if check_auth(request) != 'admin': return jsonify({"error": "Unauthorized"}), 401
    
    filename = os.path.basename(filename)
    backup_dir = os.path.join(SCRIPT_DIR, 'backups')
    fp = os.path.join(backup_dir, filename)
    
    if os.path.exists(fp):
        try:
            os.remove(fp)
            return jsonify({"status": "ok", "msg": f"File {filename} deleted"})
        except Exception as e:
            return jsonify({"status": "error", "msg": str(e)})
    return jsonify({"status": "error", "msg": "File not found"}), 404

@app.route('/api/finance/transaction/<tx_id>', methods=['PUT'])
def fin_edit_tx(tx_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    payload = request.json or {}
    amt = payload.get('amount')
    note = payload.get('note')
    
    data = load_finance()
    txs = data.get('transactions', [])
    
    found = False
    for t in txs:
        if str(t.get('id')) == str(tx_id):
            if amt is not None: t['amount'] = int(amt)
            if note is not None: t['note'] = note
            found = True
            break
            
    if not found:
        return jsonify({"error": "Transaction not found"}), 404
        
    save_finance(data)
    return jsonify({"status": "ok"})





@app.route('/api/client/<client_id>/bypass', methods=['POST'])
def toggle_client_bypass(client_id):
    if not check_auth(request):
        return jsonify({"error": "Unauthorized"}), 401
        
    db_data = load_db()
    data = request.json or {}
    target_state = data.get('bypass', None) 
    
    found = False
    new_state = False
    
    for c in db_data.get("clients", []):
        if str(c.get('id')) == str(client_id):
            if target_state is not None:
                c['bypass_billing'] = bool(target_state)
            else:
                c['bypass_billing'] = not c.get('bypass_billing', False)
                
            new_state = c['bypass_billing']
            found = True
            break
            
    if not found:
        return jsonify({"error": "Client not found"}), 404
        
    save_db(db_data)
    return jsonify({"status": "ok", "bypass": new_state})


@app.route('/api/billing/client/<client_id>/enable', methods=['POST'])
def billing_enable(client_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    db_data = load_db()
    found = False
    for c in db_data.get("clients", []):
        if str(c.get('id')) == str(client_id):
            if 'billing' not in c: c['billing'] = {}
            c['billing']['enabled'] = True
            found = True; break
    if found: 
        save_db(db_data)
        return jsonify({"status": "ok"})
    return jsonify({"error": "Not found"}), 404

@app.route('/api/billing/client/<client_id>/disable', methods=['POST'])
def billing_disable(client_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    db_data = load_db()
    found = False
    for c in db_data.get('clients', []):
        if str(c.get('id')) == str(client_id):
            if 'billing' not in c: c['billing'] = {}
            c['billing']['enabled'] = False
            found = True; break
    if found:
        save_db(db_data)
        return jsonify({"status": "ok"})
    return jsonify({"error": "Not found"}), 404

@app.route('/api/billing/bulk_toggle', methods=['POST'])
def billing_bulk_toggle():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    data = request.json
    enabled = data.get('enabled', False)
    db_data = load_db()
    for c in db_data.get('clients', []):
        if 'billing' not in c: c['billing'] = {}
        c['billing']['enabled'] = enabled
    save_db(db_data)
    return jsonify({"status": "ok", "msg": f"Billing {'Enabled' if enabled else 'Disabled'} for all clients"})

@app.route('/api/billing/client/<client_id>/settings', methods=['POST'])
def billing_settings(client_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    db_data = load_db()
    data = request.json
    for c in db_data.get('clients', []):
        if str(c.get('id')) == str(client_id):
            if 'billing' not in c: c['billing'] = {}
            if 'billing_day' in data: c['billing']['billing_day'] = int(data['billing_day'])
            if 'price' in data: c['billing']['price'] = int(data['price'])
            save_db(db_data); return jsonify({"status": "ok"})
    return jsonify({"error": "Not found"}), 404

@app.route('/api/billing/run_notify_manual', methods=['GET', 'POST'])
def api_billing_run_notify_manual():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    try:
        data = request.get_json(silent=True) or {}
        mode = data.get('template_mode', 'auto')
        run_billing_check(notify_only=True, force=True, template_mode=mode)
        return jsonify({"status": "ok", "msg": f"Proses pengiriman notifikasi manual dimulai (Mode: {mode})."})
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)})

@app.route('/api/billing/today_schedule')
def api_billing_today_schedule():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    import calendar
    db = load_db()
    now = get_local_now()
    today_str = now.strftime('%Y-%m-%d')
    curr_period = now.strftime('%m-%Y')
    
    cfg_billing = load_billing_config()
    default_day = cfg_billing.get('default_billing_day', 20)
    grace_period = cfg_billing.get('grace_period_days', 3)
    wa_pre_isolir_days = cfg_billing.get('wa_pre_isolir_days', 2)
    wa_end_month_active = cfg_billing.get('wa_end_month_enabled', True)
    wa_pre_isolir_active = cfg_billing.get('wa_pre_isolir_enabled', True)
    
    settings_root = _load_settings_raw()
    manual_arrears_list = settings_root.get('manual_arrears', [])

    schedule = []
    db_changed = False
    
    for client in db.get("clients", []):
        billing = client.get('billing', {})
        if not billing.get('enabled', False): continue
        if client.get('bypass_billing', False): continue
        
        # Proactive Sync from Finance (V3.1.4) - RESTORED
        if sync_billing_from_finance(client, db):
            db_changed = True
            billing = client.get('billing', {})
        
        paid_until_str = client.get('paid_until')
        payment_status = billing.get('payment_status', 'unpaid')
        
        # [BUG FIX V3.3.2] Dynamic Expiry Check for Dashboard at 00:00 before 09:00 engine
        if paid_until_str:
            try:
                exp_date = datetime.strptime(paid_until_str, '%Y-%m-%d').date()
                if exp_date <= now.date(): payment_status = 'unpaid'
            except: pass
            
        # Also check Manual Arrears
        manual_arr_val = sum(int(ma.get('amount', 0)) for ma in manual_arrears_list if ma.get('client_name') == client.get('name'))
        
        # SKIP TRULY PAID CLIENTS
        if payment_status == 'paid' and manual_arr_val == 0: continue
        
        track = billing.get('wa_sent_track', {})
        
        due_date_obj = None
        if paid_until_str:
            try:
                due_date_obj = datetime.strptime(paid_until_str, '%Y-%m-%d')
            except: pass
            
        if not due_date_obj:
            b_day = billing.get('billing_day') or default_day
            due_date_obj = calculate_due_date(now.year, now.month, b_day)
            
        days_overdue = (now - due_date_obj).days
        
        # Determine Potential Reasons for Today
        reasons = []
        
        # 1. REMOVED: H-3 Reminder (Per User Request)
        
        # 2. Due Date / EOM
        if wa_end_month_active:
             b_mode = cfg_billing.get('billing_mode', 'monthly')
             last_day_of_month = calendar.monthrange(now.year, now.month)[1]
             if b_mode == 'cyclic' and days_overdue == 0:
                 reasons.append({"type": "Jatuh Tempo", "key": "eom"})
             elif b_mode == 'monthly' and now.day == last_day_of_month:
                 reasons.append({"type": "Akhir Bulan", "key": "eom"})
                 
        # 3. Pre-Isolation
        if wa_pre_isolir_active:
            trigger_day = (grace_period + 1) - wa_pre_isolir_days
            if days_overdue == trigger_day:
                reasons.append({"type": f"H-{wa_pre_isolir_days} Sebelum Isolir", "key": "pre_isolir"})
                
        # 4. Isolation
        # Theoretical isolation is when grace period is exactly hit today
        if days_overdue == grace_period + 1:
            reasons.append({"type": "Auto Isolir", "key": "isolir"})

        # PERSISTENCE LOGIC
        performed_today = []
        # REMOVED: H-3 Reminder Tracking (Per User Request)
        if track.get('eom') == curr_period and days_overdue >= 0: performed_today.append({"type": "Notif Tagihan", "key": "eom"})
        if track.get('pre_isolir') == curr_period and days_overdue >= ((grace_period + 1) - wa_pre_isolir_days): 
            performed_today.append({"type": "Peringatan Isolir", "key": "pre_isolir"})
        
        # Specific check for Today's Isolation
        if billing.get('isolir_date') == today_str:
            performed_today.append({"type": "Auto Isolir", "key": "isolir"})

        # Combine items
        all_matches = reasons + performed_today
        seen_keys = set()
        unique_matches = []
        for am in all_matches:
            if am['key'] not in seen_keys:
                unique_matches.append(am)
                seen_keys.add(am['key'])

        phone_num = client.get('phone') or client.get('whatsapp_number') or client.get('wa_number') or billing.get('wa_number') or '-'
        
        for m in unique_matches:
            status = "Menunggu"
            
            # CHECK FOR COMPLETION / RESOLUTION
            if payment_status == 'paid' and manual_arr_val == 0:
                status = "Selesai / Lunas"
            elif client.get('status') == 'online' and m['key'] == 'isolir':
                status = "Selesai / Lunas"
            elif track.get(m['key']) == curr_period:
                status = "Terkirim"
            
            # Specific check for Terisolir status
            if status != "Selesai / Lunas" and m['key'] == 'isolir':
                if client.get('status') == 'isolir' or track.get('isolir_wa_sent'):
                    status = "Terisolir"
            
            schedule.append({
                "name": client.get('name', 'N/A'),
                "phone": phone_num,
                "reason": m['type'],
                "reason_key": m['key'],
                "status": status,
                "date": today_str
            })
            
    if db_changed:
        save_db(db)
            
    return jsonify({"date": now.strftime('%d %B %Y'), "schedule": schedule})

@app.route('/api/logs')
def get_logs_route():
    if not check_auth(request): return jsonify({"error":"Unauthorized"}), 401
    logs = []
    if os.path.exists(LOG_FILE):
        try:
            with open(LOG_FILE, 'r') as f: logs = json.load(f)
        except: pass
    return jsonify(logs)

@app.route('/api/logo', methods=['POST', 'DELETE'])
def handle_logo():
    """Handle logo upload and deletion"""
    # Check auth and role
    role = check_auth(request)
    if not role:
        return jsonify({"status": "error", "msg": "Unauthorized"}), 401
    if role != 'admin':
        return jsonify({"status": "error", "msg": "Admin only"}), 403
    
    if request.method == 'POST':
        # Upload logo
        if 'file' not in request.files:
            return jsonify({"status": "error", "msg": "No file"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"status": "error", "msg": "No file selected"}), 400
        
        # Validate file type
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'svg', 'webp'}
        ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if ext not in allowed_extensions:
            return jsonify({"status": "error", "msg": "Invalid file type. Allowed: PNG, JPG, GIF, SVG"}), 400
        
        # Check file size (max 500KB)
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        if file_size > 500 * 1024:
            return jsonify({"status": "error", "msg": "File too large. Max 500KB"}), 400
        
        # Save logo (always as logo.png for simplicity)
        try:
            logo_path = os.path.join(SCRIPT_DIR, 'static', 'logo.png')
            file.save(logo_path)
            return jsonify({"status": "ok"})
        except Exception as e:
            return jsonify({"status": "error", "msg": str(e)}), 500
    
    elif request.method == 'DELETE':
        # Delete logo
        try:
            logo_path = os.path.join(SCRIPT_DIR, 'static', 'logo.png')
            if os.path.exists(logo_path):
                os.remove(logo_path)
            return jsonify({"status": "ok"})
        except Exception as e:
            return jsonify({"status": "error", "msg": str(e)}), 500

@app.route('/api/qris', methods=['POST', 'DELETE'])
def handle_qris():
    """Handle QRIS upload and deletion"""
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403

    try:
        # Ensure photos directory exists
        photos_dir = os.path.join(SCRIPT_DIR, 'static', 'photos')
        if not os.path.exists(photos_dir):
            os.makedirs(photos_dir)
            
        qris_path = os.path.join(photos_dir, 'qris.jpg')

        if request.method == 'POST':
            if 'file' not in request.files:
                return jsonify({"status": "error", "message": "No file part"}), 400
            file = request.files['file']
            if file.filename == '':
                return jsonify({"status": "error", "message": "No selected file"}), 400
            
            # Simple validation
            ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
            if ext not in ['jpg', 'jpeg', 'png']:
                return jsonify({"status": "error", "message": "Format harus JPG/PNG"}), 400

            # Save directly as qris.jpg (automatic conversion by extension rename for simple usage)
            # ideally we use PIL but user asked for simple rename
            file.save(qris_path)
            return jsonify({"status": "ok"})
        
        elif request.method == 'DELETE':
            if os.path.exists(qris_path):
                os.remove(qris_path)
            return jsonify({"status": "ok"})
            
    except Exception as e:
        print(f"[QRIS] Error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


def get_router_connection(router_id):
    """Helper to get routeros_api connection"""
    try:
        if not routeros_api: return None # Library missing
        
        data = load_db()
        if router_id == 'server_utama':
            r_data = data.get("server")
        else:
            r_data = next((r for r in data.get("extra_routers", []) if r["id"] == router_id), None)
        
        if not r_data or not r_data.get("login"): return None
        login = r_data["login"]
        
        conn = routeros_api.RouterOsApiPool(
            login.get("host"), 
            username=login.get("user"), 
            password=login.get("pass"),
            port=int(login.get("port", 8728)),
            plaintext_login=True
        )
        return conn
    except Exception as e:
        print(f"[ERROR] Router Conn: {e}")
        return None

@app.route('/api/hotspot/profiles/<router_id>', methods=['GET'])
def get_hotspot_profiles(router_id):
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed or Lib Missing"}), 500
    
    try:
        api = conn.get_api()
        
        # 1. Get ALL Server Profiles first (to get DNS names)
        # Equivalent to /ip/hotspot/profile/print
        srv_profiles_res = api.get_resource('/ip/hotspot/profile').get()
        # Mapping: Profile Name -> DNS Name
        prof_dns_map = {}
        for p in srv_profiles_res:
            nm = p.get('name')
            if nm:
                dns = p.get('dns-name')
                if not dns: dns = p.get('hotspot-address')
                prof_dns_map[nm] = dns or ""

        # 2. Get ACTUAL Hotspot Servers
        # Equivalent to /ip/hotspot/print
        hotspots_res = api.get_resource('/ip/hotspot').get()
        
        servers = []
        server_dns_map = {}
        for hs in hotspots_res:
            name = hs.get('name')
            if name:
                servers.append(name)
                prof_name = hs.get('profile')
                server_dns_map[name] = prof_dns_map.get(prof_name, "")

        servers.sort()
        
        # 3. Get User Profiles
        usr_profiles_res = api.get_resource('/ip/hotspot/user/profile').get()
        usr_profiles = sorted(list(set([x.get('name') for x in usr_profiles_res if x.get('name')])))
        
        # 4. Get Address Pools
        pools_res = api.get_resource('/ip/pool').get()
        pools = sorted([x.get('name') for x in pools_res if x.get('name')])

        # 5. Get Queues & Queue Types
        queues_res = api.get_resource('/queue/simple').get()
        queues = sorted([x.get('name') for x in queues_res if x.get('name')])
        
        qtypes_res = api.get_resource('/queue/type').get()
        qtypes = sorted([x.get('name') for x in qtypes_res if x.get('name')])

        # 6. Get Firewall Address Lists
        addr_lists_res = api.get_resource('/ip/firewall/address-list').get()
        addr_lists = sorted(list(set([x.get('list') for x in addr_lists_res if x.get('list')])))

        return jsonify({
            "status": "ok",
            "server_profiles": servers,
            "server_dns_map": server_dns_map,
            "user_profiles": usr_profiles,
            "pools": pools,
            "queues": queues,
            "queue_types": qtypes,
            "address_lists": addr_lists
        })
    except Exception as e:
        print(f"[HOTSPOT] Profile Error: {e}")
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/profile/list/<router_id>', methods=['GET'])
def get_hotspot_profile_list(router_id):
    """Get detailed list of user profiles"""
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify([]), 500
    
    try:
        api = conn.get_api()
        res = api.get_resource('/ip/hotspot/user/profile').get()
        
        profiles = []
        for r in res:
            pid = r.get('.id') or r.get('id')
            profiles.append({
                ".id": pid,
                "id": pid,
                "profile_id": pid,
                "name": r.get('name'),
                "address-pool": r.get('address-pool', 'none'),
                "session-timeout": r.get('session-timeout', ''),
                "idle-timeout": r.get('idle-timeout', 'none'),
                "keepalive-timeout": r.get('keepalive-timeout', '00:02:00'),
                "status-autorefresh": r.get('status-autorefresh', '00:01:00'),
                "shared-users": r.get('shared-users', '1'),
                "rate-limit": r.get('rate-limit', ''),
                "add-mac-cookie": r.get('add-mac-cookie', 'false'),
                "mac-cookie-timeout": r.get('mac-cookie-timeout', '3d 00:00:00'),
                "address-list": r.get('address-list', ''),
                "parent-queue": r.get('parent-queue', 'none'),
                "queue-type": r.get('queue-type', 'default-small'),
                "insert-queue-before": r.get('insert-queue-before', 'first'),
                "status": r.get('status', '')
            })
        return jsonify(profiles)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/profile/add', methods=['POST'])
def add_hotspot_profile():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    name = data.get('name')
    rate_limit = data.get('rate_limit')
    shared_users = data.get('shared_users')
    session_timeout = data.get('session_timeout')
    idle_timeout = data.get('idle_timeout')
    keepalive_timeout = data.get('keepalive_timeout')
    status_autorefresh = data.get('status_autorefresh')
    address_pool = data.get('address_pool')
    add_mac_cookie = data.get('add_mac_cookie')
    mac_cookie_timeout = data.get('mac_cookie_timeout')
    address_list = data.get('address_list')
    
    # Queue fields
    parent_queue = data.get('parent_queue')
    queue_type = data.get('queue_type')
    insert_queue_before = data.get('insert_queue_before')

    if not name: return jsonify({"status": "error", "msg": "Name Required"}), 400
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res_prof = api.get_resource('/ip/hotspot/user/profile')
        
        payload = {'name': name}
        if rate_limit: payload['rate-limit'] = rate_limit
        if shared_users: payload['shared-users'] = str(shared_users)
        if session_timeout: payload['session-timeout'] = session_timeout
        if idle_timeout: payload['idle-timeout'] = idle_timeout
        if keepalive_timeout: payload['keepalive-timeout'] = keepalive_timeout
        if status_autorefresh: payload['status-autorefresh'] = status_autorefresh
        if address_pool: payload['address-pool'] = address_pool
        if add_mac_cookie is not None: payload['add-mac-cookie'] = 'yes' if add_mac_cookie else 'no'
        if mac_cookie_timeout: payload['mac-cookie-timeout'] = mac_cookie_timeout
        if address_list: payload['address-list'] = address_list

        # Queue fields
        if parent_queue: payload['parent-queue'] = parent_queue
        if queue_type: payload['queue-type'] = queue_type
        if insert_queue_before: payload['insert-queue-before'] = insert_queue_before
        
        res_prof.add(**payload)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/profile/remove', methods=['POST'])
def remove_hotspot_profile():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin': return jsonify({"error": "Admin only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    pid = data.get('id')
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res_prof = api.get_resource('/ip/hotspot/user/profile')
        res_prof.remove(id=pid)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/generate', methods=['POST'])
def generate_hotspot_users():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician':
        return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    qty = int(data.get('qty', 1))
    server = data.get('server')
    if not server or server == 'all': server = 'all'
    profile = data.get('profile')
    timelimit = data.get('timelimit')
    datalimit = data.get('datalimit')
    comment = data.get('comment', 'generated-nms')
    prefix = data.get('prefix', '')
    mode = data.get('mode', 'up') 
    char_type = data.get('char_type', 'rand')
    char_len = int(data.get('char_len', 6))
    
    import random, string
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "message": "Connection Failed"}), 500
    
    generated = []
    errors = []
    
    try:
        api = conn.get_api()
        res_user = api.get_resource('/ip/hotspot/user')
        
        # Char sets
        chars = string.digits
        if char_type == 'abc': chars = string.ascii_lowercase
        elif char_type == 'ABC': chars = string.ascii_uppercase
        elif char_type == 'mix': chars = string.ascii_lowercase + string.digits
        elif char_type == 'MIX': chars = string.ascii_letters + string.digits

        for i in range(qty):
            user_code = ''.join(random.choices(chars, k=char_len))
            username = f"{prefix}{user_code}"
            
            if mode == 'vc':
                password = ''.join(random.choices(chars, k=char_len))
            else:
                password = username
            
            payload = {
                'name': username,
                'password': password,
                'profile': profile,
                'comment': comment
            }
            
            if server and server != 'all': payload['server'] = server
            if timelimit: payload['limit-uptime'] = timelimit
            
            pbytes = parse_hotspot_limit_bytes(datalimit)
            if pbytes is not None: payload['limit-bytes-total'] = pbytes
            
            # Silently handle or log elsewhere
            pass
            try:
                res_user.add(**payload)
                # Log removed
                pass
                generated.append({
                    "name": username, 
                    "password": password, 
                    "profile": profile or "-", 
                    "limit_uptime": timelimit or "-",
                    "limit_bytes": datalimit or "-"
                })
            except Exception as e:
                # Log removed
                pass
                errors.append(f"Fail {username}: {e}")
                
        # Log removed
        pass
        
        if errors and not generated:
            return jsonify({
                "status": "error",
                "message": "Fail: " + (errors[0] if errors else "Unknown"),
                "errors": errors
            }), 400

        return jsonify({
            "status": "ok",
            "generated": len(generated),
            "errors": len(errors),
            "error_details": errors if errors else [],
            "users": generated
        })
    except Exception as e:
         print(f"[HS-GEN] FATAL ERROR: {e}")
         return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/users/<router_id>', methods=['GET'])
def get_hotspot_users(router_id):
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    comment_filter = request.args.get('comment')
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify([]), 500
    
    try:
        api = conn.get_api()
        res_user = api.get_resource('/ip/hotspot/user')
        
        if comment_filter:
            # Try exact match first
            res = res_user.get(comment=comment_filter)
            # If empty and high data volume, we should be careful. 
            # But we'll try basic filter if exact match fails.
            if not res: 
                # Limit fetch if no filter is possible or use manual filter defensively
                all_u = res_user.get()
                res = [u for u in all_u if comment_filter in u.get('comment', '')]
                # If too many items, limit to prevent UI crash? 
                # User requested 30 in UI, but API should return all or a reasonable max.
                if len(res) > 2000: res = res[:2000] # Safety cap
        else:
            # Fetch all but cap at 5000 for safety on branch nodes
            res = res_user.get()
            if len(res) > 5000: res = res[:5000] 
            
        users = []
        for r in res:
             mid = r.get('.id') or r.get('id')
             
             users.append({
                "id": mid, 
                "name": r.get('name'),
                "password": r.get('password'),
                "profile": r.get('profile'),
                "server": r.get('server', 'all'),
                "comment": r.get('comment', ''),
                "limit_uptime": r.get('limit-uptime', ''),
                "limit_bytes": r.get('limit-bytes-total') or r.get('limit-bytes-out') or '',
                "disabled": r.get('disabled') == 'true'
            })
        return jsonify(users)
    except Exception as e:
        print(f"[HS-USERS] Error: {e}")
        return jsonify([])
    finally:
        if conn: conn.disconnect()

@app.route('/api/hotspot/delete', methods=['POST'])
def delete_hotspot_users():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    ids = data.get('ids', [])
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    success = 0
    last_error = ""
    try:
        api = conn.get_api()
        res_user = api.get_resource('/ip/hotspot/user')
        
        for mid in ids:
            try:
                res_user.remove(id=mid)
                success += 1
            except Exception as e:
                # Capture the error to return it if needed
                last_error = str(e)
            
        return jsonify({"status": "ok", "deleted": success, "last_error": last_error})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/active/<router_id>', methods=['GET'])
def get_active_hotspot(router_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify([]), 500
    
    try:
        api = conn.get_api()
        # Ensure we always get a list and handle potential latency better
        res = api.get_resource('/ip/hotspot/active').get()
        
        # Normalize data for frontend to ensure keys exist and dashes are handled
        active = []
        for r in res:
            active.append({
                ".id": r.get('.id') or r.get('id'),
                "server": r.get('server', 'all'),
                "user": r.get('user', r.get('name', 'unknown')),
                "address": r.get('address', '-'),
                "uptime": r.get('uptime', '00:00:00'),
                "bytes-in": r.get('bytes-in', '0'),
                "bytes-out": r.get('bytes-out', '0')
            })
            
        return jsonify(active)
    except Exception as e:
        print(f"[HS-ACTIVE] Error: {e}")
        return jsonify([])
    finally:
        if conn: conn.disconnect()

@app.route('/api/hotspot/kick', methods=['POST'])
def kick_hotspot_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json
    router_id = data.get('router_id')
    mid = data.get('id')
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        api.get_resource('/ip/hotspot/active').remove(id=mid)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)})
    finally:
        conn.disconnect()

# ==============================================================================
#  PPPoE MANAGEMENT API
# ==============================================================================

@app.route('/api/pppoe/profiles/<router_id>', methods=['GET'])
def get_pppoe_profiles(router_id):
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        # Get PPP Profiles
        profiles_res = api.get_resource('/ppp/profile').get()
        profiles = []
        for p in profiles_res:
            if not p.get('name'): continue
            # Split combined queue types if necessary (handle ROS v6 legacy or manual combined inputs)
            q_raw = p.get('queue-type', '')
            q_up = q_raw
            q_dn = p.get('download-queue-type', '')
            
            if '/' in q_raw and not q_dn:
                parts = q_raw.split('/')
                q_up = parts[0]
                q_dn = parts[1] if len(parts) > 1 else ''

            profiles.append({
                "name": p.get('name'),
                "local_address": p.get('local-address', ''),
                "local-address": p.get('local-address', ''),
                "remote_address": p.get('remote-address', ''),
                "remote-address": p.get('remote-address', ''),
                "dns_server": p.get('dns-server', ''),
                "dns-server": p.get('dns-server', ''),
                "rate_limit": p.get('rate-limit', ''),
                "rate-limit": p.get('rate-limit', ''),
                "only_one": p.get('only-one', 'default'),
                "only-one": p.get('only-one', 'default'),
                "insert_queue_before": p.get('insert-queue-before', ''),
                "insert-queue-before": p.get('insert-queue-before', ''),
                "parent_queue": p.get('parent-queue', ''),
                "parent-queue": p.get('parent-queue', ''),
                "queue_type": q_up,
                "queue-type": q_up,
                "download_queue_type": q_dn,
                "download-queue-type": q_dn,
                "raw": p
            })
        
        # Sort by name
        profiles = sorted(profiles, key=lambda x: x['name'])

        # Fetch router identity for display in frontend dropdown
        try:
            identity_res = api.get_resource('/system/identity').get()
            router_identity = identity_res[0].get('name', '') if identity_res else ''
        except Exception:
            router_identity = ''

        return jsonify({
            "status": "ok",
            "profiles": profiles,
            "router_identity": router_identity
        })
    except Exception as e:
        print(f"[PPPoE] Profile Error: {e}")
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/pppoe/profile_options/<router_id>', methods=['GET'])
def get_pppoe_profile_options(router_id):
    """Return IP pools, queue names, and queue types for the profile creation form."""
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401

    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500

    try:
        api = conn.get_api()
        # IP Pools
        try:
            pools_raw = api.get_resource('/ip/pool').get()
            pools = [p.get('name') for p in pools_raw if p.get('name')]
        except Exception:
            pools = []
        # Queue simple list (for insert-before / parent)
        try:
            qs_raw = api.get_resource('/queue/simple').get()
            queues_simple = ['first', 'bottom'] + [q.get('name') for q in qs_raw if q.get('name')]
        except Exception:
            queues_simple = ['first', 'bottom']
        # Queue tree list (for parent)
        try:
            qt_raw = api.get_resource('/queue/tree').get()
            queues_tree = [q.get('name') for q in qt_raw if q.get('name')]
        except Exception:
            queues_tree = []
        # Queue types
        try:
            qtype_raw = api.get_resource('/queue/type').get()
            queue_types = [q.get('name') for q in qtype_raw if q.get('name')]
        except Exception:
            queue_types = ['default', 'ethernet-default', 'wireless-default', 'pcq-upload-default', 'pcq-download-default']

        # Fetch router identity for display in frontend dropdown
        router_identity = ''
        try:
            identity_res = api.get_resource('/system/identity').get()
            router_identity = identity_res[0].get('name', '') if identity_res else ''
        except Exception: pass

        return jsonify({
            "status": "ok",
            "pools": pools,
            "queues_simple": queues_simple,
            "queues_tree": queues_tree,
            "queue_types": queue_types,
            "router_identity": router_identity
        })
    except Exception as e:
        print(f"[PPPoE] profile_options Error: {e}")
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()


@app.route('/api/pppoe/profile/create', methods=['POST'])
def create_pppoe_profile():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role not in ('admin', 'technician'): return jsonify({"error": "Admin/Tech only"}), 403

    data = request.json
    router_id = data.get('router_id')
    name = data.get('name', '').strip()
    if not name:
        return jsonify({"status": "error", "msg": "Profile name required"}), 400

    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500

    try:
        api = conn.get_api()
        payload = {'name': name}
        q_up = (data.get('queue_type') or data.get('queue_type_upload') or '').strip()
        q_dn = (data.get('queue_type_download') or '').strip()
        
        optional_str = {
            'local-address': data.get('local_address', '').strip(),
            'remote-address': data.get('remote_address', '').strip(),
            'dns-server': data.get('dns_server', '').strip(),
            'rate-limit': data.get('rate_limit', '').strip(),
            'insert-queue-before': data.get('insert_queue_before', '').strip(),
            'parent-queue': data.get('parent_queue', '').strip(),
        }
        
        if q_up and q_dn:
            optional_str['queue-type'] = f"{q_up}/{q_dn}"
        elif q_up:
            optional_str['queue-type'] = q_up
        elif q_dn:
            optional_str['queue-type'] = f"default/{q_dn}"
        for k, v in optional_str.items():
            if v:
                payload[k] = v
        only_one = data.get('only_one', 'default')
        if only_one and only_one != 'default':
            payload['only-one'] = only_one

        api.get_resource('/ppp/profile').add(**payload)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()


@app.route('/api/pppoe/profile/update', methods=['POST'])
def update_pppoe_profile():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role not in ('admin', 'technician'): return jsonify({"error": "Admin/Tech only"}), 403

    data = request.json
    router_id = data.get('router_id')
    old_name = data.get('old_name', '').strip()
    new_name = data.get('name', '').strip()
    
    if not old_name:
        return jsonify({"status": "error", "msg": "Target profile name required"}), 400

    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500

    try:
        api = conn.get_api()
        res = api.get_resource('/ppp/profile')
        
        # Find the .id by name
        existing = res.get(name=old_name)
        if not existing:
            # Fallback for some Mikrotik/API versions where filtered get fails
            all_profs = res.get()
            match = [p for p in all_profs if p.get('name') == old_name]
            if match:
                existing = match
            else:
                return jsonify({"status": "error", "msg": f"Profile '{old_name}' not found"}), 404
            
        item = existing[0]
        target_id = item.get('.id') or item.get('id') or item.get('*ID')
        
        # If still no ID, maybe it's a 'default' profile that uses its name as ID in some ROS versions
        # or it's a special case. But normally we need an ID for .set()
        if not target_id:
            if old_name.lower() in ['default', 'default-encryption']:
                # For default profiles, sometimes the name itself works as ID in some API implementations
                target_id = old_name
            else:
                available_keys = list(item.keys())
                return jsonify({"status": "error", "msg": f"Could not determine ID for '{old_name}'. Keys found: {available_keys}"}), 500
        
        payload = {}
        if new_name and new_name != old_name:
            payload['name'] = new_name
            
        q_up = str(data.get('queue_type') or data.get('queue_type_upload') or '').strip()
        q_dn = str(data.get('queue_type_download') or '').strip()
            
        mapping = {
            'local-address': data.get('local_address'),
            'remote-address': data.get('remote_address'),
            'dns-server': data.get('dns_server'),
            'rate-limit': data.get('rate_limit'),
            'only-one': data.get('only_one'),
            'insert-queue-before': data.get('insert_queue_before'),
            'parent-queue': data.get('parent_queue'),
        }
        
        if q_up and q_dn: mapping['queue-type'] = f"{q_up}/{q_dn}"
        elif q_up: mapping['queue-type'] = q_up
        elif q_dn: mapping['queue-type'] = f"default/{q_dn}"
        
        for mk_key, val in mapping.items():
            if val is not None:
                val = str(val).strip()
                # Skip empty values to avoid 'ambiguous value of pool' error in Mikrotik
                if not val:
                    continue
                if mk_key == 'only-one' and val == 'default':
                    continue
                payload[mk_key] = val

        try:
            res.set(id=target_id, **payload)
        except Exception as set_err:
            # If set by ID fails, try set by name as fallback for some ROS versions/profiles
            try:
                # Some API implementations allow set by name if it's a unique key
                res.set(name=old_name, **payload)
            except:
                raise set_err # Re-raise original error if fallback also fails

        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()


@app.route('/api/pppoe/profile/delete', methods=['POST'])
def delete_pppoe_profile():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role not in ('admin', 'technician'): return jsonify({"error": "Admin/Tech only"}), 403

    data = request.json
    router_id = data.get('router_id')
    name = data.get('name', '').strip()
    if not name:
        return jsonify({"status": "error", "msg": "Profile name required"}), 400

    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500

    try:
        api = conn.get_api()
        res = api.get_resource('/ppp/profile')
        items = res.get(name=name)
        if not items:
            return jsonify({"status": "error", "msg": f"Profile '{name}' not found"}), 404
        mid = items[0].get('.id') or items[0].get('id')
        res.remove(id=mid)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()


@app.route('/api/pppoe/users/<router_id>', methods=['GET'])
def get_pppoe_users(router_id):
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    comment_filter = request.args.get('comment')
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify([]), 500
    
    try:
        api = conn.get_api()
        res_user = api.get_resource('/ppp/secret')
        
        if comment_filter:
            res = res_user.get(comment=comment_filter)
            if not res:
                all_u = res_user.get()
                res = [u for u in all_u if comment_filter in u.get('comment', '')]
        else:
            res = res_user.get()
            
        # Limit safety
        if len(res) > 5000: res = res[:5000]
            
        users = []
        for r in res:
             mid = r.get('.id') or r.get('id')
             users.append({
                "id": mid,
                "name": r.get('name'),
                "password": r.get('password'),
                "profile": r.get('profile'),
                "service": r.get('service', 'any'),
                "comment": r.get('comment', ''),
                "local_address": r.get('local-address', ''),
                "remote_address": r.get('remote-address', ''),
                "last_logged_out": r.get('last-logged-out', ''),
                "last_disconnect_reason": r.get('last-disconnect-reason', ''),
                "disabled": r.get('disabled') == 'true'
            })
        # Fetch router identity for display in frontend dropdown
        router_identity = ''
        try:
            identity_res = api.get_resource('/system/identity').get()
            router_identity = identity_res[0].get('name', '') if identity_res else ''
        except Exception: pass

        return jsonify({
            "status": "ok",
            "users": users,
            "router_identity": router_identity
        })
    except Exception as e:
        print(f"[PPPoE-USERS] Error: {e}")
        return jsonify({"status": "error", "msg": str(e), "users": []})
    finally:
        conn.disconnect()

@app.route('/api/pppoe/create', methods=['POST'])
def create_pppoe_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    name = data.get('name')
    password = data.get('password')
    profile = data.get('profile', 'default')
    service = data.get('service', 'pppoe')
    comment = data.get('comment', 'created-via-nms')
    local_address = data.get('local_address', '').strip()
    remote_address = data.get('remote_address', '').strip()
    
    if not name or not password:
        return jsonify({"status": "error", "msg": "Name and Password required"}), 400
        
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res_secret = api.get_resource('/ppp/secret')
        
        payload = {
            'name': name,
            'password': password,
            'profile': profile,
            'service': service,
            'comment': comment
        }
        if local_address:
            payload['local-address'] = local_address
        if remote_address:
            payload['remote-address'] = remote_address
        if not profile: profile = 'default'
        if not service: service = 'pppoe'
        payload['profile'] = profile
        payload['service'] = service
        
        res_secret.add(**payload)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/pppoe/update', methods=['POST'])
def update_pppoe_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    name = data.get('name')
    password = data.get('password')
    profile = data.get('profile')
    service = data.get('service')
    comment = data.get('comment')
    local_address = data.get('local_address', '').strip()
    remote_address = data.get('remote_address', '').strip()
    disabled = data.get('disabled') # Accept boolean or None
    
    if not name:
        return jsonify({"status": "error", "msg": "Name required"}), 400
        
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res_secret = api.get_resource('/ppp/secret')
        
        # Find the secret ID
        items = res_secret.get(name=name)
        if not items:
            return jsonify({"status": "error", "msg": "User not found"}), 404
        
        item = items[0]
        secret_id = item.get('.id') or item.get('id') or item.get('*ID')
        
        if not secret_id:
            available_keys = list(item.keys())
            return jsonify({"status": "error", "msg": f"Could not determine ID for user '{name}'. Keys found: {available_keys}"}), 500
        
        payload = {}
        if password is not None: payload['password'] = password
        if profile: payload['profile'] = profile
        if service: payload['service'] = service
        if comment is not None: payload['comment'] = comment
        if local_address: payload['local-address'] = local_address
        if remote_address: payload['remote-address'] = remote_address
        if disabled is not None: payload['disabled'] = 'yes' if disabled else 'no'
        
        res_secret.set(id=secret_id, **payload)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/pppoe/toggle', methods=['POST'])
def toggle_pppoe_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    name = data.get('name')
    
    if not name or not router_id:
        return jsonify({"status": "error", "msg": "Name and Router ID required"}), 400
        
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res_secret = api.get_resource('/ppp/secret')
        items = res_secret.get(name=name)
        if not items:
            # Fallback check if name was passed as ID
            items = res_secret.get(**{'.id': name})
            if not items: items = res_secret.get(id=name)
            
        if not items:
            return jsonify({"status": "error", "msg": "User not found"}), 404
        
        item = items[0]
        secret_id = item.get('.id') or item.get('id') or item.get('*ID')
        if not secret_id:
            return jsonify({"status": "error", "msg": f"Could not determine ID for user '{name}'"}), 500
            
        current_disabled = item.get('disabled') == 'true'
        new_disabled = not current_disabled
        
        res_secret.set(id=secret_id, disabled='yes' if new_disabled else 'no')
        
        # If disabling, also kick the current session
        if new_disabled:
            try: kick_pppoe_user(name, router_id)
            except: pass
            
        return jsonify({"status": "ok", "disabled": new_disabled})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()


@app.route('/api/pppoe/delete', methods=['POST'])
def delete_pppoe_users():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    ids = data.get('ids', [])
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    success = 0
    last_error = ""
    try:
        api = conn.get_api()
        res_secret = api.get_resource('/ppp/secret')
        
        for mid in ids:
            try:
                # Robust ID lookup before delete
                target_user = None
                try: target_user = res_secret.get(**{'.id': mid})
                except: pass
                
                if not target_user:
                    try: target_user = res_secret.get(id=mid)
                    except: pass
                
                if target_user:
                    actual_id = target_user[0].get('.id') or target_user[0].get('id') or target_user[0].get('*ID')
                    res_secret.remove(id=actual_id)
                    success += 1
                else:
                    # Final attempt by name
                    try:
                        res_secret.remove(name=mid)
                        success += 1
                    except:
                        last_error = f"ID/Name {mid} not found on router"
            except Exception as e:
                last_error = str(e)
            
        return jsonify({"status": "ok", "deleted": success, "last_error": last_error})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/pppoe/active/<router_id>', methods=['GET'])
def get_active_pppoe(router_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify([]), 500
    
    try:
        api = conn.get_api()
        # Equivalent to /ppp/active/print
        res = api.get_resource('/ppp/active').get()
        
        active = []
        for r in res:
            active.append({
                "id": r.get('.id') or r.get('id'),
                "name": r.get('name', 'unknown'),
                "service": r.get('service', 'pppoe'),
                "address": r.get('address', '-'),
                "uptime": r.get('uptime', '00:00:00'),
                "caller_id": r.get('caller-id', '-')
            })
            
        # Fetch router identity for display in frontend dropdown
        router_identity = ''
        try:
            identity_res = api.get_resource('/system/identity').get()
            router_identity = identity_res[0].get('name', '') if identity_res else ''
        except Exception: pass

        return jsonify({
            "status": "ok",
            "active": active,
            "router_identity": router_identity
        })
    except Exception as e:
        print(f"[PPPoE-ACTIVE] Error: {e}")
        return jsonify({"status": "error", "msg": str(e), "active": []})
    finally:
        conn.disconnect()

@app.route('/api/pppoe/kick', methods=['POST'])
def kick_pppoe_active_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json
    router_id = data.get('router_id')
    mid = data.get('id')
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res_active = api.get_resource('/ppp/active')
        
        target = None
        try: target = res_active.get(**{'.id': mid})
        except: pass
        if not target:
            try: target = res_active.get(id=mid)
            except: pass
            
        if target:
            actual_id = target[0].get('.id') or target[0].get('id') or target[0].get('*ID')
            res_active.remove(id=actual_id)
            return jsonify({"status": "ok"})
        else:
             return jsonify({"status": "error", "msg": "Active session not found"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)})
    finally:
        conn.disconnect()


@app.route('/api/hotspot/user/create', methods=['POST'])
def create_hotspot_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    name = data.get('name')
    password = data.get('password', '')
    profile = data.get('profile', 'default')
    server = data.get('server', 'all')
    comment = data.get('comment', '')
    limit_uptime = data.get('limit_uptime')
    limit_bytes = data.get('limit_bytes')
    
    if not name: return jsonify({"error": "Name required"}), 400
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res = api.get_resource('/ip/hotspot/user')
        payload = {
            "name": name,
            "password": password,
            "profile": profile,
            "server": server,
            "comment": comment
        }
        if limit_uptime: payload['limit-uptime'] = limit_uptime
        pbytes = parse_hotspot_limit_bytes(limit_bytes)
        if pbytes is not None: payload['limit-bytes-total'] = pbytes
        
        res.add(**payload)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/user/update', methods=['POST'])
def update_hotspot_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    user_id = data.get('id')
    name = data.get('name')
    password = data.get('password')
    profile = data.get('profile')
    comment = data.get('comment')
    limit_uptime = data.get('limit_uptime')
    limit_bytes = data.get('limit_bytes')
    
    if not user_id: return jsonify({"error": "ID required"}), 400
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res = api.get_resource('/ip/hotspot/user')
        
        payload = {}
        if name: payload['name'] = name
        if password is not None: payload['password'] = password
        if profile: payload['profile'] = profile
        if data.get('server'): payload['server'] = data.get('server')
        if comment is not None: payload['comment'] = comment
        if limit_uptime is not None: payload['limit-uptime'] = limit_uptime if limit_uptime else "0s"
        pbytes = parse_hotspot_limit_bytes(limit_bytes)
        if pbytes is not None: payload['limit-bytes-total'] = pbytes
        
        res.set(id=user_id, **payload)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/user/toggle', methods=['POST'])
def toggle_hotspot_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json
    router_id = data.get('router_id')
    user_id = data.get('id')
    disabled = data.get('disabled')
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res = api.get_resource('/ip/hotspot/user')
        res.set(id=user_id, disabled='yes' if disabled else 'no')
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/profile/update', methods=['POST'])
def update_hotspot_profile():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    profile_id = data.get('id')
    rate_limit = data.get('rate_limit')
    shared_users = data.get('shared_users')
    session_timeout = data.get('session_timeout')
    idle_timeout = data.get('idle_timeout')
    keepalive_timeout = data.get('keepalive_timeout')
    status_autorefresh = data.get('status_autorefresh')
    address_pool = data.get('address_pool')
    add_mac_cookie = data.get('add_mac_cookie')
    mac_cookie_timeout = data.get('mac_cookie_timeout')
    address_list = data.get('address_list')
    
    # Queue fields
    parent_queue = data.get('parent_queue')
    queue_type = data.get('queue_type')
    insert_queue_before = data.get('insert_queue_before')
    
    if not profile_id: return jsonify({"error": "ID required"}), 400
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res = api.get_resource('/ip/hotspot/user/profile')
        payload = {}
        if rate_limit is not None: payload['rate-limit'] = rate_limit
        if shared_users is not None: payload['shared-users'] = str(shared_users)
        if session_timeout is not None: payload['session-timeout'] = session_timeout
        if idle_timeout is not None: payload['idle-timeout'] = idle_timeout
        if keepalive_timeout is not None: payload['keepalive-timeout'] = keepalive_timeout
        if status_autorefresh is not None: payload['status-autorefresh'] = status_autorefresh
        if address_pool is not None: payload['address-pool'] = address_pool
        if add_mac_cookie is not None: payload['add-mac-cookie'] = 'yes' if add_mac_cookie else 'no'
        if mac_cookie_timeout is not None: payload['mac-cookie-timeout'] = mac_cookie_timeout
        if address_list is not None: payload['address-list'] = address_list

        # Queue fields
        if parent_queue: payload['parent-queue'] = parent_queue
        if queue_type: payload['queue-type'] = queue_type
        if insert_queue_before: payload['insert-queue-before'] = insert_queue_before
        
        res.set(id=profile_id, **payload)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.disconnect()

if __name__ == '__main__':
    # Flask Native Start (Threads already started above in Global)
    port = int(cfg.get('app_port', 5002))
    app.run(host='0.0.0.0', port=port, threaded=True)
